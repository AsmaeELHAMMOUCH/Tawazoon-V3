from fastapi import APIRouter, Depends
from sqlalchemy.orm import Session
from app.core.db import get_db
from app.services.capacite_nominale import get_capacite_nominale_data

router = APIRouter(prefix="/capacite-nominale", tags=["capacite-nominale"])

@router.get("/")
def get_capacite(db: Session = Depends(get_db)):
    """
    Retourne la liste des postes avec leurs capacités calculées.
    Format strictement conforme à l'UI PyQt5.
    """
    data = get_capacite_nominale_data(db)
    return data

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell
from io import BytesIO
from fastapi.responses import StreamingResponse
from pathlib import Path
from typing import Optional

from app.core.config import settings

@router.get("/export-xlsx")
def export_capacite(db: Session = Depends(get_db)):
    """
    Export Excel strictement conforme à l'UI (6 colonnes).
    ["Position","Temps/Dossier","Effectif actuel","Dossiers/Mois","Dossiers/Jour","Dossiers/Heure"]
    Ligne Total incluse.
    """
    data_dict = get_capacite_nominale_data(db)
    rows = data_dict["rows"]
    total = data_dict["total"]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.sheet_view.showGridLines = False
    ws.title = "Capacite Nominale"
    ws.row_dimensions[1].height = 60
    ws.row_dimensions[2].height = 20

    def _resolve_logo(name: str) -> Optional[str]:
        candidate = Path(name)
        if not candidate.is_absolute():
            candidate = settings.assets_dir / candidate
        candidate = candidate.resolve(strict=False)
        return str(candidate) if candidate.exists() else None

    def _add_logo(path_str: Optional[str], anchor: str, max_width: int, max_height: int):
        if not path_str:
            return
        try:
            img = Image(path_str)
            scale = min(max_width / img.width, max_height / img.height, 1)
            img.width = int(img.width * scale)
            img.height = int(img.height * scale)
            ws.add_image(img, anchor)
        except Exception:
            pass

    _add_logo(_resolve_logo(settings.LOGO_BARID), "A1", max_width=150, max_height=60)
    _add_logo(_resolve_logo(settings.LOGO_ALMAV), "F1", max_width=150, max_height=60)

    ws.merge_cells("B1:E1")
    title_cell = ws["B1"]
    title_cell.value = "Capacité nominale"
    title_cell.font = Font(bold=True, size=16, color="005EA8")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws["B2"] = ""
    ws.merge_cells("B2:E2")
    ws["B2"].alignment = Alignment(horizontal="center", vertical="center")

    # 1. En-têtes
    headers = [
        "Position",
        "Temps Actuel Par Dossier",
        "Effectif actuel",
        "Dossiers/Mois",
        "Dossiers/Jour",
        "Dossiers/Heure"
    ]
    table_start = 4
    blue_fill = PatternFill("solid", fgColor="005EA8")
    white_font = Font(bold=True, color="FFFFFF")
    center_align = Alignment(horizontal="center", vertical="center")

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=table_start, column=col_idx, value=header)
        cell.font = white_font
        cell.fill = blue_fill
        cell.alignment = center_align

    # 3. Données
    current_row = table_start + 1
    for r in rows:
        ws.cell(row=current_row, column=1, value=r["position"])
        ws.cell(row=current_row, column=2, value=r["temps_par_dossier"])
        ws.cell(row=current_row, column=3, value=r["effectif_actuel"])
        ws.cell(row=current_row, column=4, value=r["dossiers_mois"])
        ws.cell(row=current_row, column=5, value=r["dossiers_jour"])
        ws.cell(row=current_row, column=6, value=r["dossiers_heure"])
        current_row += 1

    # 4. Ligne Total
    current_row += 1
    total_values = [
        total["position"],
        total["temps_par_dossier"],
        total["effectif_actuel"],
        total["dossiers_mois"],
        total["dossiers_jour"],
        total["dossiers_heure"]
    ]
    gray_fill = PatternFill("solid", fgColor="F3F4F6")
    bold_font = Font(bold=True)
    for col_idx, value in enumerate(total_values, start=1):
        cell = ws.cell(row=current_row, column=col_idx, value=value)
        cell.font = bold_font
        cell.fill = gray_fill
        cell.alignment = center_align

    for column_cells in ws.columns:
        column_letter = None
        max_length = 0
        for cell in column_cells:
            if isinstance(cell, MergedCell):
                continue
            if column_letter is None:
                column_letter = cell.column_letter
            value = cell.value
            if value is None:
                continue
            max_length = max(max_length, len(str(value)))
        if column_letter:
            ws.column_dimensions[column_letter].width = max(12, max_length + 2)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    
    headers_resp = {
        "Content-Disposition": 'attachment; filename="capacite_nominale.xlsx"',
        "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    }
    return StreamingResponse(buf, headers=headers_resp)
