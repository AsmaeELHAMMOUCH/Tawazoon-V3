from fastapi import APIRouter, Depends, HTTPException
from sqlalchemy.orm import Session
from pydantic import BaseModel
from typing import List, Dict, Any, Optional
from app.core.db import get_db
from app.services.simulation_globale_v3 import get_simulation_globale_v3
from app.core.config import settings
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image
from io import BytesIO
from fastapi.responses import StreamingResponse
from pathlib import Path

router = APIRouter(prefix="/simulation/globale", tags=["Simulation Globale"])

class SimulationGlobaleRequest(BaseModel):
    sacs: int
    dossiers_mois: int
    productivite: float

class SimulationRow(BaseModel):
    position: str
    actuel: float
    calcule: float
    recommande: float
    ecart_calc_actuel: float
    ecart_reco_actuel: float
    ecart_reco_calcule: float

class SimulationTotals(BaseModel):
    position: str
    actuel: float
    calcule: float
    recommande: float
    ecart_calc_actuel: float
    ecart_reco_actuel: float
    ecart_reco_calcule: float

class SimulationGlobaleResponse(BaseModel):
    dossiers_jour: float
    heures_net_jour: float
    rows: List[SimulationRow]
    total: SimulationTotals
    chart_data: List[Dict[str, Any]]

@router.post("/v3/run", response_model=SimulationGlobaleResponse)
def run_simulation_globale_v3(req: SimulationGlobaleRequest, db: Session = Depends(get_db)):
    try:
        data = get_simulation_globale_v3(
            db, 
            sacs=req.sacs, 
            dossiers_mois=req.dossiers_mois, 
            productivite=req.productivite
        )
        return data
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/v3/export/excel")
def export_simulation_globale_v3_excel(
    sacs: int, 
    dossiers_mois: int, 
    productivite: float, 
    db: Session = Depends(get_db)
):
    try:
        # 1. Calculer les données
        data = get_simulation_globale_v3(db, sacs=sacs, dossiers_mois=dossiers_mois, productivite=productivite)
        
        # 2. Créer le Workbook
        wb = Workbook()
        ws = wb.active
        ws.sheet_view.showGridLines = False
        ws.title = "Simulation Globale V3"
        
        # Styles
        header_fill = PatternFill(start_color="005EA8", end_color="005EA8", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        center_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        bold_font = Font(bold=True)
        
        # Logos & Titre
        ws.row_dimensions[1].height = 60
        ws.merge_cells("C1:G1")
        ws["C1"] = "RÉSULTATS DE LA SIMULATION GLOBALE"
        ws["C1"].font = Font(size=16, bold=True, color="005EA8")
        ws["C1"].alignment = center_align
        
        def _resolve_logo_path(name: Optional[str]) -> Optional[str]:
            if not name: return None
            candidate = Path(name)
            if not candidate.is_absolute():
                candidate = settings.assets_dir / candidate
            candidate = candidate.resolve(strict=False)
            return str(candidate) if candidate.exists() else None

        def _place_logo(cell_ref, path_str, max_w, max_h):
            if not path_str: return
            try:
                img = Image(path_str)
                scale = min(max_w / img.width, max_h / img.height, 1)
                img.width = int(img.width * scale)
                img.height = int(img.height * scale)
                ws.add_image(img, cell_ref)
            except: pass

        _place_logo("A1", _resolve_logo_path(settings.LOGO_BARID), 150, 60)
        _place_logo("H1", _resolve_logo_path(settings.LOGO_ALMAV), 120, 50)
        
        # Paramètres de simulation
        ws["A3"] = "PARAMÈTRES UTILISÉS"
        ws["A3"].font = bold_font
        ws.merge_cells("A3:B3")
        
        params_row = 4
        ws.cell(row=params_row, column=1, value="Sacs / Jour :").font = bold_font
        ws.cell(row=params_row, column=2, value=sacs)
        ws.cell(row=params_row, column=4, value="Dossiers / Mois :").font = bold_font
        ws.cell(row=params_row, column=5, value=dossiers_mois)
        ws.cell(row=params_row, column=7, value="Taux Productivité :").font = bold_font
        ws.cell(row=params_row, column=8, value=f"{productivite}%")
        
        stats_row = 5
        ws.cell(row=stats_row, column=1, value="Dossiers / Jour (Calc) :").font = bold_font
        ws.cell(row=stats_row, column=2, value=data["dossiers_jour"])
        ws.cell(row=stats_row, column=4, value="Heures Travaillées / Jour :").font = bold_font
        ws.cell(row=stats_row, column=5, value=data["heures_net_jour"])
        ws.cell(row=stats_row, column=7, value="Temps Mort :").font = bold_font
        ws.cell(row=stats_row, column=8, value=0)
        
        # Tableau des résultats
        headers = [
            "Position", "Actuel", "Calculé (M)", "Recommandé (R)", 
            "Écart (M-A)", "Écart (R-A)", "Écart (R-M)"
        ]
        start_row = 7
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border
            
        cur_row = start_row + 1
        for r in data["rows"]:
            ws.cell(row=cur_row, column=1, value=r["position"]).border = thin_border
            ws.cell(row=cur_row, column=2, value=r["actuel"]).border = thin_border
            ws.cell(row=cur_row, column=3, value=r["calcule"]).border = thin_border
            ws.cell(row=cur_row, column=4, value=r["recommande"]).border = thin_border
            ws.cell(row=cur_row, column=5, value=r["ecart_calc_actuel"]).border = thin_border
            ws.cell(row=cur_row, column=6, value=r["ecart_reco_actuel"]).border = thin_border
            ws.cell(row=cur_row, column=7, value=r["ecart_reco_calcule"]).border = thin_border
            cur_row += 1
            
        # Ligne TOTAL
        t = data["total"]
        ws.cell(row=cur_row, column=1, value="TOTAL").font = bold_font
        ws.cell(row=cur_row, column=2, value=t["actuel"]).font = bold_font
        ws.cell(row=cur_row, column=3, value=t["calcule"]).font = bold_font
        ws.cell(row=cur_row, column=4, value=t["recommande"]).font = bold_font
        ws.cell(row=cur_row, column=5, value=t["ecart_calc_actuel"]).font = bold_font
        ws.cell(row=cur_row, column=6, value=t["ecart_reco_actuel"]).font = bold_font
        ws.cell(row=cur_row, column=7, value=t["ecart_reco_calcule"]).font = bold_font
        
        for col in range(1, 8):
            ws.cell(row=cur_row, column=col).border = thin_border
            ws.cell(row=cur_row, column=col).fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

        # Ajustement colonnes
        for col in range(1, 9):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
        ws.column_dimensions['A'].width = 30
        
        # Export
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=simulation_globale_v3.xlsx"}
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
