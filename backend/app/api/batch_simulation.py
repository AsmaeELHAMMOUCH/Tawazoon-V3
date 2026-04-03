# backend/app/api/batch_simulation.py
"""
Simulation Batch — Régionale & Nationale
Endpoints:
  GET  /batch/template/regional?region_id={id}   → Excel template pour une région
  GET  /batch/template/national                   → Excel template pour tout le réseau
  POST /batch/simulate                            → Lance la simulation sur le fichier importé
"""

from typing import Optional, List
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import text
import asyncio
import io
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side, Color, Protection
from openpyxl.utils import get_column_letter

from app.core.db import get_db
from app.services.bandoeng_engine import (
    run_bandoeng_simulation,
    BandoengInputVolumes,
    BandoengParameters,
)
from app.services.taches_service import auto_import_tasks_if_empty
from app.models.db_models import CentrePoste, Poste, Centre
try:
    from app.models.db_models import MappingPosteRecommande, TacheExclueOptimisee
except ImportError:
    MappingPosteRecommande = None
    TacheExclueOptimisee = None


router = APIRouter(prefix="/batch", tags=["Batch Simulation"])

# ─────────────────────────────────────────────────────────────────────────────
# Styles helpers
# ─────────────────────────────────────────────────────────────────────────────
def _styles():
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    bold = Font(bold=True)
    title_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill_blue = PatternFill("solid", fgColor="005EA8")
    header_fill_gray = PatternFill("solid", fgColor="E2E8F0")
    depart_fill = PatternFill("solid", fgColor="EFF6FF")
    arrive_fill = PatternFill("solid", fgColor="F0FDF4")
    param_fill = PatternFill("solid", fgColor="FFF7ED")
    input_fill = PatternFill("solid", fgColor="FFFBEB")
    locked_fill = PatternFill("solid", fgColor="F1F5F9")
    thin = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1"),
    )
    return dict(
        center=center, bold=bold, title_font=title_font,
        header_fill_blue=header_fill_blue, header_fill_gray=header_fill_gray,
        depart_fill=depart_fill, arrive_fill=arrive_fill,
        param_fill=param_fill, input_fill=input_fill,
        locked_fill=locked_fill, thin=thin
    )


# ─────────────────────────────────────────────────────────────────────────────
# Canvas structure — identique à BandoengGrid.jsx
# ─────────────────────────────────────────────────────────────────────────────
#
#  DÉPART block:
#    PRO:           Global | Local | Axes         (Amana, CR, CO only)
#    Particuliers:  Global | Local | Axes         (Amana only)
#    Single col:    Med                           (Ebarkia, LRH)
#
#  ARRIVÉ block (même structure):
#    PRO:           Global | Local | Axes
#    Particuliers:  Global | Local | Axes         (Amana only)
#    Single col:                                  (Ebarkia, LRH)
#
# Excel key format used for parsing:   flux_sens_segment_subsegment
# e.g.: amana_depot_gc_global, cr_med_global, ebarkia_med
#

FLUX_ROWS = ["Amana", "CR", "CO", "El Barkia", "LRH"]

# Columns per block (13 total including label)
# Block structure:  [Label] [D-PRO-Global] [D-PRO-Local] [D-PRO-Axes] [D-Part-Global] [D-Part-Local] [D-Part-Axes] [A-PRO-Global] [A-PRO-Local] [A-PRO-Axes] [A-Part-Global] [A-Part-Local] [A-Part-Axes]
#                      A         B              C             D              E              F              G              H              I              J              K              L              M

CANVAS_HEADERS = {
    "label_col": 1,  # A
    "depart_start": 2,  # B
    "depart_cols": [  # B-G
        ("PRO", "Global"),    # B
        ("PRO", "Local"),     # C
        ("PRO", "Axes"),      # D
        ("Part.", "Global"),  # E
        ("Part.", "Local"),   # F
        ("Part.", "Axes"),    # G
    ],
    "arrive_start": 8,  # H
    "arrive_cols": [  # H-M
        ("PRO", "Global"),    # H
        ("PRO", "Local"),     # I
        ("PRO", "Axes"),      # J
        ("Part.", "Global"),  # K
        ("Part.", "Local"),   # L
        ("Part.", "Axes"),    # M
    ],
}

# For each flux row, which cells are active (True) vs locked (False)
# [D-PRO-G, D-PRO-L, D-PRO-A, D-Part-G, D-Part-L, D-Part-A, A-PRO-G, A-PRO-L, A-PRO-A, A-Part-G, A-Part-L, A-Part-A]
FLUX_ACTIVE_CELLS = {
    "Amana":     [True,  True,  True,  True,  True,  True,  True,  True,  True,  True,  True,  True],
    "CR":        [True,  True,  True,  False, False, False, True,  True,  True,  False, False, False],
    "CO":        [True,  True,  True,  False, False, False, True,  True,  True,  False, False, False],
    "El Barkia": [True,  False, False, False, False, False, True,  False, False, False, False, False],
    "LRH":       [True,  False, False, False, False, False, True,  False, False, False, False, False],
}

# Excel cell keys for parsing (same order as FLUX_ACTIVE_CELLS)
FLUX_CELL_KEYS = {
    "Amana":     [
        "amana_depot_gc_global", "amana_depot_gc_local", "amana_depot_gc_axes",
        "amana_depot_part_global", "amana_depot_part_local", "amana_depot_part_axes",
        "amana_recu_gc_global",  "amana_recu_gc_local",  "amana_recu_gc_axes",
        "amana_recu_part_global","amana_recu_part_local","amana_recu_part_axes",
    ],
    "CR":        [
        "cr_med_global", "cr_med_local", "cr_med_axes",
        None, None, None,
        "cr_arrive_global", "cr_arrive_local", "cr_arrive_axes",
        None, None, None,
    ],
    "CO":        [
        "co_med_global", "co_med_local", "co_med_axes",
        None, None, None,
        "co_arrive_global", "co_arrive_local", "co_arrive_axes",
        None, None, None,
    ],
    "El Barkia": [
        "ebarkia_med", None, None, None, None, None,
        "ebarkia_arrive", None, None, None, None, None,
    ],
    "LRH":       [
        "lrh_med", None, None, None, None, None,
        "lrh_arrive", None, None, None, None, None,
    ],
}

GLOBAL_PARAMS = [
    ("productivite",        "Productivité",              100.0,   "%"),
    ("idle_minutes",        "Temps Mort",                  0.0,  "min"),
    ("shift",               "Nombre de Shifts",           1,     ""),
    ("coeff_circ",          "Coeff. Circulaire",           1.0,   ""),
    ("coeff_geo",           "Coeff. Géographique",         1.0,   ""),
    ("duree_trajet",        "Durée Trajet",                0.0,  "min"),
]

AMANA_PARAMS = [
    ("amana_pct_collecte",      "% Collecte Amana",        0.0,   "%"),
    ("amana_pct_marche_ordinaire", "% Marché Ordinaire Amana", 0.0, "%"),
    ("amana_pct_guichet",       "% Guichet Amana",         0.0,   "%"),
    ("amana_pct_axes_arrivee",  "Axes Amana",              0.0,   "%"),
    ("amana_pct_axes_depart",   "Local Amana",             0.0,   "%"),
    ("amana_pct_national",      "% National Amana",        0.0,   "%"),
    ("amana_pct_international", "% International Amana",   0.0,   "%"),
    ("amana_pct_crbt",          "% CRBT Amana",            0.0,   "%"),
    ("amana_pct_hors_crbt",     "% Hors CRBT Amana",       0.0,   "%"),
    # Paramètres non affichés initialement mais nécessaires
    ("amana_pct_retour",        "% Retour Amana",          0.0,   "%"),
    ("ed_percent",              "% ED (Global)",           40.0,   "%"),
    ("colis_amana_par_canva_sac", "Colis Amana/Sac",      35.0,   ""),
]

CO_PARAMS = [
    ("co_pct_collecte",         "% Collecte CO",           0.0,   "%"),
    ("co_pct_marche_ordinaire", "% Marché Ordinaire CO",   0.0,   "%"),
    ("co_pct_guichet",          "% Guichet CO",            0.0,   "%"),
    ("co_pct_axes_arrivee",     "Axes CO",                 0.0,   "%"),
    ("co_pct_axes_depart",      "Local CO",                0.0,   "%"),
    ("co_pct_vague_master",     "% Vaguemestre CO",       0.0,   "%"),
    ("co_pct_boite_postale",    "% Boîte Postale CO",      0.0,   "%"),
    # Paramètres non affichés initialement
    ("co_pct_retour",           "% Retour CO",             0.0,   "%"),
    ("nbr_co_sac",              "Nbr CO/Sac",            350.0,   ""),
]

CR_PARAMS = [
    ("cr_pct_collecte",         "% Collecte CR",           0.0,   "%"),
    ("cr_pct_marche_ordinaire", "% Marché Ordinaire CR",   0.0,   "%"),
    ("cr_pct_guichet",          "% Guichet CR",            0.0,   "%"),
    ("cr_pct_axes_arrivee",     "Axes CR",                 0.0,   "%"),
    ("cr_pct_axes_depart",      "Local CR",                0.0,   "%"),
    ("cr_pct_national",         "% National CR",           0.0,   "%"),
    ("cr_pct_international",    "% International CR",      0.0,   "%"),
    ("cr_pct_crbt",             "% Accusé de Réc.",               0.0,   "%"),
    ("cr_pct_hors_crbt",        "% Hors Accusé de Réc.",          0.0,   "%"),
    ("cr_pct_vague_master",     "% Vaguemestre CR",       0.0,   "%"),
    # Paramètres non affichés
    ("cr_pct_retour",           "% Retour CR",             0.0,   "%"),
    ("cr_par_caisson",          "CR par Caisson",        500.0,   ""),
]

# Liste technique pour itérer le parser si besoin de compatibilité
PARAMS_FLUX_MAP = {
    "GLOBALE": (GLOBAL_PARAMS, "F97316"), # Orange
    "AMANA":   (AMANA_PARAMS,  "0EA5E9"), # Blue
    "COURRIER ORDINAIRE": (CO_PARAMS, "64748B"), # Slate
    "COURRIER RECOMMANDÉ": (CR_PARAMS, "10B981"), # Emerald
}

# ─────────────────────────────────────────────────────────────────────────────
# Core: generate template workbook
# ─────────────────────────────────────────────────────────────────────────────
def _build_template_workbook(centres: list) -> openpyxl.Workbook:
    s = _styles()
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ── Guide sheet ─────────────────────────────────────────────────────────
    ws_g = wb.create_sheet("Guide", 0)
    ws_g["A1"] = "GUIDE — SIMULATION RÉGIONALE / NATIONALE"
    ws_g["A1"].font = Font(bold=True, size=14, color="005EA8")
    ws_g["A3"] = "⚠️  NE PAS RENOMMER NI SUPPRIMER LES ONGLETS"
    ws_g["A3"].font = Font(bold=True, color="CC0000")
    ws_g["A4"] = "Le nom de chaque onglet identifie le centre. Toute modification provoquera une erreur d'import."
    ws_g["A6"] = "REMPLISSAGE"
    ws_g["A6"].font = Font(bold=True)
    ws_g["A7"] = "• Cellules jaunes = volumes à saisir (volumes annuels)"
    ws_g["A8"] = "• Cellules grises = non applicables pour ce flux (laisser vides)"
    ws_g["A9"] = "• Section PARAMÈTRES : modifier les valeurs de la colonne B uniquement"
    ws_g["A11"] = "STRUCTURE CANVAS (identique à l'interface Simulation Wizard)"
    ws_g["A11"].font = Font(bold=True)
    ws_g["A12"] = "Bloc DÉPART   → colonnes B à G  (PRO: Global/Local/Axes | Part.: Global/Local/Axes)"
    ws_g["A13"] = "Bloc ARRIVÉ   → colonnes H à M  (même structure)"
    ws_g.column_dimensions["A"].width = 90

    used_names = set()

    for c in centres:
        c_id = c["id"]
        c_label = c["label"]
        region_label = c.get("region_label", "")
        # Pré-remplissage "comme Wizard" depuis la BD (Ville liée au centre)
        # - coeff_geo  <- ville.geographie
        # - coeff_circ <- ville.circulation
        # - duree_trajet <- ville.trajet
        geo_db = c.get("coeff_geo_db", c.get("geographie", None))
        circ_db = c.get("coeff_circ_db", c.get("circulation", None))
        trajet_db = c.get("duree_trajet_db", c.get("trajet", None))

        safe = "".join(x for x in c_label if x not in r'[]:*?/\\')[:28]
        name = safe
        idx = 1
        while name in used_names:
            name = f"{safe[:26]}{idx}"
            idx += 1
        used_names.add(name)

        ws = wb.create_sheet(name)
        # Par défaut, une feuille protégée verrouille toutes les cellules.
        # On va ensuite déverrouiller explicitement les cellules de saisie (volumes Global + paramètres autorisés).
        ws.protection.sheet = True
        ws.protection.enable()

        # Row 1: centre info
        ws.merge_cells("A1:M1")
        ws["A1"] = f"CENTRE : {c_label}  |  RÉGION : {region_label}  |  ID : {c_id}"
        ws["A1"].font = s["title_font"]
        ws["A1"].fill = s["header_fill_blue"]
        ws["A1"].alignment = s["center"]

        # Row 2: block headers
        ws.merge_cells("B2:G2")
        ws["B2"] = "DÉPART"
        ws["B2"].font = Font(bold=True, color="FFFFFF")
        ws["B2"].fill = PatternFill("solid", fgColor="0EA5E9")
        ws["B2"].alignment = s["center"]

        ws.merge_cells("H2:M2")
        ws["H2"] = "ARRIVÉ"
        ws["H2"].font = Font(bold=True, color="FFFFFF")
        ws["H2"].fill = PatternFill("solid", fgColor="10B981")
        ws["H2"].alignment = s["center"]

        # Row 3: sub-block headers (PRO / Particuliers)
        for start_col, label in [(2, "PRO"), (5, "Particuliers"), (8, "PRO"), (11, "Particuliers")]:
            ws.merge_cells(
                start_row=3, start_column=start_col,
                end_row=3, end_column=start_col + 2
            )
            cell = ws.cell(row=3, column=start_col, value=label)
            cell.font = Font(bold=True)
            cell.fill = s["header_fill_gray"]
            cell.alignment = s["center"]
            cell.border = s["thin"]

        # Row 4: column headers
        ws.cell(row=4, column=1, value="Flux").font = Font(bold=True)
        ws.cell(row=4, column=1).fill = s["header_fill_gray"]
        ws.cell(row=4, column=1).alignment = s["center"]
        ws.cell(row=4, column=1).border = s["thin"]

        for i, (_, subcol) in enumerate(CANVAS_HEADERS["depart_cols"]):
            col = CANVAS_HEADERS["depart_start"] + i
            c_cell = ws.cell(row=4, column=col, value=subcol)
            c_cell.font = Font(bold=True)
            c_cell.fill = s["depart_fill"]
            c_cell.alignment = s["center"]
            c_cell.border = s["thin"]

        for i, (_, subcol) in enumerate(CANVAS_HEADERS["arrive_cols"]):
            col = CANVAS_HEADERS["arrive_start"] + i
            c_cell = ws.cell(row=4, column=col, value=subcol)
            c_cell.font = Font(bold=True)
            c_cell.fill = s["arrive_fill"]
            c_cell.alignment = s["center"]
            c_cell.border = s["thin"]

        # Rows 5-9: data rows per flux
        for r_idx, flux in enumerate(FLUX_ROWS):
            row = 5 + r_idx
            active = FLUX_ACTIVE_CELLS[flux]

            # Label cell
            lbl = ws.cell(row=row, column=1, value=flux)
            lbl.font = Font(bold=True)
            lbl.alignment = Alignment(vertical="center")
            lbl.border = s["thin"]
            lbl.fill = s["header_fill_gray"]

            # 12 data cells
            for col_offset, is_active in enumerate(active):
                col = 2 + col_offset
                # Only "Global" cells are manual inputs. "Local" and "Axes" are computed from %
                is_global = (col_offset % 3) == 0
                cell = ws.cell(row=row, column=col, value=0 if (is_active and is_global) else None)
                cell.border = s["thin"]
                if is_active:
                    if is_global:
                        cell.fill = s["input_fill"]
                        cell.alignment = Alignment(horizontal="center")
                        cell.number_format = "#,##0"
                        cell.protection = Protection(locked=False)
                    else:
                        # computed cells: locked style (formula will be set after params section)
                        cell.fill = s["locked_fill"]
                        cell.alignment = Alignment(horizontal="center")
                        cell.number_format = "#,##0"
                        cell.protection = Protection(locked=True)
                else:
                    cell.fill = s["locked_fill"]
                    cell.protection = Protection(locked=True)

        # Column widths
        ws.column_dimensions["A"].width = 12
        for col in range(2, 14):
            ws.column_dimensions[get_column_letter(col)].width = 11
        ws.row_dimensions[1].height = 22
        ws.row_dimensions[2].height = 18

        # ── PARAMS section ──────────────────────────────────────────────────
        curr_row = 12
        param_value_cell = {}
        param_override_values = {
            "coeff_geo": geo_db,
            "coeff_circ": circ_db,
            "duree_trajet": trajet_db,
        }
        read_only_param_keys = {"coeff_geo", "coeff_circ", "duree_trajet"}
        for section_name, (params, color) in PARAMS_FLUX_MAP.items():
            ws.merge_cells(f"A{curr_row}:D{curr_row}")
            ws[f"A{curr_row}"] = f"SECTION : {section_name}"
            ws[f"A{curr_row}"].font = Font(bold=True, color="FFFFFF")
            ws[f"A{curr_row}"].fill = PatternFill("solid", fgColor=color)
            ws[f"A{curr_row}"].alignment = s["center"]
            
            curr_row += 1
            ws.cell(row=curr_row, column=1, value="Clé").font = Font(bold=True)
            ws.cell(row=curr_row, column=2, value="Paramètre").font = Font(bold=True)
            ws.cell(row=curr_row, column=3, value="Valeur").font = Font(bold=True)
            ws.cell(row=curr_row, column=4, value="Unité").font = Font(bold=True)
            for col in range(1, 5):
                ws.cell(row=curr_row, column=col).fill = s["header_fill_gray"]
                ws.cell(row=curr_row, column=col).border = s["thin"]
                ws.cell(row=curr_row, column=col).alignment = s["center"]
            
            curr_row += 1
            for key, label, default_val, unit in params:
                ws.cell(row=curr_row, column=1, value=key).border = s["thin"]
                ws.cell(row=curr_row, column=2, value=label).border = s["thin"]
                v = param_override_values.get(key, default_val)
                v_cell = ws.cell(row=curr_row, column=3, value=v if v is not None else default_val)
                v_cell.fill = s["input_fill"]
                v_cell.border = s["thin"]
                v_cell.alignment = Alignment(horizontal="center")
                ws.cell(row=curr_row, column=4, value=unit).border = s["thin"]
                param_value_cell[key] = v_cell.coordinate
                # Déverrouiller uniquement les paramètres autorisés.
                # Les paramètres "ville" restent en lecture seule.
                v_cell.protection = Protection(locked=(key in read_only_param_keys))
                curr_row += 1
            
            curr_row += 1 # Espacement entre sections

        ws.column_dimensions["B"].width = 28
        ws.column_dimensions["C"].width = 12
        ws.column_dimensions["D"].width = 8

        # ── Formulas for Local/Axes (Wizard-like) ───────────────────────────
        # Local = ROUND(Global * pct_local/100, 0)
        # Axes  = ROUND(Global * pct_axes/100, 0)
        # For each flux, use its specific parameters when present; fallback to 0.
        def _pct_cells_for_flux(fx: str):
            fx_u = (fx or "").strip().upper()
            if fx_u == "AMANA":
                return (
                    param_value_cell.get("amana_pct_axes_depart"),
                    param_value_cell.get("amana_pct_axes_arrivee"),
                )
            if fx_u == "CR":
                return (
                    param_value_cell.get("cr_pct_axes_depart"),
                    param_value_cell.get("cr_pct_axes_arrivee"),
                )
            if fx_u == "CO":
                return (
                    param_value_cell.get("co_pct_axes_depart"),
                    param_value_cell.get("co_pct_axes_arrivee"),
                )
            return (None, None)

        # Only these rows have local/axes columns in the canvas
        flux_row_map = {"Amana": 5, "CR": 6, "CO": 7}
        for fx, row in flux_row_map.items():
            pct_local_cell, pct_axes_cell = _pct_cells_for_flux(fx)
            if not pct_local_cell or not pct_axes_cell:
                continue

            # For each 3-col group, apply formula:
            # group Global col -> Local col (global+1) and Axes col (global+2)
            for global_col in [2, 5, 8, 11]:  # B, E, H, K
                g_addr = ws.cell(row=row, column=global_col).coordinate
                local_addr = ws.cell(row=row, column=global_col + 1).coordinate
                axes_addr = ws.cell(row=row, column=global_col + 2).coordinate

                ws[local_addr].value = f"=ROUND({g_addr}*{pct_local_cell}/100,0)"
                ws[axes_addr].value = f"=ROUND({g_addr}*{pct_axes_cell}/100,0)"

        # ── Wizard Step2 rules in PARAMS (keep totals at 100%) ──────────────
        # Convention: l'utilisateur saisit la "valeur principale", la complémentaire est calculée.
        # - Axes/Local: on saisit Axes (arrivée), Local = 100 - Axes
        # - National/International: on saisit National, International = 100 - National
        # - CRBT/Hors CRBT: on saisit CRBT, Hors = 100 - CRBT
        # - Collecte/Marche/Guichet: on saisit Collecte + Marche, Guichet = 100 - Collecte - Marche
        def _set_calc_param(target_key: str, formula: str):
            coord = param_value_cell.get(target_key)
            if not coord:
                return
            cell = ws[coord]
            cell.value = formula
            cell.fill = s["locked_fill"]
            cell.alignment = Alignment(horizontal="center")
            cell.number_format = "0.00"
            cell.protection = Protection(locked=True)

        def _c(key: str) -> str | None:
            return param_value_cell.get(key)

        def _max0(expr: str) -> str:
            # Excel: MAX(0, expr) to avoid negative percentages
            return f"=MAX(0,{expr})"

        # Axes/Local by flux
        for axes_key, local_key in [
            ("amana_pct_axes_arrivee", "amana_pct_axes_depart"),
            ("co_pct_axes_arrivee", "co_pct_axes_depart"),
            ("cr_pct_axes_arrivee", "cr_pct_axes_depart"),
        ]:
            a = _c(axes_key)
            if a:
                _set_calc_param(local_key, _max0(f"100-{a}"))

        # National/International (Amana + CR)
        for nat_key, intl_key in [
            ("amana_pct_national", "amana_pct_international"),
            ("cr_pct_national", "cr_pct_international"),
        ]:
            n = _c(nat_key)
            if n:
                _set_calc_param(intl_key, _max0(f"100-{n}"))

        # CRBT / Hors CRBT (Amana + CR)
        for crbt_key, hors_key in [
            ("amana_pct_crbt", "amana_pct_hors_crbt"),
            ("cr_pct_crbt", "cr_pct_hors_crbt"),
        ]:
            c1 = _c(crbt_key)
            if c1:
                _set_calc_param(hors_key, _max0(f"100-{c1}"))

        # Collecte/Marche/Guichet (Amana + CO + CR)
        for collecte_key, marche_key, guichet_key in [
            ("amana_pct_collecte", "amana_pct_marche_ordinaire", "amana_pct_guichet"),
            ("co_pct_collecte", "co_pct_marche_ordinaire", "co_pct_guichet"),
            ("cr_pct_collecte", "cr_pct_marche_ordinaire", "cr_pct_guichet"),
        ]:
            c_col = _c(collecte_key)
            c_mar = _c(marche_key)
            if c_col and c_mar:
                _set_calc_param(guichet_key, _max0(f"100-{c_col}-{c_mar}"))

    return wb


# ─────────────────────────────────────────────────────────────────────────────
# Endpoint 1: Template Régional
# ─────────────────────────────────────────────────────────────────────────────
@router.get("/template/regional")
def get_regional_template(
    region_id: int = Query(..., description="ID de la région"),
    db: Session = Depends(get_db),
):
    region = db.execute(
        text("SELECT id, label FROM dbo.regions WHERE id = :rid"),
        {"rid": region_id}
    ).mappings().first()

    if not region:
        raise HTTPException(status_code=404, detail="Région introuvable")

    centres = db.execute(
        text("""
            SELECT
                c.id,
                c.label,
                r.label AS region_label,
                v.geographie AS coeff_geo_db,
                v.circulation AS coeff_circ_db,
                v.trajet AS duree_trajet_db
            FROM dbo.centres c
            JOIN dbo.regions r ON r.id = c.region_id
            LEFT JOIN dbo.Ville v ON v.Code = c.code_ville
            WHERE c.region_id = :rid
              AND c.categorie_id IS NOT NULL
            ORDER BY c.label
        """),
        {"rid": region_id}
    ).mappings().all()

    if not centres:
        raise HTTPException(status_code=404, detail="Aucun centre dans cette région")

    wb = _build_template_workbook([dict(c) for c in centres])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    region_name = region["label"].replace(" ", "_")
    filename = f"template_simulation_{region_name}.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# ─────────────────────────────────────────────────────────────────────────────
# Endpoint 2: Template National
# ─────────────────────────────────────────────────────────────────────────────
@router.get("/template/national")
def get_national_template(db: Session = Depends(get_db)):
    centres = db.execute(
        text("""
            SELECT
                c.id,
                c.label,
                r.label AS region_label,
                v.geographie AS coeff_geo_db,
                v.circulation AS coeff_circ_db,
                v.trajet AS duree_trajet_db
            FROM dbo.centres c
            JOIN dbo.regions r ON r.id = c.region_id
            LEFT JOIN dbo.Ville v ON v.Code = c.code_ville
            WHERE c.categorie_id IS NOT NULL
            ORDER BY r.label, c.label
        """)
    ).mappings().all()

    if not centres:
        raise HTTPException(status_code=404, detail="Aucun centre trouvé")

    wb = _build_template_workbook([dict(c) for c in centres])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=template_simulation_national.xlsx"},
    )


# ─────────────────────────────────────────────────────────────────────────────
# Parser: Excel → grid_values dict (same format as wizard)
# ─────────────────────────────────────────────────────────────────────────────
def _parse_sheet_grid(ws) -> dict:
    """Parse the canvas rows 5-9, returns grid_values dict matching wizard format."""

    def _v(row, col):
        v = ws.cell(row=row, column=col).value
        try:
            return float(v) if v is not None else 0.0
        except (TypeError, ValueError):
            return 0.0

    grid = {
        "amana": {
            "depot": {
                "gc":   {"global": _v(5, 2), "local": _v(5, 3), "axes": _v(5, 4)},
                "part": {"global": _v(5, 5), "local": _v(5, 6), "axes": _v(5, 7)},
            },
            "recu": {
                "gc":   {"global": _v(5, 8), "local": _v(5, 9), "axes": _v(5, 10)},
                "part": {"global": _v(5, 11), "local": _v(5, 12), "axes": _v(5, 13)},
            },
        },
        "cr": {
            "med":    {"global": _v(6, 2), "local": _v(6, 3), "axes": _v(6, 4)},
            "arrive": {"global": _v(6, 8), "local": _v(6, 9), "axes": _v(6, 10)},
        },
        "co": {
            "med":    {"global": _v(7, 2), "local": _v(7, 3), "axes": _v(7, 4)},
            "arrive": {"global": _v(7, 8), "local": _v(7, 9), "axes": _v(7, 10)},
        },
        "ebarkia": {"med": _v(8, 2), "arrive": _v(8, 8)},
        "lrh":     {"med": _v(9, 2), "arrive": _v(9, 8)},
    }
    return grid


def _recalculate_grid_values_like_wizard(grid_values: dict, raw_params: dict) -> dict:
    """
    Recalcule Local/Axes depuis Global selon les % Axes/Local du bloc paramètres.
    Nécessaire car lors de l'import, les formules Excel peuvent ne pas être évaluées
    (openpyxl `data_only=True` peut renvoyer None/0 pour les cellules à formule).
    """
    if not grid_values:
        return grid_values

    p = raw_params or {}

    def pct(key: str, default: float = 0.0) -> float:
        try:
            v = p.get(key, default)
            return float(v) if v is not None else float(default)
        except Exception:
            return float(default)

    def round0(x: float) -> float:
        try:
            return float(int(round(float(x or 0))))
        except Exception:
            return 0.0

    def recalc_flow(flow_obj: dict, pct_local: float, pct_axes: float):
        if not isinstance(flow_obj, dict):
            return

        # AMANA nested (gc/part)
        if "gc" in flow_obj and "part" in flow_obj:
            for sub in ("gc", "part"):
                sub_obj = flow_obj.get(sub)
                if not isinstance(sub_obj, dict):
                    continue
                g = float(sub_obj.get("global") or 0)
                sub_obj["local"] = round0(g * (pct_local / 100.0))
                sub_obj["axes"] = round0(g * (pct_axes / 100.0))
            return

        # Simple global/local/axes
        if "global" in flow_obj:
            g = float(flow_obj.get("global") or 0)
            flow_obj["local"] = round0(g * (pct_local / 100.0))
            flow_obj["axes"] = round0(g * (pct_axes / 100.0))

    # AMANA
    amana_local = pct("amana_pct_axes_depart", 0.0)
    amana_axes = pct("amana_pct_axes_arrivee", 0.0)
    if isinstance(grid_values.get("amana"), dict):
        for k in ("depot", "recu"):
            recalc_flow(grid_values["amana"].get(k), amana_local, amana_axes)

    # CR
    cr_local = pct("cr_pct_axes_depart", 0.0)
    cr_axes = pct("cr_pct_axes_arrivee", 0.0)
    if isinstance(grid_values.get("cr"), dict):
        for k in ("med", "arrive"):
            recalc_flow(grid_values["cr"].get(k), cr_local, cr_axes)

    # CO
    co_local = pct("co_pct_axes_depart", 0.0)
    co_axes = pct("co_pct_axes_arrivee", 0.0)
    if isinstance(grid_values.get("co"), dict):
        for k in ("med", "arrive"):
            recalc_flow(grid_values["co"].get(k), co_local, co_axes)

    return grid_values


def _parse_sheet_params(ws, start_row: int = 12) -> dict:
    """Scan rows from start_row looking for parameter keys in column 1."""
    params = {}
    # On scanne un range large car la section des paramètres s'est agrandie
    for row in range(start_row, start_row + 100):
        key = ws.cell(row=row, column=1).value
        val = ws.cell(row=row, column=3).value
        if key and isinstance(key, str) and not key.startswith("SECTION"):
            try:
                params[key.strip()] = float(val) if val is not None else None
            except (TypeError, ValueError):
                params[key.strip()] = None
    return params


def _normalize_params_like_wizard(params: dict) -> dict:
    """
    Applique les règles Step2 du Wizard sur les % pour garantir la cohérence:
    - complémentaire = 100 - valeur
    - guichet = 100 - collecte - marché
    """
    p = dict(params or {})

    def f(key: str, default: float = 0.0) -> float:
        try:
            v = p.get(key, default)
            return float(v) if v is not None else float(default)
        except Exception:
            return float(default)

    def clamp01_100(x: float) -> float:
        try:
            return max(0.0, min(100.0, float(x)))
        except Exception:
            return 0.0

    def set_comp(primary_key: str, comp_key: str):
        v = clamp01_100(f(primary_key, 0.0))
        p[primary_key] = v
        p[comp_key] = clamp01_100(100.0 - v)

    def set_triple(collecte_key: str, marche_key: str, guichet_key: str):
        c = clamp01_100(f(collecte_key, 0.0))
        m = clamp01_100(f(marche_key, 0.0))
        p[collecte_key] = c
        p[marche_key] = m
        p[guichet_key] = clamp01_100(100.0 - c - m)

    # Axes/Local
    set_comp("amana_pct_axes_arrivee", "amana_pct_axes_depart")
    set_comp("co_pct_axes_arrivee", "co_pct_axes_depart")
    set_comp("cr_pct_axes_arrivee", "cr_pct_axes_depart")

    # National/International
    set_comp("amana_pct_national", "amana_pct_international")
    set_comp("cr_pct_national", "cr_pct_international")

    # CRBT / Hors CRBT
    set_comp("amana_pct_crbt", "amana_pct_hors_crbt")
    set_comp("cr_pct_crbt", "cr_pct_hors_crbt")

    # Collecte / Marché / Guichet
    set_triple("amana_pct_collecte", "amana_pct_marche_ordinaire", "amana_pct_guichet")
    set_triple("co_pct_collecte", "co_pct_marche_ordinaire", "co_pct_guichet")
    set_triple("cr_pct_collecte", "cr_pct_marche_ordinaire", "cr_pct_guichet")

    return p


def _fill_city_params_from_db_if_missing(db: Session, centre_id: int, raw_params: dict) -> dict:
    """
    Dans le Wizard, coeff_geo/coeff_circ/duree_trajet sont pré-remplis depuis la BD (Ville du centre).
    Dans Excel, on les pré-remplit dans le template, mais à l'import on garantit aussi le fallback BD
    si l'utilisateur laisse vide ou si le fichier n'a pas conservé ces cellules.
    """
    p = dict(raw_params or {})

    def is_missing(k: str) -> bool:
        return k not in p or p.get(k) is None

    if not (is_missing("coeff_geo") or is_missing("coeff_circ") or is_missing("duree_trajet")):
        return p

    row = db.execute(
        text(
            """
            SELECT
                v.geographie AS coeff_geo_db,
                v.circulation AS coeff_circ_db,
                v.trajet AS duree_trajet_db
            FROM dbo.centres c
            LEFT JOIN dbo.Ville v ON v.Code = c.code_ville
            WHERE c.id = :cid
            """
        ),
        {"cid": int(centre_id)},
    ).mappings().first()

    if not row:
        return p

    if is_missing("coeff_geo") and row.get("coeff_geo_db") is not None:
        p["coeff_geo"] = float(row["coeff_geo_db"])
    if is_missing("coeff_circ") and row.get("coeff_circ_db") is not None:
        p["coeff_circ"] = float(row["coeff_circ_db"])
    if is_missing("duree_trajet") and row.get("duree_trajet_db") is not None:
        p["duree_trajet"] = float(row["duree_trajet_db"])

    return p


def _params_to_engine(p: dict) -> BandoengParameters:
    def g(k, default=0.0):
        v = p.get(k)
        return float(v) if v is not None else default

    return BandoengParameters(
        # Globals
        shift=int(g("shift", 1)),
        productivite=g("productivite", 100.0),
        idle_minutes=g("idle_minutes", 0.0),
        coeff_circ=g("coeff_circ", 1.0),
        coeff_geo=g("coeff_geo", 1.0),
        has_guichet=1, # Fixe à 1 (logique backend conservée selon specification)
        duree_trajet=g("duree_trajet", 0.0),

        # Mappage des paramètres (Flux vers les champs déjà existants dans le moteur)
        ed_percent=g("ed_percent", 40.0),
        
        # Amana
        amana_pct_collecte=g("amana_pct_collecte", 0.0),
        amana_pct_guichet=g("amana_pct_guichet", 0.0),
        amana_pct_retour=g("amana_pct_retour", 0.0),
        amana_pct_axes_arrivee=g("amana_pct_axes_arrivee", 0.0),
        amana_pct_axes_depart=g("amana_pct_axes_depart", 0.0),
        amana_pct_national=g("amana_pct_national", 0.0),
        amana_pct_international=g("amana_pct_international", 0.0),
        amana_pct_marche_ordinaire=g("amana_pct_marche_ordinaire", 0.0),
        amana_pct_crbt=g("amana_pct_crbt", 0.0),
        amana_pct_hors_crbt=g("amana_pct_hors_crbt", 0.0),
        colis_amana_par_canva_sac=g("colis_amana_par_canva_sac", 35.0),

        # CO
        co_pct_collecte=g("co_pct_collecte", 0.0),
        co_pct_guichet=g("co_pct_guichet", 0.0),
        co_pct_marche_ordinaire=g("co_pct_marche_ordinaire", 0.0),
        co_pct_axes_arrivee=g("co_pct_axes_arrivee", 0.0),
        co_pct_axes_depart=g("co_pct_axes_depart", 0.0),
        co_pct_vague_master=g("co_pct_vague_master", 0.0),
        co_pct_boite_postale=g("co_pct_boite_postale", 0.0),
        nbr_co_sac=g("nbr_co_sac", 350.0),

        # CR
        cr_pct_collecte=g("cr_pct_collecte", 0.0),
        cr_pct_guichet=g("cr_pct_guichet", 0.0),
        cr_pct_retour=g("cr_pct_retour", 0.0),
        cr_pct_axes_arrivee=g("cr_pct_axes_arrivee", 0.0),
        cr_pct_axes_depart=g("cr_pct_axes_depart", 0.0),
        cr_pct_national=g("cr_pct_national", 0.0),
        cr_pct_international=g("cr_pct_international", 0.0),
        cr_pct_marche_ordinaire=g("cr_pct_marche_ordinaire", 0.0),
        cr_pct_vague_master=g("cr_pct_vague_master", 0.0),
        cr_pct_crbt=g("cr_pct_crbt", 0.0),
        cr_pct_hors_crbt=g("cr_pct_hors_crbt", 0.0),
        # Consolidation : on utilise cr_par_caisson pour les deux champs du moteur
        nbr_cr_sac=g("cr_par_caisson", 500.0),
        cr_par_caisson=g("cr_par_caisson", 500.0),
    )


# ─────────────────────────────────────────────────────────────────────────────
# Endpoint 3: Simulate Batch
# ─────────────────────────────────────────────────────────────────────────────
@router.post("/simulate")
async def simulate_batch(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    region_id: Optional[int] = Query(default=None, description="Filtre centres par région (pour template régional)"),
    process_mode: str = Query(default="actuel", description="Mode de calcul : actuel | recommande | optimise"),
):
    """
    Parse the imported Excel file and run Bandoeng simulation for each sheet.
    Returns results per centre + aggregated by region + national total.
    """
    content = await file.read()

    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Fichier Excel invalide : {str(e)}")

    # Build mapping sheet_name -> centre (strict matching)
    # On limite d'abord les centres à la région si `region_id` est fourni.
    if region_id:
        centres_db = db.execute(
            text(
                """
                SELECT c.id, c.label, r.id AS region_id, r.label AS region_label
                FROM dbo.centres c
                JOIN dbo.regions r ON r.id = c.region_id
                WHERE c.region_id = :rid
                  AND c.categorie_id IS NOT NULL
                ORDER BY c.label
                """
            ),
            {"rid": int(region_id)},
        ).mappings().all()
    else:
        centres_db = db.execute(
            text(
                """
                SELECT c.id, c.label, r.id AS region_id, r.label AS region_label
                FROM dbo.centres c
                JOIN dbo.regions r ON r.id = c.region_id
                WHERE c.categorie_id IS NOT NULL
                ORDER BY r.label, c.label
                """
            ),
        ).mappings().all()

    def _safe_sheet_name(c_label: str, used_names: set) -> str:
        safe = "".join(x for x in c_label if x not in r'[]:*?/\\')[:28]
        name = safe
        idx = 1
        while name in used_names:
            name = f"{safe[:26]}{idx}"
            idx += 1
        used_names.add(name)
        return name

    def _normalize_sheet_key(name: str) -> str:
        """
        Normalise un nom d'onglet ou de centre pour le matching:
        - trim
        - passe en minuscule
        - supprime tous les espaces (y compris multiples)
        """
        if not name:
            return ""
        # "".join(split()) supprime tous les espaces, tabulations, etc.
        return "".join(str(name).split()).lower()

    used_names: set = set()
    sheet_name_to_centre = {}
    for c in centres_db:
        c_label = (c.get("label") or "").strip()
        sheet_name = _safe_sheet_name(c_label, used_names)
        key = _normalize_sheet_key(sheet_name)
        sheet_name_to_centre[key] = dict(c)

    results_par_centre = []
    errors = []

    # ─── Pré-chargement des données selon le process_mode ────────────────────
    # Consolidé : chargement global des mappings de postes (source → cible)
    role_mapping_global: Optional[dict] = None
    if process_mode == "recommande" and MappingPosteRecommande is not None:
        try:
            mappings = db.query(MappingPosteRecommande).all()
            role_mapping_global = {
                m.poste_source.Code: m.poste_cible.Code
                for m in mappings
                if m.poste_source and m.poste_cible
            }
        except Exception:
            role_mapping_global = {}

    # Optimisé : chargement des exclusions par centre et par quadruplet
    # On construit une map { centre_id: (excluded_ids, excluded_quadruplets) }
    optimise_excl_by_centre: dict = {}
    if process_mode == "optimise" and TacheExclueOptimisee is not None:
        try:
            def _clean(s):
                if not s: return ""
                return "".join(str(s).split()).lower()

            # Exclusions par ID (liées à un centre spécifique)
            all_id_excl = (
                db.query(TacheExclueOptimisee.centre_id, TacheExclueOptimisee.tache_id)
                .filter(TacheExclueOptimisee.centre_id.isnot(None),
                        TacheExclueOptimisee.tache_id.isnot(None))
                .all()
            )
            for cid, tid in all_id_excl:
                optimise_excl_by_centre.setdefault(cid, {"ids": [], "quads": []})["ids"].append(tid)

            # Exclusions par quadruplet (liées à une catégorie, appliquées à tout centre de la catégorie)
            all_quad_excl = (
                db.query(
                    TacheExclueOptimisee.categorie_id,
                    TacheExclueOptimisee.nom_tache,
                    TacheExclueOptimisee.produit,
                    TacheExclueOptimisee.famille_uo,
                    TacheExclueOptimisee.unite_mesure,
                )
                .filter(TacheExclueOptimisee.nom_tache.isnot(None))
                .all()
            )
            # Map catégorie_id → list of quadruplets
            quads_by_cat = {}
            for cat_id, nom, produit, famille, unite in all_quad_excl:
                if cat_id not in quads_by_cat:
                    quads_by_cat[cat_id] = []
                quads_by_cat[cat_id].append(
                    (_clean(nom), _clean(produit), _clean(famille), _clean(unite))
                )
        except Exception:
            pass

    for sheet_name in wb.sheetnames:
        if sheet_name == "Guide":
            continue

        ws = wb[sheet_name]

        # Matching insensible aux espaces et à la casse
        lookup_key = _normalize_sheet_key(sheet_name)
        match = sheet_name_to_centre.get(lookup_key)
        if not match:
            errors.append({"sheet": sheet_name, "error": "Centre non trouvé (matching strict onglet)"})
            continue

        centre_id = match["id"]
        centre_label = match["label"]
        region_id = match["region_id"]
        region_label = match["region_label"]

        try:
            # ✅ AUTO-IMPORT if empty
            auto_import_tasks_if_empty(db, centre_id)

            raw_params = _parse_sheet_params(ws)
            raw_params = _normalize_params_like_wizard(raw_params)
            raw_params = _fill_city_params_from_db_if_missing(db, centre_id, raw_params)
            grid_values = _parse_sheet_grid(ws)
            grid_values = _recalculate_grid_values_like_wizard(grid_values, raw_params)
            engine_params = _params_to_engine(raw_params)

            # Résoudre role_mapping et exclusions pour ce centre selon le mode
            role_mapping = role_mapping_global  # None ou dict (recommande)
            excluded_task_ids = None
            excluded_task_quadruplets = None
            if process_mode == "optimise" and TacheExclueOptimisee is not None:
                centre_excl = optimise_excl_by_centre.get(centre_id, {})
                excluded_task_ids = centre_excl.get("ids") or None
                # Quadruplets par catégorie du centre
                centre_obj = db.query(Centre).filter(Centre.id == centre_id).first()
                if centre_obj and centre_obj.categorie_id and centre_obj.categorie_id in quads_by_cat:
                    excluded_task_quadruplets = quads_by_cat[centre_obj.categorie_id]

            volumes = BandoengInputVolumes(grid_values=grid_values)

            # Récupérer les effectifs actuels MOI/MOD/APS (alignement avec Wizard)
            rows_postes = (
                db.query(
                    CentrePoste.effectif_actuel,
                    CentrePoste.aps,
                    Poste.type_poste,
                    Poste.label,
                    Poste.charge_salaire,
                )
                .join(Poste, CentrePoste.code_resp == Poste.Code)
                .filter(CentrePoste.centre_id == centre_id)
                .all()
            )
            actual_moi = 0.0
            actual_mod = 0.0
            actual_aps = 0.0
            actual_aps_mod = 0.0
            effectifs_par_poste: dict = {}
            for eff, aps, type_poste, poste_label, _charge in rows_postes:
                eff_val = float(eff or 0)
                aps_val = float(aps or 0)
                total_poste = eff_val + aps_val
                t = (type_poste or "").upper()
                is_moi = t in ["MOI", "INDIRECT", "STRUCTURE"]
                if is_moi:
                    actual_moi += eff_val
                else:
                    actual_mod += eff_val
                    actual_aps_mod += aps_val
                actual_aps += aps_val
                if not is_moi and total_poste > 0 and poste_label:
                    effectifs_par_poste[poste_label] = round(total_poste, 2)

            result = run_bandoeng_simulation(
                db, centre_id, volumes, engine_params,
                role_mapping=role_mapping,
                excluded_task_ids=excluded_task_ids,
                excluded_task_quadruplets=excluded_task_quadruplets,
            )

            rpp = result.ressources_par_poste or {}

            def _rpp_match_etp(lab: str) -> float:
                if not lab:
                    return 0.0
                L = str(lab).strip()
                if L in rpp:
                    return float(rpp[L] or 0)
                lu = L.upper()
                for k, v in rpp.items():
                    if str(k).strip().upper() == lu:
                        return float(v or 0)
                return 0.0

            postes_chiffrage: list = []
            for eff, aps, type_poste, poste_label, charge_salaire in rows_postes:
                eff_val = float(eff or 0)
                aps_val = float(aps or 0)
                total_poste = eff_val + aps_val
                t = (type_poste or "").upper()
                is_moi = t in ["MOI", "INDIRECT", "STRUCTURE"]
                if is_moi:
                    continue
                plab = (poste_label or "").strip()
                if not plab:
                    continue
                sim_etp = int(round(_rpp_match_etp(plab)))
                sal = float(charge_salaire or 0)
                if total_poste <= 0 and sim_etp <= 0:
                    continue
                postes_chiffrage.append({
                    "label": plab,
                    "type_poste": type_poste or "",
                    "charge_salaire": sal,
                    "actuel_etp": round(total_poste, 2),
                    "simule_etp": sim_etp,
                })

            fte_calcule = round(result.fte_calcule, 2)
            # Cohérence visuelle: `arrondi` doit correspondre à `calculé`
            fte_total_calcule = fte_calcule
            fte_total_arrondi = int(round(fte_calcule))
            fte_arrondi = int(round(fte_calcule))

            results_par_centre.append({
                "centre_id": centre_id,
                "centre_label": centre_label,
                "region_id": region_id,
                "region_label": region_label,
                "fte_calcule": fte_calcule,
                "fte_arrondi": fte_arrondi,
                "fte_total_calcule": fte_total_calcule,
                "fte_total_arrondi": fte_total_arrondi,
                "actual_moi": round(actual_moi, 2),
                "actual_mod": round(actual_mod, 2),
                "actual_aps": round(actual_aps, 2),
                "actual_aps_mod": round(actual_aps_mod, 2),
                "total_heures": round(result.total_heures, 2),
                "ressources_par_poste": result.ressources_par_poste,
                "effectifs_par_poste": effectifs_par_poste,
                "postes_chiffrage": postes_chiffrage,
            })

        except Exception as e:
            errors.append({"sheet": sheet_name, "centre": centre_label, "error": str(e)})

    # Aggregate by region
    par_region: dict = {}
    for r in results_par_centre:
        rid = r["region_id"]
        if rid not in par_region:
            par_region[rid] = {
                "region_id": rid,
                "region_label": r["region_label"],
                "total_fte_calcule": 0.0,
                "total_fte_arrondi": 0,
                "total_fte_total_calcule": 0.0,
                "total_fte_total_arrondi": 0,
                "nb_centres": 0,
            }
        par_region[rid]["total_fte_calcule"] += r["fte_calcule"]
        par_region[rid]["total_fte_arrondi"] += r["fte_arrondi"]
        par_region[rid]["total_fte_total_calcule"] += r.get("fte_total_calcule", r["fte_calcule"])
        par_region[rid]["total_fte_total_arrondi"] += r.get("fte_total_arrondi", r["fte_arrondi"])
        par_region[rid]["nb_centres"] += 1

    national_total = sum(r["fte_calcule"] for r in results_par_centre)
    national_arrondi = sum(r["fte_arrondi"] for r in results_par_centre)
    national_total_total = sum(r.get("fte_total_calcule", r["fte_calcule"]) for r in results_par_centre)
    national_arrondi_total = sum(r.get("fte_total_arrondi", r["fte_arrondi"]) for r in results_par_centre)

    return {
        "par_centre": results_par_centre,
        "par_region": list(par_region.values()),
        "national": {
            "total_fte_calcule": round(national_total, 2),
            "total_fte_arrondi": national_arrondi,
            "total_fte_total_calcule": round(national_total_total, 2),
            "total_fte_total_arrondi": national_arrondi_total,
            "nb_centres": len(results_par_centre),
        },
        "errors": errors,
    }


@router.post("/simulate-comparatif")
async def simulate_batch_comparatif(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    region_id: Optional[int] = Query(default=None),
):
    """
    Lance la simulation batch pour les 3 modes (actuel / recommande / optimise)
    avec le même fichier Excel et renvoie les 3 résultats pour comparatif.
    """
    content = await file.read()

    async def _run_mode(mode: str):
        mock_file = UploadFile(
            filename=file.filename,
            file=io.BytesIO(content),
        )
        return await simulate_batch(
            file=mock_file,
            db=db,
            region_id=region_id,
            process_mode=mode,
        )

    actuel, recommande, optimise = await asyncio.gather(
        _run_mode("actuel"),
        _run_mode("recommande"),
        _run_mode("optimise"),
    )

    return {
        "actuel":     actuel,
        "recommande": recommande,
        "optimise":   optimise,
    }

