from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter

from app.core.db import get_db
from app.core.config import settings
from app.schemas.comparaison_effectif import (
    ComparaisonEffectifRequest,
    ComparaisonEffectifResponse,
)
from app.services.comparaison_effectif import compute_comparaison_effectif

router = APIRouter(prefix="/comparaison-effectif", tags=["comparaison-effectif"])


@router.post("/simuler", response_model=ComparaisonEffectifResponse)
def simuler(payload: ComparaisonEffectifRequest, db: Session = Depends(get_db)):
    if payload.productivite_pct <= 0:
        raise HTTPException(status_code=400, detail="Productivité doit être > 0")
    return compute_comparaison_effectif(db, payload.sacs_jour, payload.dossiers_mois, payload.productivite_pct)


@router.get("/export-xlsx")
def export_xlsx(
    sacs_jour: int = Query(..., ge=0),
    dossiers_mois: int = Query(..., ge=0),
    productivite_pct: float = Query(..., gt=0),
    db: Session = Depends(get_db),
):
    data = compute_comparaison_effectif(db, sacs_jour, dossiers_mois, productivite_pct)
    rows = data["rows"]
    total = data["total"]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.sheet_view.showGridLines = False
    ws.title = "Comparaison"
    ws.row_dimensions[1].height = 60
    ws.row_dimensions[2].height = 28
    ws.row_dimensions[3].height = 20

    ws.merge_cells("B1:E1")
    title_cell = ws["B1"]
    title_cell.value = "Comparatif Effectifs"
    title_cell.font = Font(size=16, bold=True, color="005EA8")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    ws["B2"] = ""
    ws["B2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("B2:C2")

    start_param_row = 4
    ws.cell(row=start_param_row, column=1, value="Paramètres de simulation").font = Font(size=12, bold=True, color="1F2937", underline="single")
    ws.merge_cells(f"A{start_param_row}:D{start_param_row}")

    ws.cell(row=start_param_row + 1, column=1, value="Sacs / Jour").font = Font(bold=True)
    ws.cell(row=start_param_row + 1, column=2, value=sacs_jour).alignment = Alignment(horizontal="left")

    ws.cell(row=start_param_row + 2, column=1, value="Dossiers / Mois").font = Font(bold=True)
    ws.cell(row=start_param_row + 2, column=2, value=dossiers_mois).alignment = Alignment(horizontal="left")

    ws.cell(row=start_param_row + 3, column=1, value="Taux Productivité").font = Font(bold=True)
    ws.cell(row=start_param_row + 3, column=2, value=f"{productivite_pct}%").alignment = Alignment(horizontal="left")

    ws.cell(row=start_param_row + 4, column=1, value="Temps Mort").font = Font(bold=True)
    ws.cell(row=start_param_row + 4, column=2, value=0).alignment = Alignment(horizontal="left")

    ws.cell(row=start_param_row + 5, column=1, value="Heures Travaillées / Jour").font = Font(bold=True)
    ws.cell(row=start_param_row + 5, column=2, value=(8.0 * productivite_pct) / 100.0).alignment = Alignment(horizontal="left")

    try:
        img_barid = Image(str(settings.assets_dir / settings.LOGO_BARID))
        img_barid.height = 60
        img_barid.width = 150
        ws.add_image(img_barid, "A1")
    except Exception as exc:
        print(f"⚠️ Impossible de charger le logo Barid: {exc}")

    try:
        img_almav = Image(str(settings.assets_dir / settings.LOGO_ALMAV))
        img_almav.height = 50
        img_almav.width = 120
        ws.add_image(img_almav, "F1")
    except Exception as exc:
        print(f"⚠️ Impossible de charger le logo Almav: {exc}")

    headers = [
        "Position",
        "Effectif Actuel",
        "FTE Calculé",
        "FTE Calculé Arrondi",
        "Écart (FTE)",
        "Écart Arrondi",
    ]
    table_start = start_param_row + 7
    for col_index, header_title in enumerate(headers, start=1):
        cell = ws.cell(row=table_start, column=col_index, value=header_title)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="005EA8")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    data_start = table_start + 1
    current_row = data_start
    for r in rows:
        ws.cell(row=current_row, column=1, value=r["position"])
        ws.cell(row=current_row, column=2, value=r["effectif_actuel"])
        ws.cell(row=current_row, column=3, value=r["fte_calcule"])
        ws.cell(row=current_row, column=4, value=r["fte_arrondi"])
        ws.cell(row=current_row, column=5, value=r["ecart_fte"])
        ws.cell(row=current_row, column=6, value=r["ecart_arrondi"])
        current_row += 1

    current_row += 1
    total_row = [
        "TOTAL GÉNÉRAL",
        total["effectif_actuel"],
        total["fte"],
        total["fte_arrondi"],
        total["ecart_fte"],
        total["ecart_arrondi"],
    ]
    for col_index, value in enumerate(total_row, start=1):
        cell = ws.cell(row=current_row, column=col_index, value=value)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="e8f0fe")

    for idx, col in enumerate(ws.columns, start=1):
        max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
        ws.column_dimensions[get_column_letter(idx)].width = max(12, max_len + 2)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    headers_resp = {
        "Content-Disposition": 'attachment; filename="comparaison_effectif.xlsx"',
        "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    }
    return StreamingResponse(buf, headers=headers_resp)
