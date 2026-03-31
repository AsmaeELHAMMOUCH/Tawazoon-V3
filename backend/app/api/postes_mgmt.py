# app/api/postes_mgmt.py
"""
API pour la gestion CRUD des postes par centre.
Permet de gérer la relation centre_postes (effectifs actuels par poste dans un centre).
"""
# Force reload
from typing import List, Optional
from fastapi import APIRouter, Depends, HTTPException, Query, File, UploadFile
from fastapi.responses import StreamingResponse
import pandas as pd
import io
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session
from sqlalchemy import text
from pydantic import BaseModel

from app.core.db import get_db
from app.models.db_models import CentrePoste, Poste, Centre

router = APIRouter(prefix="/pm", tags=["Postes Management"])

@router.get("/global/postes", response_model=List[dict])
def get_all_centre_postes(db: Session = Depends(get_db)):
    """
    Récupère tous les postes de tous les centres.
    """
    centre_postes = db.query(CentrePoste).all()
    
    result = []
    for cp in centre_postes:
        c = cp.centre
        p = cp.poste
        
        if not c or not p: continue
        
        region_label = "N/A"
        if c.region:
             region_label = c.region.label
        
        result.append({
            "centre_poste_id": cp.id,
            "poste_id": p.id,
            "poste_label": p.label,
            "centre_id": c.id,
            "centre_label": c.label,
            "region_label": region_label,
            "effectif_actuel": cp.effectif_actuel,
            "effectif_aps": cp.aps if cp.aps is not None else 0
        })
    return result

class ApsUpdate(BaseModel):
    aps: float

@router.put("/centres/{centre_id}/aps")
def update_centre_aps(centre_id: int, update: ApsUpdate, db: Session = Depends(get_db)):
    """
    Met à jour l'APS d'un centre spécifié.
    """
    centre = db.query(Centre).filter(Centre.id == centre_id).first()
    if not centre:
        raise HTTPException(status_code=404, detail="Centre non trouvé")
    
    centre.aps = update.aps
    db.commit()
    return {"status": "success", "message": f"APS du centre {centre.label} mis à jour"}

@router.get("/aps/export-template")
def export_aps_template(
    region_id: Optional[int] = Query(None),
    typologie_id: Optional[int] = Query(None),
    db: Session = Depends(get_db)
):
    """
    Exporte un template Excel des centres filtrés avec leur APS actuel.
    """
    query = db.query(Centre)
    if region_id:
        query = query.filter(Centre.region_id == region_id)
    if typologie_id:
        query = query.filter(Centre.categorie_id == typologie_id)
    
    centres = query.all()
    
    data = []
    for c in centres:
        data.append({
            "Région": c.region.label if c.region else "N/A",
            "Centre": c.label,
            "APS Actuel": c.aps if c.aps is not None else 0.0
        })
    
    df = pd.DataFrame(data)
    
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Update_APS')
        workbook = writer.book
        worksheet = writer.sheets['Update_APS']
        
        # Styles
        header_fill = PatternFill(start_color="005EA8", end_color="005EA8", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        center_align = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin', color='DDDDDD'),
            right=Side(style='thin', color='DDDDDD'),
            top=Side(style='thin', color='DDDDDD'),
            bottom=Side(style='thin', color='DDDDDD')
        )
        
        # Formater l'en-tête
        for col_num, column_title in enumerate(df.columns, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = border
            
            # Largeur automatique
            max_length = max(df[column_title].astype(str).map(len).max(), len(column_title)) + 5
            worksheet.column_dimensions[get_column_letter(col_num)].width = max_length
            
        # Figer les volets
        worksheet.freeze_panes = 'A2'
    
    output.seek(0)
    
    filename = "template_aps_global.xlsx"
    if region_id:
        region = db.query(text("label from dbo.regions where id = :id")).params(id=region_id).first()
        if region:
            filename = f"template_aps_{region[0]}.xlsx"
            
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@router.post("/aps/import")
async def import_aps(file: UploadFile = File(...), db: Session = Depends(get_db)):
    """
    Importe un fichier Excel pour mettre à jour les APS des centres par lot.
    """
    try:
        contents = await file.read()
        df = pd.read_excel(io.BytesIO(contents))
        
        required_cols = ["Centre", "APS Actuel"]
        if not all(col in df.columns for col in required_cols):
            raise HTTPException(status_code=400, detail=f"Colonnes manquantes. Requis: {required_cols}")
        
        updated_count = 0
        errors = []
        
        for index, row in df.iterrows():
            centre_name_raw = str(row["Centre"]).strip()
            new_aps_raw = row["APS Actuel"]
            
            try:
                if pd.isna(new_aps_raw) or str(new_aps_raw).strip() == "":
                    new_aps_float = 0.0
                else:
                    new_aps_float = float(str(new_aps_raw).replace(',', '.'))
            except ValueError:
                errors.append(f"Ligne {index+2}: Valeur APS invalide '{new_aps_raw}'")
                continue
            
            # Recherche du centre par nom (normalisé)
            centre = db.query(Centre).filter(text("LOWER(REPLACE(label, ' ', '')) = LOWER(REPLACE(:name, ' ', ''))")).params(name=centre_name_raw).first()
            
            if centre:
                centre.aps = new_aps_float
                updated_count += 1
            else:
                errors.append(f"Ligne {index+2}: Centre '{centre_name_raw}' non trouvé")
        
        db.commit()
        
        if errors:
            error_data = [{"Erreur": err} for err in errors]
            df_errors = pd.DataFrame(error_data)
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df_errors.to_excel(writer, index=False, sheet_name='Lignes_Rejetees')
            output.seek(0)
            
            headers = {
                "Content-Disposition": "attachment; filename=rejets_import_aps.xlsx",
                "Access-Control-Expose-Headers": "X-Error-Count, X-Updated-Count",
                "X-Error-Count": str(len(errors)),
                "X-Updated-Count": str(updated_count)
            }
            return StreamingResponse(
                output,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers=headers
            )
            
        return {
            "status": "success", 
            "message": f"{updated_count} centres mis à jour",
            "errors": []
        }
    except Exception as e:
        import traceback
        traceback.print_exc()
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/effectifs/export-template")
def export_effectifs_template(
    region_id: Optional[int] = Query(None),
    typologie_id: Optional[int] = Query(None),
    centre_id: Optional[int] = Query(None),
    db: Session = Depends(get_db)
):
    """
    Exporte un template Excel des postes par centre selon les filtres.
    """
    if centre_id:
        # Si un centre est spécifié, on utilise LEFT JOIN pour qu'il apparaisse même s'il n'a pas encore de postes
        query = text("""
            SELECT 
                r.label as region,
                c.label as centre,
                p.label as poste,
                COALESCE(cp.effectif_actuel, 0) as effectif,
                COALESCE(cp.aps, 0) as aps
            FROM dbo.centres c
            JOIN dbo.regions r ON r.id = c.region_id
            JOIN dbo.centre_postes cp ON cp.centre_id = c.id
            JOIN dbo.postes p ON p.Code = cp.code_resp
            WHERE c.id = :centre_id AND (COALESCE(cp.effectif_actuel, 0) > 0 OR COALESCE(cp.aps, 0) > 0)
        """)
        rows = db.execute(query, {"centre_id": centre_id}).mappings().all()
    else:
        # Sinon, on liste tous les couples Centre/Poste existants filtrés par région/typologie
        filters = []
        params = {}
        if region_id:
            filters.append("c.region_id = :region_id")
            params["region_id"] = region_id
        if typologie_id:
            filters.append("c.categorie_id = :typologie_id")
            params["typologie_id"] = typologie_id
        
        filters.append("(COALESCE(cp.effectif_actuel, 0) > 0 OR COALESCE(cp.aps, 0) > 0)")
        filter_str = " AND ".join(filters)
        
        query = text(f"""
            SELECT 
                r.label as region,
                c.label as centre,
                p.label as poste,
                COALESCE(cp.effectif_actuel, 0) as effectif,
                COALESCE(cp.aps, 0) as aps
            FROM dbo.centres c
            JOIN dbo.regions r ON r.id = c.region_id
            JOIN dbo.centre_postes cp ON cp.centre_id = c.id
            JOIN dbo.postes p ON p.Code = cp.code_resp
            WHERE {filter_str}
            ORDER BY c.label, p.label
        """)
        rows = db.execute(query, params).mappings().all()

    data = []
    for row in rows:
        data.append({
            "Région": row["region"],
            "Centre": row["centre"],
            "Poste": row["poste"] if row["poste"] else "",
            "Statutaires": row["effectif"],
            "APS": row["aps"]
        })
    
    # Sincérité des colonnes même si data est vide
    df = pd.DataFrame(data, columns=["Région", "Centre", "Poste", "Statutaires", "APS"])
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Update_Effectifs', startrow=1)
        workbook = writer.book
        worksheet = writer.sheets['Update_Effectifs']
        
        # Ajouter l'instruction sur la première ligne
        worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))
        instruction_cell = worksheet.cell(row=1, column=1)
        instruction_cell.value = "💡 NOTE : Pour affecter un nouveau poste à un centre, ajoutez simplement une ligne dans le fichier avec le nom du centre, du poste et les effectifs."
        instruction_cell.font = Font(italic=True, color="444444", size=10, bold=True)
        instruction_cell.alignment = Alignment(horizontal="left", vertical="center")
        worksheet.row_dimensions[1].height = 25
        
        # Styles pour les en-têtes (à la ligne 2)
        header_fill = PatternFill(start_color="005EA8", end_color="005EA8", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        center_align = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin', color='DDDDDD'),
            right=Side(style='thin', color='DDDDDD'),
            top=Side(style='thin', color='DDDDDD'),
            bottom=Side(style='thin', color='DDDDDD')
        )
        
        # Formater l'en-tête (ligne 2)
        for col_num, column_title in enumerate(df.columns, 1):
            cell = worksheet.cell(row=2, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = border
            
            # Largeur automatique
            max_length = max(df[column_title].astype(str).map(len).max(), len(column_title)) + 8
            worksheet.column_dimensions[get_column_letter(col_num)].width = max_length
            
        # Figer les volets
        worksheet.freeze_panes = 'A3'
    
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=template_effectifs.xlsx"}
    )

@router.post("/effectifs/import")
async def import_effectifs(file: UploadFile = File(...), db: Session = Depends(get_db)):
    """
    Importe un fichier Excel pour mettre à jour ou ajouter des effectifs (Upsert).
    """
    try:
        contents = await file.read()
        
        # Détection automatique du début du tableau
        df_temp = pd.read_excel(io.BytesIO(contents), nrows=5)
        skip = 0
        if "Centre" not in df_temp.columns:
            df_check = pd.read_excel(io.BytesIO(contents), skiprows=1, nrows=5)
            if "Centre" in df_check.columns:
                skip = 1
        
        df = pd.read_excel(io.BytesIO(contents), skiprows=skip)
        
        required_cols = ["Centre", "Poste", "Statutaires"]
        # On rend "APS" optionnel pour la compatibilité avec les anciens templates
        has_aps_col = "APS" in df.columns
        
        updated_count = 0
        created_count = 0
        errors = []
        
        # Track processed combinations to prevent IntegrityError on duplicates
        processed_combinations = set()
        # Track which centers were touched and which posts were actually in the file
        affected_center_ids = set()
        successfully_processed_keys = set() # Set of (centre_id, poste_id)
        
        import re
        def norm(s):
            if not s: return ""
            return re.sub(r'\s+', '', str(s)).lower()

        # Pré-charger les centres et les postes pour éviter trop de requêtes unitaires
        all_centres = {norm(c.label): c.id for c in db.query(Centre).all()}
        all_postes_by_label = {norm(p.label): (p.id, p.Code) for p in db.query(Poste).all()}
        
        for index, row in df.iterrows():
            centre_name_raw = str(row["Centre"]).strip()
            poste_label_raw = str(row["Poste"]).strip()
            new_val = row["Statutaires"]
            try:
                if pd.isna(new_val) or str(new_val).strip() == "":
                    new_val_float = 0.0
                else:
                    new_val_float = float(str(new_val).replace(',', '.'))
                
                new_aps_float = 0.0
                if has_aps_col:
                    aps_val = row["APS"]
                    if not pd.isna(aps_val) and str(aps_val).strip() != "":
                        new_aps_float = float(str(aps_val).replace(',', '.'))
            except ValueError:
                errors.append(f"Ligne {index+2}: Valeur d'effectif ou APS invalide")
                continue
            centre_key = norm(centre_name_raw)
            centre_id = all_centres.get(centre_key)
            
            poste_id = None
            poste_code = None
            
            # Identification par libellé
            poste_key = norm(poste_label_raw)
            p_info = all_postes_by_label.get(poste_key)
            if p_info:
                poste_id, poste_code = p_info

            if centre_id and poste_id:
                combo_key = (centre_id, poste_id)
                if combo_key in processed_combinations:
                    continue
                processed_combinations.add(combo_key)
                affected_center_ids.add(centre_id)
                successfully_processed_keys.add(combo_key)
            
            # Données de base de la ligne pour le fichier de rejets
            row_context = {
                "Région": row.get("Région", ""),
                "Centre": centre_name_raw,
                "Poste": poste_label_raw,
                "Statutaires": new_val_float,
                "APS": new_aps_float
            }
            
            if not centre_id:
                row_context["Erreur"] = f"Centre '{centre_name_raw}' non trouvé"
                errors.append(row_context)
                continue
            if not poste_id:
                row_context["Erreur"] = f"Poste '{poste_label_raw}' non trouvé dans le référentiel"
                errors.append(row_context)
                continue
            
            # Recherche de l'association existante
            # On cherche par poste_id ou par code_resp pour être résilient
            cp = db.query(CentrePoste).filter(
                CentrePoste.centre_id == centre_id,
                (CentrePoste.poste_id == poste_id) | (CentrePoste.code_resp == poste_code)
            ).first()
            
            if cp:
                cp.effectif_actuel = new_val_float
                if has_aps_col:
                    cp.aps = new_aps_float
                updated_count += 1
            else:
                # Création d'une nouvelle association
                new_cp = CentrePoste(
                    centre_id=centre_id,
                    poste_id=poste_id,
                    code_resp=poste_code,
                    effectif_actuel=new_val_float,
                    aps=new_aps_float if has_aps_col else 0.0
                )
                db.add(new_cp)
                created_count += 1
        
        # Logic to zero-out missing posts for affected centers
        zeroed_count = 0
        if affected_center_ids:
            # Find all existing CentrePoste relations for these centers
            existing_relations = db.query(CentrePoste).filter(CentrePoste.centre_id.in_(list(affected_center_ids))).all()
            for rel in existing_relations:
                key = (rel.centre_id, rel.poste_id)
                if key not in successfully_processed_keys:
                    # This post was in the DB for one of the uploaded centers, but not in the Excel
                    if rel.effectif_actuel != 0 or rel.aps != 0:
                        rel.effectif_actuel = 0
                        rel.aps = 0
                        zeroed_count += 1

        db.commit()
        
        if errors:
            # S'il y a des erreurs, on génère un fichier excel de rejets
            df_errors = pd.DataFrame(errors)
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df_errors.to_excel(writer, index=False, sheet_name='Lignes_Rejetees')
            output.seek(0)
            
            headers = {
                "Content-Disposition": "attachment; filename=rejets_import_effectifs.xlsx",
                "Access-Control-Expose-Headers": "X-Error-Count, X-Updated-Count, X-Created-Count, X-Zeroed-Count",
                "X-Error-Count": str(len(errors)),
                "X-Updated-Count": str(updated_count),
                "X-Created-Count": str(created_count),
                "X-Zeroed-Count": str(zeroed_count)
            }
            return StreamingResponse(
                output,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers=headers
            )
            
        return {
            "status": "success", 
            "message": f"{updated_count} effectifs mis à jour, {created_count} nouveaux postes affectés, {zeroed_count} postes remis à 0 car absents du fichier.",
            "errors": []
        }
    except Exception as e:
        import traceback
        traceback.print_exc()
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))

@router.delete("/effectifs/clear")
def clear_effectifs(
    region_id: Optional[int] = Query(None),
    typologie_id: Optional[int] = Query(None),
    centre_id: Optional[int] = Query(None),
    db: Session = Depends(get_db)
):
    """
    Supprime massivement les effectifs (entrées centre_postes) selon les filtres.
    ATTENTION: Cette opération est destructive et supprimera aussi les tâches associées.
    """
    try:
        if centre_id:
            # Cas d'un centre unique
            db.query(CentrePoste).filter(CentrePoste.centre_id == centre_id).delete(synchronize_session=False)
        else:
            # Cas filtré par région ou typologie
            # On récupère les IDs des centres concernés
            centres_query = db.query(Centre.id)
            if region_id:
                centres_query = centres_query.filter(Centre.region_id == region_id)
            if typologie_id:
                centres_query = centres_query.filter(Centre.categorie_id == typologie_id)
            
            centre_ids = [c[0] for c in centres_query.all()]
            
            if centre_ids:
                db.query(CentrePoste).filter(CentrePoste.centre_id.in_(centre_ids)).delete(synchronize_session=False)
        
        db.commit()
        return {"status": "success", "message": "Les effectifs pour le périmètre sélectionné ont été réinitialisés."}
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))

# ==================== SCHEMAS ====================
class PosteCentreItem(BaseModel):
    """Représente un poste dans un centre avec ses effectifs."""
    centre_poste_id: int
    poste_id: int
    code_poste: Optional[str] = None
    poste_label: str
    type_poste: Optional[str] = "MOD"
    effectif_actuel: float = 0.0
    effectif_aps: Optional[float] = None


class PosteCentreUpdate(BaseModel):
    """Données pour mise à jour d'un poste dans un centre."""
    effectif_actuel: Optional[float] = None
    effectif_aps: Optional[float] = None


class PosteCentreCreate(BaseModel):
    """Données pour créer un nouveau poste dans un centre."""
    poste_id: int
    effectif_actuel: float = 0.0
    effectif_aps: float = 0.0


@router.patch("/centre-postes/{cp_id}")
def update_centre_poste(cp_id: int, update: PosteCentreUpdate, db: Session = Depends(get_db)):
    """
    Met à jour les effectifs d'un poste spécifique.
    """
    cp = db.query(CentrePoste).filter(CentrePoste.id == cp_id).first()
    if not cp:
        raise HTTPException(status_code=404, detail="Association poste-centre non trouvée")
    
    if update.effectif_actuel is not None:
        cp.effectif_actuel = update.effectif_actuel
    if update.effectif_aps is not None:
        cp.aps = update.effectif_aps
    
    db.commit()
    return {"status": "success", "message": "Poste mis à jour"}


# ==================== ENDPOINTS ====================




@router.get("/centres/{centre_id}/postes", response_model=List[PosteCentreItem])
def get_postes_by_centre(
    centre_id: int,
    db: Session = Depends(get_db)
):
    """
    Récupère tous les postes d'un centre avec leurs effectifs.
    """
    centre = db.query(Centre).filter(Centre.id == centre_id).first()
    if not centre:
        raise HTTPException(status_code=404, detail=f"Centre {centre_id} non trouvé")
    
    query = text("""
        SELECT 
            cp.id as centre_poste_id,
            cp.poste_id,
            p.Code as code_poste,
            p.label as poste_label,
            COALESCE(p.type_poste, 'MOD') as type_poste,
            COALESCE(cp.effectif_actuel, 0) as effectif,
            COALESCE(cp.aps, 0) as effectif_aps
        FROM dbo.centre_postes cp
        INNER JOIN dbo.postes p ON p.Code = cp.code_resp
        WHERE cp.centre_id = :centre_id
        ORDER BY p.label
    """)
    
    rows = db.execute(query, {"centre_id": centre_id}).mappings().all()
    
    return [
        PosteCentreItem(
            centre_poste_id=row["centre_poste_id"],
            poste_id=row["poste_id"],
            code_poste=row["code_poste"],
            poste_label=row["poste_label"],
            type_poste=row["type_poste"],
            effectif_actuel=float(row["effectif"] or 0),
            effectif_aps=float(row["effectif_aps"]) if row["effectif_aps"] is not None else None
        )
        for row in rows
    ]


@router.get("/postes/available", response_model=List[dict])
def get_available_postes(
    db: Session = Depends(get_db)
):
    """
    Récupère la liste de tous les postes référencés.
    """
    postes = db.query(Poste).order_by(Poste.label).all()
    
    return [
        {
            "id": p.id,
            "label": p.label,
            "type_poste": p.type_poste,
            "code_poste": p.Code
        }
        for p in postes
    ]

