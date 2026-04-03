import io
import math
import logging
from pathlib import Path
from typing import List, Optional

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, validator
from sqlalchemy import text
from sqlalchemy.orm import Session

from app.core.db import get_db
from app.core.config import settings

logger = logging.getLogger(__name__)

router = APIRouter(
    prefix="/comparaison/actuel-vs-recommande",
    tags=["Comparaison Actuel vs Recommandé"],
    responses={404: {"description": "Not found"}},
)

class ComparaisonInput(BaseModel):
    sacs_jour: int = 50
    dossiers_mois: int = 6500
    productivite_pct: float = 100.0

    @validator("productivite_pct")
    def verifier_productivite(cls, v):
        if v <= 0:
            raise ValueError("Productivité > 0 requise.")
        return v

class ComparaisonRow(BaseModel):
    position: str
    effectif_actuel: float
    recommande: float
    ecart: float

class ComparaisonTotals(BaseModel):
    effectif_actuel: float
    recommande: float
    ecart: float

class ComparaisonResponse(BaseModel):
    dossiers_jour: float
    heures_net: float
    rows: List[ComparaisonRow]
    totaux: ComparaisonTotals

@router.post("/simuler", response_model=ComparaisonResponse)
def simuler_comparaison(payload: ComparaisonInput, db: Session = Depends(get_db)):
    dossiers_jour = payload.dossiers_mois / 22.0
    heures_net = (8.0 * payload.productivite_pct) / 100.0

    IGNORER_METIERS = ["client", "agence", "automatisation", "compta"]
    FIXED_RECO = ["chef service", "chargé saisie", "chargé contrôle", "détaché agence", "chargé réclamation et reporting"]

    try:
        metiers = db.execute(text("SELECT id_metier_recommandee, nom_metier_recomande, effectif_Actuel_rec FROM METIER_recommande")).fetchall()
    except Exception as e:
        logger.error(f"Error fetching metiers: {e}")
        raise HTTPException(status_code=500, detail="Erreur chargement métiers.")

    lignes = []

    for m in metiers:
        nom_metier = m.nom_metier_recomande or ""
        nom_lower = nom_metier.lower().strip()
        
        if nom_lower in IGNORER_METIERS:
            continue

        try:
            taches_rec = db.execute(
                text("SELECT minutes_tache_rec, secondes_tache_rec, unite_rec FROM Tache_rec WHERE id_metier_rec = :id"),
                {"id": m.id_metier_recommandee}
            ).fetchall()
        except Exception:
            taches_rec = []

        heures_reco = 0.0
        for t in taches_rec:
            mins = float(t.minutes_tache_rec or 0)
            secs = float(t.secondes_tache_rec or 0)
            unite = (t.unite_rec or "").strip()
            sec_total = mins * 60 + secs

            if nom_lower == "chargé numérisation":
                qte = 0.0
            else:
                mapping_qte = {"Sac": dossiers_jour, "Demande": dossiers_jour}
                qte = mapping_qte.get(unite, 0.0)

            heures_reco += (sec_total * qte) / 3600.0

        if nom_lower in FIXED_RECO:
            fte_reco_non_arrondi = 1.0
        else:
            if heures_net > 0:
                ratio = heures_reco / heures_net
                fte_reco_non_arrondi = 0.0 if ratio <= 0.05 else ratio
            else:
                fte_reco_non_arrondi = 0.0

        fte_reco = math.ceil(fte_reco_non_arrondi)
        actuel = float(m.effectif_Actuel_rec or 0)
        ecart = round(fte_reco - actuel, 2)

        lignes.append({
            "position": nom_metier,
            "effectif_actuel": actuel,
            "recommande": fte_reco,
            "ecart": ecart
        })

    # Totals
    tot_actuel = sum(r["effectif_actuel"] for r in lignes)
    tot_reco = sum(r["recommande"] for r in lignes)
    tot_ecart = round(sum(r["ecart"] for r in lignes), 2)

    return ComparaisonResponse(
        dossiers_jour=round(dossiers_jour, 1),
        heures_net=round(heures_net, 2),
        rows=[ComparaisonRow(**r) for r in lignes],
        totaux=ComparaisonTotals(
            effectif_actuel=tot_actuel,
            recommande=tot_reco,
            ecart=tot_ecart
        )
    )

import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.cell.cell import MergedCell

def _resolve_logo_filename(name: Optional[str]) -> Optional[str]:
    if not name:
        return None
    candidate = Path(name)
    if not candidate.is_absolute():
        candidate = settings.assets_dir / candidate
    candidate = candidate.resolve(strict=False)
    return str(candidate) if candidate.exists() else None

def _place_logo(ws, path_str: Optional[str], anchor: str, max_width: int, max_height: int):
    if not path_str:
        return
    try:
        logo = Image(path_str)
        scale = min(max_width / logo.width, max_height / logo.height, 1)
        logo.width = int(logo.width * scale)
        logo.height = int(logo.height * scale)
        ws.add_image(logo, anchor)
    except Exception as e:
        logger.error(f"Error placing logo {path_str}: {e}")

@router.get("/export")
def export_comparaison(sacs_jour: int = 50, dossiers_mois: int = 6500, productivite_pct: float = 100.0, db: Session = Depends(get_db)):
    payload = ComparaisonInput(sacs_jour=sacs_jour, dossiers_mois=dossiers_mois, productivite_pct=productivite_pct)
    data = simuler_comparaison(payload, db)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.sheet_view.showGridLines = False
    ws.title = "Comparaison Effectifs"
    
    # 1. Logos & Title (Row 1)
    ws.row_dimensions[1].height = 65
    _place_logo(ws, _resolve_logo_filename(settings.LOGO_BARID), "A1", 160, 60)
    _place_logo(ws, _resolve_logo_filename(settings.LOGO_ALMAV), "D1", 150, 55)

    ws.merge_cells("B1:C1")
    ws["B1"] = "Comparaison Effectif Actuel vs Recommandé"
    ws["B1"].font = Font(bold=True, size=14, color="005EA8")
    ws["B1"].alignment = Alignment(horizontal="center", vertical="center")

    # 2. Simulation Parameters (Row 3-4)
    ws["A3"] = "PARAMÈTRES DE SIMULATION"
    ws["A3"].font = Font(bold=True)
    
    ws["A4"] = f"Sacs/jour : {sacs_jour}"
    ws["B4"] = f"Dossiers/Mois : {dossiers_mois}"
    ws["C4"] = f"Taux Productivité : {productivite_pct}%"
    ws["D4"] = "Temps Mort : 0"
    
    # Outcomes
    ws["A5"] = f"Dossiers/Jour : {data.dossiers_jour}"
    ws["B5"] = f"Heures Travaillées / Jour : {data.heures_net}"

    # 3. Main Data Table (Row 7)
    table_start = 7
    headers = ["Position", "Effectif Actuel", "Recommandé", "Écart Actuel - Reco"]
    
    header_fill = PatternFill("solid", fgColor="005EA8")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col_idx, header_title in enumerate(headers, start=1):
        cell = ws.cell(row=table_start, column=col_idx, value=header_title)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    row_idx = table_start + 1
    for r in data.rows:
        ws.cell(row=row_idx, column=1, value=r.position)
        ws.cell(row=row_idx, column=2, value=r.effectif_actuel)
        ws.cell(row=row_idx, column=3, value=r.recommande)
        ws.cell(row=row_idx, column=4, value=r.ecart)
        row_idx += 1
        
    row_idx += 1 # Empty line
    
    # Total General
    total_fill = PatternFill("solid", fgColor="E8F0FE")
    total_font = Font(bold=True)
    
    cells_data = ["TOTAL GÉNÉRAL", data.totaux.effectif_actuel, data.totaux.recommande, data.totaux.ecart]
    for col_idx, val in enumerate(cells_data, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        cell.font = total_font
        cell.fill = total_fill
        if col_idx > 1:
            cell.alignment = Alignment(horizontal="center")

    # Auto-adjust column width
    for column_cells in ws.columns:
        column_letter = None
        max_length = 0
        for cell in column_cells:
            if isinstance(cell, MergedCell): continue
            if column_letter is None: column_letter = cell.column_letter
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        if column_letter:
            ws.column_dimensions[column_letter].width = max_length + 3

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    
    headers_dict = {
        "Content-Disposition": 'attachment; filename="comparaison_effectifs_recommande.xlsx"',
        "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    }
    return StreamingResponse(buf, headers=headers_dict)

