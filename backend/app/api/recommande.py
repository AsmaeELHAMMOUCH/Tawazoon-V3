import logging
from pathlib import Path
from typing import List, Optional

import pyodbc
from fastapi import APIRouter, Depends, HTTPException
from pydantic import BaseModel, validator
from sqlalchemy import text
from sqlalchemy.exc import DBAPIError
from sqlalchemy.orm import Session

from app.core.config import settings
from app.core.db import get_db

logger = logging.getLogger(__name__)

router = APIRouter(
    prefix="/recommande",
    tags=["Dimensionnement Recommandé"],
    responses={404: {"description": "Not found"}},
)


class MetierItem(BaseModel):
    id: int
    nom: str


class TacheItem(BaseModel):
    id: int
    nom: str
    minutes: float
    secondes: float
    unite: str


class SimulationInput(BaseModel):
    metier_id: int
    nb_sac: int
    nb_dossier_mois: int
    productivite: float

    @validator("productivite")
    def productivite_positive(cls, value: float) -> float:
        if value <= 0:
            raise ValueError("La productivité doit être > 0")
        return value


class TaskResult(BaseModel):
    nom: str
    qte: int
    heures: float


class SimulationResult(BaseModel):
    dossiers_jour: float
    heures_net_jour: float
    donnees_taches: List[TaskResult]
    total_heures: float
    effectif: float
    effectif_arrondi: int


# --- DTOs Globaux ---
class SimulationGlobalInput(BaseModel):
    nb_sac: int = 50
    nb_dossier_mensuel: int = 6500
    productivite: float = 100.0

class GlobalRow(BaseModel):
    nom_metier: str
    heures: float
    fte: float
    fte_arrondi: int

class GlobalTotals(BaseModel):
    total_heures: float
    total_fte: float
    total_fte_arrondi: int

class SimulationGlobalResponse(BaseModel):
    simulation_id: int
    dossiers_jours: float
    heures_net_jour: float
    rows: List[GlobalRow]
    totals: GlobalTotals

class GraphBar(BaseModel):
    nom: str
    minutes: float
    secondes: float
    total_seconds: float
    total_minutes: float
    label_mmss: str
    is_total: bool = False


class GraphResponse(BaseModel):
    bars: List[GraphBar]
    max_minutes: float
    x_limit: float


def _handle_db_error(exc: Exception, context: str = "Erreur base de données"):
    logger.error("Erreur SQL dans /recommande : %s", exc)
    if isinstance(exc, DBAPIError):
        orig = exc.orig
        if isinstance(orig, pyodbc.Error):
            code = str(orig.args[0]) if orig.args else ""
            if code.startswith("28000"):
                raise HTTPException(
                    status_code=401,
                    detail="Authentification SQL Server refusée (code 28000)",
                )
    if isinstance(exc, pyodbc.Error):
        code = str(exc.args[0]) if exc.args else ""
        if code.startswith("28000"):
            raise HTTPException(
                status_code=401,
                detail="Authentification SQL Server refusée (code 28000)",
            )
    raise HTTPException(status_code=500, detail=context)


def _format_mmss(seconds: float) -> str:
    total = max(0, int(round(seconds)))
    minutes = total // 60
    sec = total % 60
    return f"{minutes:02d}:{sec:02d}"


def _resolve_logo(path_value: Optional[str]) -> Optional[str]:
    if not path_value:
        return None
    candidate = Path(path_value)
    if not candidate.exists():
        candidate = settings.assets_dir / candidate.name
    if not candidate.exists():
        return None
    try:
        relative = candidate.relative_to(settings.assets_dir)
    except ValueError:
        relative = Path(candidate.name)
    return f"/assets/{relative.as_posix()}"



@router.get("/logos")
def get_logos():
    return {
        "logo_barid": _resolve_logo(settings.LOGO_BARID),
        "logo_almav": _resolve_logo(settings.LOGO_ALMAV),
    }


@router.get("/metiers", response_model=List[MetierItem])
def get_metiers(db: Session = Depends(get_db)):
    query = text(
        """
        SELECT id_metier_recommandee, nom_metier_recomande
        FROM METIER_recommande
        ORDER BY id_metier_recommandee
        """
    )
    try:
        rows = db.execute(query).fetchall()
    except Exception as exc:
        _handle_db_error(exc, "Impossible de charger la liste des métiers.")

    return [{"id": row[0], "nom": row[1]} for row in rows]


@router.get("/metiers/{metier_id}/taches", response_model=List[TacheItem])
def get_taches(metier_id: int, db: Session = Depends(get_db)):
    query = text(
        """
        SELECT id_tache_rec, nom_tache_rec, minutes_tache_rec, secondes_tache_rec, unite_rec
        FROM Tache_rec
        WHERE id_metier_rec = :metier_id
        ORDER BY id_tache_rec
        """
    )
    try:
        rows = db.execute(query, {"metier_id": metier_id}).fetchall()
    except Exception as exc:
        _handle_db_error(exc, "Impossible de charger les tâches recommandées.")

    return [
        {
            "id": row[0],
            "nom": row[1],
            "minutes": float(row[2] or 0),
            "secondes": float(row[3] or 0),
            "unite": row[4] or "",
        }
        for row in rows
    ]


@router.get("/taches/all")
def get_all_taches_recommande(db: Session = Depends(get_db)):
    query = text(
        """
        SELECT t.id_tache_rec, t.nom_tache_rec, t.minutes_tache_rec, t.secondes_tache_rec, t.unite_rec, m.nom_metier_recomande
        FROM Tache_rec t
        JOIN METIER_recommande m ON t.id_metier_rec = m.id_metier_recommandee
        ORDER BY m.id_metier_recommandee, t.id_tache_rec
        """
    )
    try:
        rows = db.execute(query).fetchall()
    except Exception as exc:
        _handle_db_error(exc, "Impossible de charger toutes les tâches recommandées.")

    return [
        {
            "id": r[0],
            "nom": r[1],
            "minutes": float(r[2] or 0),
            "secondes": float(r[3] or 0),
            "unite": r[4] or "",
            "metierNom": r[5]
        }
        for r in rows
    ]


@router.post("/simuler", response_model=SimulationResult)
def simuler(payload: SimulationInput, db: Session = Depends(get_db)):
    query = text(
        """
        SELECT nom_tache_rec, minutes_tache_rec, secondes_tache_rec, unite_rec
        FROM Tache_rec
        WHERE id_metier_rec = :metier_id
        ORDER BY id_tache_rec
        """
    )
    try:
        rows = db.execute(query, {"metier_id": payload.metier_id}).fetchall()
    except Exception as exc:
        _handle_db_error(exc, "Impossible de charger les tâches pour la simulation.")

    if not rows:
        raise HTTPException(status_code=404, detail="Aucune tâche trouvée pour ce métier.")

    dossiers_jour_exact = payload.nb_dossier_mois / 22.0
    heures_net_jour_exact = (8.0 * payload.productivite) / 100.0

    nb_unites = {
        "Sac": payload.nb_sac,
        "Demande": dossiers_jour_exact,
    }

    tasks_result: List[TaskResult] = []
    total_heures = 0.0

    for row in rows:
        nom = row[0]
        minutes = float(row[1] or 0)
        secondes = float(row[2] or 0)
        unite = row[3] or ""

        temps_unitaire_sec = minutes * 60 + secondes
        qte_base = nb_unites.get(unite, 0)
        qte = int(round(qte_base))

        temps_total_sec = temps_unitaire_sec * qte
        heures = temps_total_sec / 3600.0
        total_heures += heures

        tasks_result.append(
            TaskResult(
                nom=nom,
                qte=qte,
                heures=heures,
            )
        )

    effectif = 0.0
    if heures_net_jour_exact > 0:
        effectif = total_heures / heures_net_jour_exact

    return SimulationResult(
        dossiers_jour=round(dossiers_jour_exact, 2),
        heures_net_jour=round(heures_net_jour_exact, 2),
        donnees_taches=tasks_result,
        total_heures=total_heures,
        effectif=effectif,
        effectif_arrondi=round(effectif),
    )


@router.get("/graphe/{metier_id}", response_model=GraphResponse)
def get_graph_data(metier_id: int, db: Session = Depends(get_db)):
    query = text(
        """
        SELECT nom_tache_rec, minutes_tache_rec, secondes_tache_rec
        FROM Tache_rec
        WHERE id_metier_rec = :metier_id
        ORDER BY id_tache_rec
        """
    )
    try:
        rows = db.execute(query, {"metier_id": metier_id}).fetchall()
    except Exception as exc:
        _handle_db_error(exc, "Impossible de construire les données du graphique.")

    bars: List[GraphBar] = []
    total_seconds = 0.0

    for row in rows:
        nom = row[0] or "Tâche inconnue"
        minutes = float(row[1] or 0)
        secondes = float(row[2] or 0)
        duree_sec = minutes * 60 + secondes
        if duree_sec <= 0:
            continue
        total_seconds += duree_sec
        bars.append(
            GraphBar(
                nom=nom,
                minutes=minutes,
                secondes=secondes,
                total_seconds=duree_sec,
                total_minutes=duree_sec / 60,
                label_mmss=_format_mmss(duree_sec),
            )
        )

    if not bars:
        raise HTTPException(status_code=404, detail="Aucune durée unitaire valide.")

    bars.append(
        GraphBar(
            nom="TOTAL Temps",
            minutes=total_seconds / 60,
            secondes=0.0,
            total_seconds=total_seconds,
            total_minutes=total_seconds / 60,
            label_mmss=_format_mmss(total_seconds),
            is_total=True,
        )
    )

    max_minutes = max(bar.total_minutes for bar in bars)
    x_limit = max_minutes * 1.4 if max_minutes > 0 else 1.0

    return GraphResponse(bars=bars, max_minutes=max_minutes, x_limit=x_limit)


import math
import io
import csv
from fastapi.responses import StreamingResponse

@router.post("/global/simuler", response_model=SimulationGlobalResponse)
def simuler_globale(payload: SimulationGlobalInput, db: Session = Depends(get_db)):
    dossiers_jours = payload.nb_dossier_mensuel / 22.0
    heures_net_jour = (8.0 * payload.productivite) / 100.0
    
    ROLES_A_IGNORER = ["client", "agence", "compta", "automatisation"]
    ROLES_FIXES_FTE_1 = ["chef service", "chargé réclamation et reporting", "chargé contrôle", "détaché agence"]
    
    try:
        metiers = db.execute(text("SELECT id_metier_recommandee, nom_metier_recomande FROM METIER_recommande")).fetchall()
    except Exception as exc:
        _handle_db_error(exc, "Impossible de charger la liste des métiers.")
        
    resultats_bruts = []
    
    for m in metiers:
        nom_metier = m.nom_metier_recomande or ""
        nom_metier_lower = nom_metier.lower().strip()
        if nom_metier_lower in ROLES_A_IGNORER:
            continue
            
        taches = db.execute(text("SELECT minutes_tache_rec, secondes_tache_rec, unite_rec FROM Tache_rec WHERE id_metier_rec = :id"), {"id": m.id_metier_recommandee}).fetchall()
        
        total_heures_metier = 0.0
        
        for t in taches:
            mins = float(t.minutes_tache_rec or 0)
            secs = float(t.secondes_tache_rec or 0)
            unite = (t.unite_rec or "").strip()
            
            duree_minute = mins + (secs / 60.0)
            
            volume = dossiers_jours if unite in ["Sac", "Demande"] else 0
            temps_en_heures = (volume * duree_minute) / 60.0
            total_heures_metier += temps_en_heures
            
        if nom_metier_lower in ROLES_FIXES_FTE_1:
            fte = 1.0
        else:
            if heures_net_jour > 0:
                ratio = total_heures_metier / heures_net_jour
                fte = math.ceil(round(ratio, 2) * 100) / 100.0
                if fte <= 0.05:
                    fte = 0.0
            else:
                fte = 0.0
                
        fte_arrondi = round(fte)
        
        resultats_bruts.append({
            "nom_metier": nom_metier,
            "heures": round(total_heures_metier, 2),
            "fte": fte,
            "fte_arrondi": fte_arrondi
        })
        
    # FUSION
    fusion_dict = {}
    for r in resultats_bruts:
        nom = r["nom_metier"]
        if nom not in fusion_dict:
            fusion_dict[nom] = {"heures": 0.0, "fte": 0.0}
        fusion_dict[nom]["heures"] += r["heures"]
        fusion_dict[nom]["fte"] = max(fusion_dict[nom]["fte"], r["fte"])

    lignes_finales = []
    total_heures = 0.0
    total_fte = 0.0
    total_fte_arrondi = 0
    
    for nom, data in fusion_dict.items():
        h = round(data["heures"], 2)
        f = data["fte"]
        f_arr = round(f)
        
        lignes_finales.append({"nom_metier": nom, "heures": h, "fte": f, "fte_arrondi": f_arr})
        
        total_heures += h
        total_fte += f
        total_fte_arrondi += f_arr
        
    try:
        # DB Insert
        cursor = db.execute(text("INSERT INTO SimulationResult (date_simulation) OUTPUT INSERTED.id_simulation VALUES (GETDATE())"))
        sim_id_row = cursor.fetchone()
        sim_id = sim_id_row[0] if sim_id_row else 0
        
        if sim_id:
            for row in lignes_finales:
                db.execute(text("""
                    INSERT INTO SimulationDetail (id_simulation, nom_metier, heures, fte, fte_arrondi)
                    VALUES (:sim, :nom, :h, :f, :f_arr)
                """), {
                    "sim": sim_id, "nom": row["nom_metier"], "h": row["heures"], "f": row["fte"], "f_arr": row["fte_arrondi"]
                })
            db.commit()
    except Exception as exc:
        db.rollback()
        logger.error(f"Error DB insert global: {exc}")
        sim_id = 0

    return SimulationGlobalResponse(
        simulation_id=sim_id,
        dossiers_jours=round(dossiers_jours, 2),
        heures_net_jour=round(heures_net_jour, 2),
        rows=[GlobalRow(**r) for r in lignes_finales],
        totals=GlobalTotals(total_heures=total_heures, total_fte=total_fte, total_fte_arrondi=total_fte_arrondi)
    )

@router.get("/global/{id}/export-csv")
def export_global_csv(id: int, db: Session = Depends(get_db)):
    rows = db.execute(text("""
        SELECT nom_metier, heures, fte, fte_arrondi 
        FROM SimulationDetail 
        WHERE id_simulation = :id
    """), {"id": id}).fetchall()
    
    if not rows:
        raise HTTPException(status_code=404, detail="Simulation not found")
        
    total_h = sum(r[1] or 0 for r in rows)
    total_fte = sum(r[2] or 0 for r in rows)
    total_fte_arr = sum(r[3] or 0 for r in rows)
    
    output = io.StringIO()
    writer = csv.writer(output, delimiter=';')
    
    writer.writerow(["Position", "Heures", "Besoin Effectifs", "Besoin Effectifs Arrondi"])
    for r in rows:
        writer.writerow([r[0], f"{r[1]:.2f}", f"{r[2]:.2f}", r[3]])
        
    writer.writerow([])
    writer.writerow([f"Total heures nécessaires (Activités/jour)", f"{total_h:.2f}", "", ""])
    writer.writerow([f"Effectif nécessaire", "", f"{total_fte:.2f}", ""])
    writer.writerow([f"Effectif nécessaire Arrondi", "", "", int(total_fte_arr)])
    
    response = StreamingResponse(iter([output.getvalue().encode('utf-8-sig')]), media_type="text/csv")
    response.headers["Content-Disposition"] = f"attachment; filename=simulation_recommandee_globale_{id}.csv"
    return response

@router.get("/global/{id}/graph")
def graph_global(id: int, db: Session = Depends(get_db)):
    rows = db.execute(text("""
        SELECT nom_metier, fte_arrondi, fte 
        FROM SimulationDetail 
        WHERE id_simulation = :id AND fte_arrondi > 0
    """), {"id": id}).fetchall()
    
    if not rows:
        raise HTTPException(status_code=404, detail="Simulation not found or empty")
    
    data = [{"name": r[0], "effectif": float(r[1] or 0)} for r in rows]
    total_exact = sum(float(r[2] or 0) for r in rows)
    data.append({"name": "Total", "effectif": round(total_exact)})
    
    return data

class NormeRecommandeeRow(BaseModel):
    activite: str
    responsable: str
    minutes: str
    secondes: str
    unite: str

@router.get("/normes", response_model=List[NormeRecommandeeRow])
def get_normes_recommandees(db: Session = Depends(get_db)):
    query = text("""
        SELECT t.nom_tache_rec, m.nom_metier_recomande, t.minutes_tache_rec, t.secondes_tache_rec, t.unite_rec
        FROM Tache_rec t
        JOIN METIER_recommande m ON t.id_metier_rec = m.id_metier_recommandee
        WHERE LOWER(m.nom_metier_recomande) NOT IN ('agence', 'chef service')
        ORDER BY t.id_metier_rec
    """)
    try:
        rows = db.execute(query).fetchall()
    except Exception as exc:
        _handle_db_error(exc, "Impossible de charger les normes de dimensionnement recommandées.")

    result = []
    for row in rows:
        activite = row[0] or ""
        responsable = row[1] or ""
        minutes = int(row[2] or 0)
        secondes = float(row[3] or 0)
        
        total_seconds = int(minutes * 60 + secondes)
        minutes_norm = total_seconds // 60
        secondes_norm = total_seconds % 60
        
        result.append(NormeRecommandeeRow(
            activite=activite,
            responsable=responsable,
            minutes=str(minutes_norm),
            secondes=str(int(secondes_norm)),
            unite=row[4] or ""
        ))

    return result

import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.cell.cell import MergedCell

@router.get("/normes/export-xlsx")
def export_normes_recommandees(db: Session = Depends(get_db)):
    normes = get_normes_recommandees(db)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.sheet_view.showGridLines = False
    ws.title = "Normes dimensionnement Reco"
    ws.row_dimensions[1].height = 60
    ws.row_dimensions[2].height = 22

    def _resolve_logo_filename(name: Optional[str]) -> Optional[str]:
        if not name:
            return None
        candidate = Path(name)
        if not candidate.is_absolute():
            candidate = settings.assets_dir / candidate
        candidate = candidate.resolve(strict=False)
        return str(candidate) if candidate.exists() else None

    def _place_logo(path_str: Optional[str], anchor: str, max_width: int, max_height: int):
        if not path_str:
            return
        try:
            logo = Image(path_str)
            scale = min(max_width / logo.width, max_height / logo.height, 1)
            logo.width = int(logo.width * scale)
            logo.height = int(logo.height * scale)
            ws.add_image(logo, anchor)
        except Exception as e:
            logger.error(f"Error placing logo {path_str}: {e}")

    _place_logo(_resolve_logo_filename(settings.LOGO_BARID), "A1", max_width=150, max_height=60)
    _place_logo(_resolve_logo_filename(settings.LOGO_ALMAV), "E1", max_width=140, max_height=55)

    ws.merge_cells("B1:D1")
    ws["B1"] = "Normes de dimensionnement - Recommandé"
    ws["B1"].font = Font(bold=True, size=16, color="005EA8")
    ws["B1"].alignment = Alignment(horizontal="center", vertical="center")
    
    headers = ["Activité", "Responsable", "Minutes", "Secondes", "Unité"]
    table_start = 4
    header_fill = PatternFill("solid", fgColor="005EA8")
    header_font = Font(bold=True, color="FFFFFF")
    header_align = Alignment(horizontal="center", vertical="center")
    
    for col_idx, header_title in enumerate(headers, start=1):
        cell = ws.cell(row=table_start, column=col_idx, value=header_title)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align

    for row_idx, row in enumerate(normes, start=table_start + 1):
        ws.cell(row=row_idx, column=1, value=row.activite)
        ws.cell(row=row_idx, column=2, value=row.responsable)
        ws.cell(row=row_idx, column=3, value=row.minutes)
        ws.cell(row=row_idx, column=4, value=row.secondes)
        ws.cell(row=row_idx, column=5, value=row.unite)

    for column_cells in ws.columns:
        column_letter = None
        max_length = 0
        for cell in column_cells:
            if isinstance(cell, MergedCell):
                continue
            if column_letter is None:
                column_letter = cell.column_letter
            value = cell.value
            if value is None:
                continue
            max_length = max(max_length, len(str(value)))
        if column_letter:
            ws.column_dimensions[column_letter].width = max(max_length + 2, 12)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    
    headers_dict = {
        "Content-Disposition": 'attachment; filename="normes_dimensionnement_recommandees.xlsx"',
        "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    }
    return StreamingResponse(buf, headers=headers_dict)
