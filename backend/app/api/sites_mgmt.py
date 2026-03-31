from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import text
import pandas as pd
import io
import re
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from typing import List, Optional
from pydantic import BaseModel

from app.core.db import get_db
from app.models.db_models import Centre, AttachedSite, Region

router = APIRouter(tags=["Sites Rattachés"])

# ==================== SCHEMAS ====================
class AttachedSiteItem(BaseModel):
    id: int
    label: str
    code: str
    centre_id: int
    count_taches: Optional[int] = 0

    class Config:
        from_attributes = True

class AttachedSiteCreate(BaseModel):
    label: str
    code: Optional[str] = None
    centre_id: int

class AttachedSiteUpdate(BaseModel):
    label: Optional[str] = None
    code: Optional[str] = None
    centre_id: Optional[int] = None

# ==================== UTILS ====================
def norm(s):
    if not s: return ""
    return re.sub(r'\s+', '', str(s)).lower()

# ==================== ENDPOINTS ====================

@router.get("/sites/centre/{centre_id}", response_model=List[AttachedSiteItem])
def get_sites_by_centre(centre_id: int, db: Session = Depends(get_db)):
    """Lister les sites rattachés à un centre."""
    sites = db.query(AttachedSite).filter(AttachedSite.centre_id == centre_id).all()
    return sites

@router.post("/sites", response_model=AttachedSiteItem)
def create_site(site: AttachedSiteCreate, db: Session = Depends(get_db)):
    """Créer un nouveau site rattaché."""
    if not site.code:
        # Générer un code unique basé sur le label
        base_code = norm(site.label)[:10].upper()
        if not base_code: base_code = "SITE"
        
        # S'assurer de l'unicité
        code = base_code
        counter = 1
        while db.query(AttachedSite).filter(AttachedSite.code == code).first():
            code = f"{base_code}_{counter}"
            counter += 1
        site.code = code
    else:
        # Vérifier l'unicité du code fourni
        existing = db.query(AttachedSite).filter(AttachedSite.code == site.code).first()
        if existing:
            raise HTTPException(status_code=400, detail=f"Le code site '{site.code}' est déjà utilisé.")
    
    new_site = AttachedSite(**site.dict())
    db.add(new_site)
    db.commit()
    db.refresh(new_site)
    return new_site

@router.patch("/sites/{site_id}", response_model=AttachedSiteItem)
def update_site(site_id: int, update: AttachedSiteUpdate, db: Session = Depends(get_db)):
    """Modifier un site rattaché."""
    site = db.query(AttachedSite).filter(AttachedSite.id == site_id).first()
    if not site:
        raise HTTPException(status_code=404, detail="Site non trouvé")
    
    if update.code is not None and update.code != site.code:
        existing = db.query(AttachedSite).filter(AttachedSite.code == update.code).first()
        if existing:
            raise HTTPException(status_code=400, detail=f"Le code site '{update.code}' est déjà utilisé.")
        site.code = update.code
    
    if update.label is not None:
        site.label = update.label
    if update.centre_id is not None:
        site.centre_id = update.centre_id
        
    db.commit()
    db.refresh(site)
    return site

@router.delete("/sites/{site_id}")
def delete_site(site_id: int, db: Session = Depends(get_db)):
    """Supprimer un site rattaché."""
    site = db.query(AttachedSite).filter(AttachedSite.id == site_id).first()
    if not site:
        raise HTTPException(status_code=404, detail="Site non trouvé")
    
    db.delete(site)
    db.commit()
    return {"status": "success", "message": "Site supprimé"}

@router.get("/sites/export-template")
def export_sites_template(
    region_id: Optional[int] = Query(None),
    centre_id: Optional[int] = Query(None),
    db: Session = Depends(get_db)
):
    """Générer un template Excel pour l'import des sites."""
    query_str = """
        SELECT r.label as region, c.label as centre, s.label as site_label
        FROM dbo.centres c
        JOIN dbo.regions r ON r.id = c.region_id
        LEFT JOIN dbo.attached_sites s ON s.centre_id = c.id
        WHERE 1=1
    """
    params = {}
    if region_id:
        query_str += " AND c.region_id = :region_id"
        params["region_id"] = region_id
    if centre_id:
        query_str += " AND c.id = :centre_id"
        params["centre_id"] = centre_id
        
    query_str += " ORDER BY r.label, c.label, s.label"
    
    rows = db.execute(text(query_str), params).mappings().all()
    
    data = []
    # Ajouter deux lignes d'exemple réelles pour montrer la structure
    data.append({
        "Région": "AGADIR (EXEMPLE)",
        "Centre": "AGADIR CENTRE DE TRAITEMENT ET DISTRIBUTION (EXEMPLE)",
        "Nom Site": "SITE AGADIR 1 (EXEMPLE)"
    })
    data.append({
        "Région": "AGADIR (EXEMPLE)",
        "Centre": "AGADIR CENTRE DE TRAITEMENT ET DISTRIBUTION (EXEMPLE)",
        "Nom Site": "SITE AGADIR 2 (EXEMPLE)"
    })
    
    for row in rows:
        data.append({
            "Région": row["region"],
            "Centre": row["centre"],
            "Nom Site": row["site_label"] if row["site_label"] else ""
        })
    
    if len(data) == 1: # Seulement l'exemple
        data.append({"Région": "", "Centre": "", "Nom Site": ""})
        
    df = pd.DataFrame(data)
    output = io.BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # On commence à écrire à la ligne 1 (0-indexed) pour laisser place à la note à la ligne 0
        df.to_excel(writer, index=False, sheet_name='Template_Sites', startrow=1)
        workbook = writer.book
        worksheet = writer.sheets['Template_Sites']
        
        # Ajouter l'instruction sur la première ligne
        worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))
        instruction_cell = worksheet.cell(row=1, column=1)
        instruction_cell.value = "💡 NOTE : Pour ajouter plusieurs sites à un même centre : dupliquez la ligne du centre et saisissez le nom du nouveau site."
        instruction_cell.font = Font(italic=True, color="444444", size=10, bold=True)
        instruction_cell.alignment = Alignment(horizontal="left", vertical="center")
        worksheet.row_dimensions[1].height = 25
        
        # Styles pour les en-têtes (maintenant à la ligne 2)
        header_fill = PatternFill(start_color="005EA8", end_color="005EA8", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        example_font = Font(color="808080", italic=True)
        center_align = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin', color='DDDDDD'),
            right=Side(style='thin', color='DDDDDD'),
            top=Side(style='thin', color='DDDDDD'),
            bottom=Side(style='thin', color='DDDDDD')
        )
        
        # Formater l'en-tête (ligne 2)
        for col_num, column_title in enumerate(df.columns, 1):
            cell = worksheet.cell(row=2, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = border
            
            # Largeur automatique
            max_length = max(df[column_title].astype(str).map(len).max(), len(column_title)) + 8
            worksheet.column_dimensions[get_column_letter(col_num)].width = max_length
            
        # Formater les lignes d'exemple (lignes 3 et 4)
        for row_num in [3, 4]:
            for col_num in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.font = example_font
                cell.border = border
                
        # Figer les volets sous les en-têtes
        worksheet.freeze_panes = 'A3'
        
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=template_sites_rattaches.xlsx"}
    )

@router.post("/sites/import")
async def import_sites(file: UploadFile = File(...), db: Session = Depends(get_db)):
    """Importer des sites depuis un fichier Excel (Génération auto du code)."""
    try:
        contents = await file.read()
        # On tente de lire l'en-tête soit à la ligne 0 soit à la ligne 1 (si une note existe)
        df_temp = pd.read_excel(io.BytesIO(contents), nrows=5)
        skip = 0
        if "Centre" not in df_temp.columns:
            # On vérifie si les colonnes sont à la ligne suivante
            df_check = pd.read_excel(io.BytesIO(contents), skiprows=1, nrows=5)
            if "Centre" in df_check.columns:
                skip = 1
        
        df = pd.read_excel(io.BytesIO(contents), skiprows=skip)
        
        required_cols = ["Centre", "Nom Site"]
        for col in required_cols:
            if col not in df.columns:
                raise HTTPException(status_code=400, detail=f"Colonne manquante: {col}")
        
        # Pré-charger les centres
        all_centres = {norm(c.label): c.id for c in db.query(Centre).all()}
        
        # 1. Collecter les données et identifier les centres concernés
        centers_to_clear = set()
        sites_to_insert = [] # Liste de tuples (centre_id, label)
        errors = []
        
        for index, row in df.iterrows():
            centre_name = str(row["Centre"]).strip()
            site_label = str(row["Nom Site"]).strip()
            
            # Ignorer la ligne d'exemple ou les lignes vides
            if "(EXEMPLE)" in centre_name.upper() or not centre_name or centre_name == 'nan':
                continue
            
            centre_id = all_centres.get(norm(centre_name))
            if not centre_id:
                errors.append({"Ligne": index+2, "Centre": centre_name, "Erreur": "Centre non trouvé"})
                continue
                
            centers_to_clear.add(centre_id)
            if site_label and site_label != 'nan':
                sites_to_insert.append((centre_id, site_label))
        
        # 2. Supprimer les sites existants pour les centres trouvés dans le fichier
        if centers_to_clear:
            db.query(AttachedSite).filter(AttachedSite.centre_id.in_(list(centers_to_clear))).delete(synchronize_session=False)
            db.flush() # Appliquer les suppressions pour libérer les codes
            
        # 3. Rafraîchir la liste des codes utilisés après suppression
        used_codes = {s.code for s in db.query(AttachedSite.code).all()}
        
        # 4. Insertion des nouveaux sites
        success_count = 0
        for centre_id, label in sites_to_insert:
            # Génération auto du code
            clean_label = norm(label)
            base_code = f"S{centre_id}_{clean_label[:8].upper()}"
            
            code = base_code
            counter = 1
            while code in used_codes:
                code = f"{base_code}_{counter}"
                counter += 1
            
            new_site = AttachedSite(label=label, code=code, centre_id=centre_id)
            db.add(new_site)
            used_codes.add(code)
            success_count += 1
            
        db.commit()
        
        if errors:
            # Générer rapport de rejets
            df_errors = pd.DataFrame(errors)
            err_output = io.BytesIO()
            with pd.ExcelWriter(err_output, engine='openpyxl') as writer:
                df_errors.to_excel(writer, index=False, sheet_name='Rejets')
            err_output.seek(0)
            
            return StreamingResponse(
                err_output,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={
                    "Content-Disposition": "attachment; filename=rejets_sites.xlsx",
                    "X-Success-Count": str(success_count),
                    "X-Error-Count": str(len(errors))
                }
            )
            
        return {"status": "success", "message": f"{success_count} sites traités avec succès."}
        
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))
