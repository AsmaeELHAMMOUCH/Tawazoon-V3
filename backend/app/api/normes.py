from io import BytesIO
from pathlib import Path
from typing import Dict, List, Optional

import openpyxl
from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.cell.cell import MergedCell
from sqlalchemy.orm import Session

from app.core.config import settings
from app.core.db import get_db
from app.services.normes_dimensionnement import get_normes_dimensionnement

router = APIRouter(prefix="/normes-dimensionnement", tags=["normes"])


@router.get("/")
def read_normes(db: Session = Depends(get_db)):
    rows = get_normes_dimensionnement(db)
    return {"rows": rows}


def _build_normes_workbook(rows: List[Dict[str, str]]) -> BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Normes dimensionnement"
    ws.row_dimensions[1].height = 60
    ws.row_dimensions[2].height = 22

    def _resolve_logo_filename(name: Optional[str]) -> Optional[str]:
        if not name:
            return None
        candidate = Path(name)
        if not candidate.is_absolute():
            candidate = settings.assets_dir / candidate
        candidate = candidate.resolve(strict=False)
        return str(candidate) if candidate.exists() else None

    def _place_logo(path_str: Optional[str], anchor: str, max_width: int, max_height: int):
        if not path_str:
            return
        try:
            logo = Image(path_str)
            scale = min(max_width / logo.width, max_height / logo.height, 1)
            logo.width = int(logo.width * scale)
            logo.height = int(logo.height * scale)
            ws.add_image(logo, anchor)
        except Exception:
            pass

    _place_logo(_resolve_logo_filename(settings.LOGO_BARID), "A1", max_width=150, max_height=60)
    _place_logo(_resolve_logo_filename(settings.LOGO_ALMAV), "F1", max_width=140, max_height=55)

    ws.merge_cells("B1:E1")
    ws["B1"] = "Normes de dimensionnement"
    ws["B1"].font = Font(bold=True, size=16, color="005EA8")
    ws["B1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["B2"] = ""
    ws.merge_cells("B2:E2")
    ws["B2"].alignment = Alignment(horizontal="center", vertical="center")
    headers = ["Activité", "Responsable", "Intitulé RH", "Minutes", "Secondes", "Unité"]
    table_start = 4
    header_fill = PatternFill("solid", fgColor="005EA8")
    header_font = Font(bold=True, color="FFFFFF")
    header_align = Alignment(horizontal="center", vertical="center")
    for col_idx, header_title in enumerate(headers, start=1):
        cell = ws.cell(row=table_start, column=col_idx, value=header_title)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align

    for row_idx, row in enumerate(rows, start=table_start + 1):
        ws.cell(row=row_idx, column=1, value=row.get("activite"))
        ws.cell(row=row_idx, column=2, value=row.get("responsable"))
        ws.cell(row=row_idx, column=3, value=row.get("intitule_rh"))
        ws.cell(row=row_idx, column=4, value=row.get("minutes"))
        ws.cell(row=row_idx, column=5, value=row.get("secondes"))
        ws.cell(row=row_idx, column=6, value=row.get("unite"))

    for column_cells in ws.columns:
        column_letter = None
        max_length = 0
        for cell in column_cells:
            if isinstance(cell, MergedCell):
                continue
            if column_letter is None:
                column_letter = cell.column_letter
            value = cell.value
            if value is None:
                continue
            max_length = max(max_length, len(str(value)))
        if column_letter:
            ws.column_dimensions[column_letter].width = max(max_length + 2, 12)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


@router.get("/export-xlsx")
def export_normes(db: Session = Depends(get_db)):
    rows = get_normes_dimensionnement(db)
    buf = _build_normes_workbook(rows)
    headers = {
        "Content-Disposition": 'attachment; filename="normes_dimensionnement.xlsx"',
        "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    }
    return StreamingResponse(buf, headers=headers)
