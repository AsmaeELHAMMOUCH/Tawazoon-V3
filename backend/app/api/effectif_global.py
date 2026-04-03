from fastapi import APIRouter, Depends, HTTPException, Query, Response
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from io import StringIO
import csv
import logging
import math
import os
import time
from typing import Optional

from app.core.db import get_db
from app.schemas.effectif_global import (
    EffectifGlobalRequest,
    EffectifGlobalResponse,
    EffectifGlobalRow,
    EffectifGlobalTotals,
)
from app.services.effectif_global import compute_effectif_global, insert_simulation

router = APIRouter(prefix="/effectif-global", tags=["effectif-global"])


@router.post("/simuler", response_model=EffectifGlobalResponse)
def simuler_effectif_global(payload: EffectifGlobalRequest, db: Session = Depends(get_db)):
    if payload.productivite_pct <= 0:
        raise HTTPException(status_code=400, detail="Productivité doit être > 0")
    result = compute_effectif_global(
        db=db,
        sacs_jour=payload.sacs_jour,
        dossiers_mois=payload.dossiers_mois,
        productivite_pct=payload.productivite_pct,
    )
    # Persist simulation
    insert_simulation(db, result["rows"], result["dossiers_jour"], result["heures_net_jour"])
    return result


from io import BytesIO
import openpyxl
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.styles import Font, Alignment, PatternFill
from pathlib import Path

from app.core.config import settings

logger = logging.getLogger(__name__)
uvicorn_logger = logging.getLogger("uvicorn.error")

VALID_LOGO_EXT = {".png", ".jpg", ".jpeg"}

def _insert_logo(ws, logo_name: str, anchor: str, max_width: int = 120, max_height: int = 50) -> bool:
    path = _resolve_logo_path(logo_name)
    if not path:
        logger.debug(f"Logo insertion failed: path resolution failed for {logo_name}")
        return False
    try:
        img = OpenpyxlImage(path)
        scale = min(max_width / img.width, max_height / img.height, 1)
        img.width = int(img.width * scale)
        img.height = int(img.height * scale)
        ws.add_image(img, anchor)
        logger.debug(f"Logo inserted successfully: {logo_name} at {anchor}")
        return True
    except Exception as exc:
        logger.error(f"Logo insertion failed: {logo_name} at {anchor} ({exc})")
        return False


def _resolve_logo_path(path_value: str) -> Optional[str]:
    if not path_value:
        logger.debug("Logo path resolved: None (no value provided)")
        return None
    candidate = Path(path_value)
    if not candidate.is_absolute():
        candidate = settings.assets_dir / candidate
    candidate = candidate.resolve(strict=False)
    if not candidate.exists():
        logger.error(f"Logo path resolved: {candidate} (file not found)")
        return None
    if candidate.suffix.lower() not in VALID_LOGO_EXT:
        logger.error(f"Logo path resolved: {candidate} (unsupported format)")
        return None
    resolved = str(candidate)
    logger.debug(f"Logo path resolved: {resolved}")
    return resolved

@router.get("/export")
def export_effectif_global(
    sacs_jour: int = Query(..., ge=0),
    dossiers_mois: int = Query(..., ge=0),
    productivite_pct: float = Query(..., gt=0),
    db: Session = Depends(get_db),
):
    uvicorn_logger.warning(
        "[EXPORT_XLSX_HANDLER] CALLED pid=%s ts=%s",
        os.getpid(),
        time.time(),
    )
    res = compute_effectif_global(db, sacs_jour, dossiers_mois, productivite_pct)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.sheet_view.showGridLines = False
    ws.title = "Effectif Global"

    # Settings and headers styles
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F497D")
    param_font = Font(bold=True)

    # 1. Logos and title header
    ws.row_dimensions[1].height = 60
    ws.row_dimensions[2].height = 18

    _insert_logo(ws, settings.LOGO_BARID, "A1")
    _insert_logo(ws, settings.LOGO_ALMAV, "D1")

    ws["B1"] = "Simulation Globale de Processus Actuel"
    ws["B1"].font = Font(bold=True, size=16, color="1F497D")
    ws["B1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("B1:C1")

   
    ws["B2"].font = Font(size=11, color="1F497D")
    ws["B2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("B2:C2")

    # 2. Simulation Parameters
    start_row = 4
    ws.cell(row=start_row, column=1, value="Paramètres de simulation").font = Font(bold=True, size=12, underline="single")
    ws.merge_cells(f"A{start_row}:D{start_row}")
    
    ws.cell(row=start_row+1, column=1, value="Sacs / Jour").font = param_font
    ws.cell(row=start_row+1, column=2, value=sacs_jour).alignment = Alignment(horizontal="left")
    
    ws.cell(row=start_row+2, column=1, value="Dossiers / Mois").font = param_font
    ws.cell(row=start_row+2, column=2, value=dossiers_mois).alignment = Alignment(horizontal="left")
    
    ws.cell(row=start_row+3, column=1, value="Taux Productivité").font = param_font
    ws.cell(row=start_row+3, column=2, value=f"{productivite_pct}%").alignment = Alignment(horizontal="left")
    
    ws.cell(row=start_row+4, column=1, value="Temps Mort").font = param_font
    ws.cell(row=start_row+4, column=2, value=0).alignment = Alignment(horizontal="left")

    ws.cell(row=start_row+5, column=1, value="Heures Travaillées / Jour").font = param_font
    ws.cell(row=start_row+5, column=2, value=res["heures_net_jour"]).alignment = Alignment(horizontal="left")

    # Render table headers
    table_start = start_row + 7
    headers = ["Position", "Heures", "Besoin Effectifs", "Besoin Effectifs Arrondi"]
    for col_idx, col_name in enumerate(headers, 1):
        cell = ws.cell(row=table_start, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Values
    current_row = table_start + 1
    for r in res["rows"]:
        ws.cell(row=current_row, column=1, value=r["position"])
        ws.cell(row=current_row, column=2, value=round(r["heures"], 2))
        ws.cell(row=current_row, column=3, value=round(r["fte"], 2))
        ws.cell(row=current_row, column=4, value=r["fte_arrondi"])
        current_row += 1

    current_row += 1
    tot = res["totaux"]
    
    # Footer Totals
    ws.cell(row=current_row, column=1, value="Total heures nécessaires (Activités/jour)").font = param_font
    ws.cell(row=current_row, column=2, value=round(tot["total_heures_calculees"], 2))
    current_row += 1
    
    ws.cell(row=current_row, column=1, value=f"Effectif nécessaire (base {tot['net_jrs']:.2f} h/jour)").font = param_font
    ws.cell(row=current_row, column=2, value=round(tot["total_fte"], 2))
    current_row += 1
    
    ws.cell(row=current_row, column=1, value=f"Effectif nécessaire Arrondi").font = param_font
    ws.cell(row=current_row, column=2, value=tot["total_fte_arrondi"]).font = Font(bold=True)
    
    # Adjust column widths
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 22

    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    content = output.getvalue()
    output.close()
    
    return Response(
        content=content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="effectif_global.xlsx"'}
    )

