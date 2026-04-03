from datetime import datetime
import csv
import io

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from app.core.db import get_db
from app.schemas.ratios_productivite import (
    AdequationIndex,
    ExportPayload,
    RatiosSimulationResponse,
    SimulationParams,
)
from app.services.ratios_productivite_service import simulate_ratios_productivite

router = APIRouter(prefix="/ratios-productivite", tags=["ratios-productivite"])


@router.post("/simuler", response_model=RatiosSimulationResponse)
def simulate(payload: SimulationParams, db: Session = Depends(get_db)) -> RatiosSimulationResponse:
    """Lance la simulation principale et retourne toutes les lignes + graphiques"""
    return simulate_ratios_productivite(db, payload)


@router.post("/indice-adequation", response_model=AdequationIndex)
def indice_adequation(payload: SimulationParams, db: Session = Depends(get_db)) -> AdequationIndex:
    """Endpoint complémentaire pour afficher uniquement les indices"""
    result = simulate_ratios_productivite(db, payload)
    return result.adequation


@router.post("/export-csv")
def export_csv(payload: ExportPayload, db: Session = Depends(get_db)) -> StreamingResponse:
    """Génère un CSV dans un format compatible Excel FR"""
    if payload.format != "csv":
        raise HTTPException(status_code=400, detail="Le format CSV est requis pour cet endpoint")

    result = simulate_ratios_productivite(db, payload)
    buffer = io.StringIO()
    writer = csv.writer(buffer, delimiter=";")

    # — En-tête paramètres —
    writer.writerow(["PARAMÈTRES DE SIMULATION"])
    writer.writerow(["Sacs / Jour", payload.sacs_jour])
    writer.writerow(["Dossiers / Mois", payload.dossiers_mois])
    writer.writerow(["Productivité (%)", payload.productivite])
    writer.writerow(["Dossiers / Jour (calculé)", result.calculated_fields.dossiers_par_jour])
    writer.writerow(["Heures Net / Jour", result.calculated_fields.heures_net_par_jour])
    writer.writerow([])

    # — Choix des colonnes selon le scope —
    scope = payload.scope
    if scope == "global":
        writer.writerow([
            "Poste", "Effectif Actuel", "Effectif Calculé", "Effectif Recommandé",
            "Vol Moy/Mois (Actuel)", "Vol Moy/Mois (Calculé)", "Vol Moy/Mois (Recommandé)",
            "Vol Moy/Jour (Actuel)", "Vol Moy/Jour (Calculé)", "Vol Moy/Jour (Recommandé)",
            "Vol Moy/Heure (Actuel)", "Vol Moy/Heure (Calculé)", "Vol Moy/Heure (Recommandé)",
        ])
        for row in result.rows:
            writer.writerow([
                row.position,
                row.effectif_actuel, row.effectif_calcule, row.effectif_recommande,
                row.volume_moyen_mois_actuel, row.volume_moyen_mois_calcule, row.volume_moyen_mois_recommande,
                row.volume_moyen_jour_actuel, row.volume_moyen_jour_calcule, row.volume_moyen_jour_recommande,
                row.volume_moyen_heure_actuel, row.volume_moyen_heure_calcule, row.volume_moyen_heure_recommande,
            ])
    elif scope == "mois":
        writer.writerow(["Poste", "Vol Moy/Mois (Actuel)", "Vol Moy/Mois (Calculé)", "Vol Moy/Mois (Recommandé)"])
        for row in result.rows:
            writer.writerow([row.position, row.volume_moyen_mois_actuel, row.volume_moyen_mois_calcule, row.volume_moyen_mois_recommande])
    elif scope == "jour":
        writer.writerow(["Poste", "Vol Moy/Jour (Actuel)", "Vol Moy/Jour (Calculé)", "Vol Moy/Jour (Recommandé)"])
        for row in result.rows:
            writer.writerow([row.position, row.volume_moyen_jour_actuel, row.volume_moyen_jour_calcule, row.volume_moyen_jour_recommande])
    elif scope == "heure":
        writer.writerow(["Poste", "Vol Moy/Heure (Actuel)", "Vol Moy/Heure (Calculé)", "Vol Moy/Heure (Recommandé)"])
        for row in result.rows:
            writer.writerow([row.position, row.volume_moyen_heure_actuel, row.volume_moyen_heure_calcule, row.volume_moyen_heure_recommande])

    # — Ligne TOTAL —
    writer.writerow([])
    s = result.summary
    if scope == "global":
        writer.writerow([
            "TOTAL",
            s.effectif_actuel, s.effectif_calcule, s.effectif_recommande,
            s.volume_moyen_mois_actuel, s.volume_moyen_mois_calcule, s.volume_moyen_mois_recommande,
            s.volume_moyen_jour_actuel, s.volume_moyen_jour_calcule, s.volume_moyen_jour_recommande,
            s.volume_moyen_heure_actuel, s.volume_moyen_heure_calcule, s.volume_moyen_heure_recommande,
        ])
    elif scope == "mois":
        writer.writerow(["TOTAL", s.volume_moyen_mois_actuel, s.volume_moyen_mois_calcule, s.volume_moyen_mois_recommande])
    elif scope == "jour":
        writer.writerow(["TOTAL", s.volume_moyen_jour_actuel, s.volume_moyen_jour_calcule, s.volume_moyen_jour_recommande])
    elif scope == "heure":
        writer.writerow(["TOTAL", s.volume_moyen_heure_actuel, s.volume_moyen_heure_calcule, s.volume_moyen_heure_recommande])

    buffer.seek(0)
    filename = f"ratios_productivite_{scope}_{datetime.utcnow():%Y%m%dT%H%M%SZ}.csv"
    return StreamingResponse(
        buffer,
        media_type="text/csv; charset=utf-8-sig",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.post("/export-excel")
def export_excel(payload: ExportPayload, db: Session = Depends(get_db)) -> StreamingResponse:
    """Génère un fichier Excel formaté avec en-tête, paramètres et tableau par scope"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import (
            Alignment,
            Border,
            Font,
            PatternFill,
            Side,
        )
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl non disponible sur le serveur")

    result = simulate_ratios_productivite(db, payload)
    scope = payload.scope

    wb = Workbook()
    ws = wb.active
    ws.title = f"Ratios Productivité – {scope.capitalize()}"

    # — Styles —
    BLUE = "005EA8"
    DARK_BLUE = "072956"
    LIGHT_BLUE = "EBF3FB"
    GREY = "F5F5F5"
    WHITE = "FFFFFF"

    h1_font = Font(name="Calibri", size=16, bold=True, color=DARK_BLUE)
    h2_font = Font(name="Calibri", size=11, bold=True, color=DARK_BLUE)
    header_font = Font(name="Calibri", size=10, bold=True, color=WHITE)
    body_font = Font(name="Calibri", size=10)
    bold_font = Font(name="Calibri", size=10, bold=True)
    header_fill = PatternFill(start_color=BLUE, end_color=BLUE, fill_type="solid")
    total_fill = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")
    param_fill = PatternFill(start_color=GREY, end_color=GREY, fill_type="solid")
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    right = Alignment(horizontal="right", vertical="center")
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def _apply(cell, font=None, fill=None, alignment=None, b=None):
        if font:    cell.font = font
        if fill:    cell.fill = fill
        if alignment: cell.alignment = alignment
        if b:       cell.border = b

    # — Titre —
    ws.merge_cells("A1:I1")
    ws["A1"] = "RÉSULTATS DE LA SIMULATION — RATIOS DE PRODUCTIVITÉ"
    _apply(ws["A1"], font=h1_font, alignment=center)
    ws.row_dimensions[1].height = 32

    now = datetime.now()
    ws.merge_cells("A2:I2")
    ws["A2"] = f"Exporté le {now.strftime('%d/%m/%Y')} à {now.strftime('%H:%M')} — Portée : {scope.upper()}"
    _apply(ws["A2"], font=Font(name="Calibri", size=9, color="888888"), alignment=center)
    ws.row_dimensions[2].height = 20

    # — Paramètres —
    ws.merge_cells("A4:B4")
    ws["A4"] = "PARAMÈTRES UTILISÉS"
    _apply(ws["A4"], font=h2_font, fill=param_fill, alignment=left)

    params_data = [
        ("Sacs / Jour", payload.sacs_jour),
        ("Dossiers / Mois", payload.dossiers_mois),
        ("Productivité (%)", f"{payload.productivite}%"),
        ("Dossiers / Jour (calculé)", result.calculated_fields.dossiers_par_jour),
        ("Heures Net / Jour", result.calculated_fields.heures_net_par_jour),
    ]
    for i, (k, v) in enumerate(params_data):
        row = 5 + i
        ws.cell(row=row, column=1, value=k).font = bold_font
        ws.cell(row=row, column=2, value=v).font = body_font
        ws.cell(row=row, column=1).fill = param_fill
        ws.cell(row=row, column=2).fill = param_fill

    # — En-têtes du tableau —
    if scope == "global":
        headers = [
            "Poste",
            "Effectif Actuel", "Effectif Calculé", "Effectif Recommandé",
            "Vol/Mois (Actuel)", "Vol/Mois (Calculé)", "Vol/Mois (Recommandé)",
            "Vol/Jour (Actuel)", "Vol/Jour (Calculé)", "Vol/Jour (Recommandé)",
            "Vol/Heure (Actuel)", "Vol/Heure (Calculé)", "Vol/Heure (Recommandé)",
        ]
    elif scope == "mois":
        headers = ["Poste", "Vol Moy/Mois (Actuel)", "Vol Moy/Mois (Calculé)", "Vol Moy/Mois (Recommandé)"]
    elif scope == "jour":
        headers = ["Poste", "Vol Moy/Jour (Actuel)", "Vol Moy/Jour (Calculé)", "Vol Moy/Jour (Recommandé)"]
    else:  # heure
        headers = ["Poste", "Vol Moy/Heure (Actuel)", "Vol Moy/Heure (Calculé)", "Vol Moy/Heure (Recommandé)"]

    header_row = 11
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=h)
        _apply(cell, font=header_font, fill=header_fill, alignment=center, b=border)
    ws.row_dimensions[header_row].height = 24

    # — Données —
    def _get_row_vals(row):
        if scope == "global":
            return [
                row.position,
                row.effectif_actuel, row.effectif_calcule, row.effectif_recommande,
                row.volume_moyen_mois_actuel, row.volume_moyen_mois_calcule, row.volume_moyen_mois_recommande,
                row.volume_moyen_jour_actuel, row.volume_moyen_jour_calcule, row.volume_moyen_jour_recommande,
                row.volume_moyen_heure_actuel, row.volume_moyen_heure_calcule, row.volume_moyen_heure_recommande,
            ]
        elif scope == "mois":
            return [row.position, row.volume_moyen_mois_actuel, row.volume_moyen_mois_calcule, row.volume_moyen_mois_recommande]
        elif scope == "jour":
            return [row.position, row.volume_moyen_jour_actuel, row.volume_moyen_jour_calcule, row.volume_moyen_jour_recommande]
        else:
            return [row.position, row.volume_moyen_heure_actuel, row.volume_moyen_heure_calcule, row.volume_moyen_heure_recommande]

    cur = header_row + 1
    for row in result.rows:
        vals = _get_row_vals(row)
        for col_idx, v in enumerate(vals, 1):
            cell = ws.cell(row=cur, column=col_idx, value=v)
            cell.font = body_font
            cell.border = border
            cell.alignment = left if col_idx == 1 else right
        cur += 1

    # — TOTAL —
    s = result.summary
    total_vals = _get_row_vals(s)
    total_vals[0] = "TOTAL GÉNÉRAL"
    for col_idx, v in enumerate(total_vals, 1):
        cell = ws.cell(row=cur, column=col_idx, value=v)
        _apply(cell, font=bold_font, fill=total_fill, alignment=left if col_idx == 1 else right, b=border)

    # — Largeurs colonnes —
    ws.column_dimensions["A"].width = 34
    for col in range(2, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18

    # — Freeze header —
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    # — Export —
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"ratios_productivite_{scope}_{datetime.utcnow():%Y%m%dT%H%M%SZ}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
