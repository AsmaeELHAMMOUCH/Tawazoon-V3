
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from typing import Optional
from io import BytesIO
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.core.db import get_db
from app.services.simulation_run import get_simulation_history

router = APIRouter(tags=["export"])

@router.get("/export/history/excel")
def export_history_excel(
    centre_id: Optional[int] = None,
    user_id: Optional[int] = None,
    limit: int = 1000, # Limite plus grande pour l'export
    db: Session = Depends(get_db)
):
    # 1. Récupérer les données
    simulations = get_simulation_history(
        db, 
        centre_id=centre_id, 
        user_id=user_id, 
        limit=limit
    )
    
    # 2. Créer le fichier Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Historique Simulations"
    
    # --- CONFIGURATION DU STYLE ---
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="007BFF", end_color="007BFF", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    # --- EN-TÊTE AVEC LOGOS ---
    # On laisse de l'espace pour les logos (lignes 1 à 4)
    ws.row_dimensions[1].height = 60
    
    # Titre du document au centre
    ws.merge_cells('C1:F1')
    cell_title = ws['C1']
    cell_title.value = "HISTORIQUE DES SIMULATIONS"
    cell_title.font = Font(size=16, bold=True, color="005EA8")
    cell_title.alignment = center_align
    
    # Insertion des Logos
    # NOTE: Les chemins doivent être absolus sur le serveur
    try:
        # Logo Barid (Gauche)
        img_barid = Image(r"c:\Users\Aelhammouch\simulateur-rh-V2\frontend\src\assets\BaridLogo.png")
        img_barid.height = 60
        img_barid.width = 150 # Ajuster selon ratio
        ws.add_image(img_barid, 'A1')
        
        # Logo Almav (Droite)
        img_almav = Image(r"c:\Users\Aelhammouch\simulateur-rh-V2\frontend\src\assets\AlmavLogo.png")
        img_almav.height = 50
        img_almav.width = 120 # Ajuster
        ws.add_image(img_almav, 'G1')
    except Exception as e:
        print(f"⚠️ Impossible de charger les logos: {e}")
        # On continue sans logos si erreur
    
    # --- TABLEAU DE DONNÉES (commence ligne 5) ---
    headers = ["ID", "Centre", "Date", "Productivité", "Heures Calc.", "ETP Calculé", "ETP Arrondi", "Commentaire"]
    start_row = 5
    
    # Écriture des en-têtes
    for col_num, header_title in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col_num, value=header_title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
        
    # Écriture des données
    for i, sim in enumerate(simulations, 1):
        row_num = start_row + i
        
        sim_id = sim.get("simulation_id")
        centre = sim.get("centre_label") or f"ID {sim.get('centre_id')}"
        date_str = sim.get("launched_at").strftime("%d/%m/%Y %H:%M") if sim.get("launched_at") else "-"
        prod = f"{sim.get('productivite')}%"
        heures = sim.get("heures_necessaires")
        etp = sim.get("etp_calcule")
        etp_arr = sim.get("etp_arrondi")
        comm = sim.get("commentaire") or ""
        
        row_data = [sim_id, centre, date_str, prod, heures, etp, etp_arr, comm]
        
        for col_num, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")
            if col_num in [1, 4, 7]: # Centrer ID, Prod, ETP Arr
                cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Ajustement largeur colonnes
    col_widths = [10, 30, 20, 15, 15, 15, 15, 50]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
        
    # --- RETOUR DU FICHIER ---
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"Historique_Simulations_{centre_id if centre_id else 'Global'}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@router.get("/export/bandoeng/template")
def generate_bandoeng_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "Modèle Import Volumes"
    
    # Styles
    header_fill = PatternFill(start_color="007BFF", end_color="007BFF", fill_type="solid")
    subheader_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    subheader_font = Font(bold=True)
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    # --- AMANA ---
    ws.merge_cells('B1:G1')
    ws['B1'] = "AMANA DEPOT"
    ws['B1'].fill = header_fill
    ws['B1'].font = header_font
    ws['B1'].alignment = center_align

    ws.merge_cells('H1:M1')
    ws['H1'] = "AMANA REÇU"
    ws['H1'].fill = header_fill
    ws['H1'].font = header_font
    ws['H1'].alignment = center_align

    ws.merge_cells('B2:D2'); ws['B2'] = "GC"; ws['B2'].fill = subheader_fill; ws['B2'].alignment = center_align
    ws.merge_cells('E2:G2'); ws['E2'] = "Particuliers"; ws['E2'].fill = subheader_fill; ws['E2'].alignment = center_align
    ws.merge_cells('H2:J2'); ws['H2'] = "GC"; ws['H2'].fill = subheader_fill; ws['H2'].alignment = center_align
    ws.merge_cells('K2:M2'); ws['K2'] = "Particuliers"; ws['K2'].fill = subheader_fill; ws['K2'].alignment = center_align

    sub_headers = ["Global", "Local", "Axes"]
    for i, col in enumerate(range(2, 14)): # Columns B to M
        cell = ws.cell(row=3, column=col, value=sub_headers[i % 3])
        cell.alignment = center_align
        cell.border = thin_border

    ws.cell(row=4, column=1, value="Volumes Amana").font = Font(bold=True)
    # Placeholder row for user input
    for col in range(2, 14):
        cell = ws.cell(row=4, column=col)
        cell.border = thin_border
        cell.number_format = '#,##0'

    # --- AUTRES FLUX (CR, CO) ---
    start_row = 7
    ws.merge_cells(f'B{start_row}:D{start_row}')
    ws[f'B{start_row}'] = "MED"
    ws[f'B{start_row}'].fill = header_fill
    ws[f'B{start_row}'].font = header_font
    ws[f'B{start_row}'].alignment = center_align

    ws.merge_cells(f'E{start_row}:G{start_row}')
    ws[f'E{start_row}'] = "ARRIVÉ"
    ws[f'E{start_row}'].fill = header_fill
    ws[f'E{start_row}'].font = header_font
    ws[f'E{start_row}'].alignment = center_align

    for i, col in enumerate(range(2, 8)): # Columns B to G
        cell = ws.cell(row=start_row+1, column=col, value=sub_headers[i % 3])
        cell.alignment = center_align
        cell.border = thin_border

    rows = ["CR", "CO"]
    for i, label in enumerate(rows):
        r = start_row + 2 + i
        ws.cell(row=r, column=1, value=label).font = Font(bold=True)
        for col in range(2, 8):
            cell = ws.cell(row=r, column=col)
            cell.border = thin_border
            cell.number_format = '#,##0'
    
    # --- EL BARKIA ET LRH (structure simplifiée: MED et Arrivé uniquement) ---
    simple_start_row = start_row + 2 + len(rows) + 1  # After CR and CO
    ws.cell(row=simple_start_row, column=2, value="MED").fill = subheader_fill
    ws.cell(row=simple_start_row, column=2).alignment = center_align
    ws.cell(row=simple_start_row, column=2).border = thin_border
    
    ws.cell(row=simple_start_row, column=3, value="ARRIVÉ").fill = subheader_fill
    ws.cell(row=simple_start_row, column=3).alignment = center_align
    ws.cell(row=simple_start_row, column=3).border = thin_border
    
    simple_rows = ["El Barkia", "LRH"]
    for i, label in enumerate(simple_rows):
        r = simple_start_row + 1 + i
        ws.cell(row=r, column=1, value=label).font = Font(bold=True)
        for col in range(2, 4):  # Only MED and Arrivé columns
            cell = ws.cell(row=r, column=col)
            cell.border = thin_border
            cell.number_format = '#,##0'

    # Instructions
    ws['A15'] = "Instructions:"
    ws['A15'].font = Font(bold=True)
    ws['A16'] = "1. Remplissez les cases blanches avec les volumes."
    ws['A17'] = "2. Ne modifiez pas la structure du fichier."
    ws['A18'] = "3. Une fois rempli, importez ce fichier dans l'application."

    # Column widths
    ws.column_dimensions['A'].width = 20
    for col in range(2, 14):
        ws.column_dimensions[get_column_letter(col)].width = 10

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = "Modele_Import_Volumes_Bandoeng.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@router.get("/export/bandoeng/tasks-template")
def generate_bandoeng_tasks_template():
    """
    Génère un modèle Excel pour l'importation des tâches Bandoeng.
    Colonnes: Nom de tâche, Produit, Famille, Unité de mesure, Responsable 1, Responsable 2, Temps_min, Temps_sec
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Modèle Import Tâches"
    
    # Styles
    header_fill = PatternFill(start_color="005EA8", end_color="005EA8", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    thin_border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )
    
    # --- TITRE ---
    ws.merge_cells('A1:H1')
    title_cell = ws['A1']
    title_cell.value = "MODÈLE D'IMPORTATION DES TÂCHES - BANDOENG"
    title_cell.font = Font(size=14, bold=True, color="005EA8")
    title_cell.alignment = center_align
    ws.row_dimensions[1].height = 25
    
    # --- EN-TÊTES ---
    headers = [
        "Nom de tâche",
        "Produit", 
        "Famille",
        "Unité de mesure",
        "Responsable 1",
        "Responsable 2",
        "Temps_min",
        "Temps_sec"
    ]
    
    for col_num, header_title in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_num, value=header_title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
    
    # --- LIGNES D'EXEMPLE (optionnel) ---
    example_data = [
        ["Tri courrier ordinaire", "CO", "Tri", "Sac", "Trieur", "Chef d'équipe", "5", "30"],
        ["Réception colis", "Colis", "Réception", "Colis", "Magasinier", "Responsable logistique", "2", "15"],
    ]
    
    for row_idx, row_data in enumerate(example_data, 4):
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_num, value=value)
            cell.border = thin_border
            if col_num <= 6:
                cell.alignment = left_align
            else:
                cell.alignment = center_align
    
    # --- INSTRUCTIONS ---
    instructions_start = 8
    ws[f'A{instructions_start}'] = "INSTRUCTIONS :"
    ws[f'A{instructions_start}'].font = Font(bold=True, size=11, color="D32F2F")
    
    instructions = [
        "1. Remplissez les colonnes avec les informations des tâches à mettre à jour.",
        "2. Les colonnes 'Nom de tâche', 'Produit', 'Famille' et 'Unité de mesure' servent à identifier la tâche dans la base.",
        "3. Les colonnes 'Responsable 1', 'Responsable 2', 'Temps_min' et 'Temps_sec' seront mises à jour.",
        "4. Temps_min = minutes, Temps_sec = secondes (ex: 5 min 30 sec).",
        "5. Ne modifiez pas les en-têtes de colonnes.",
        "6. Supprimez les lignes d'exemple avant l'importation.",
        "7. Une fois rempli, importez ce fichier dans l'application Bandoeng."
    ]
    
    for i, instruction in enumerate(instructions, instructions_start + 1):
        ws[f'A{i}'] = instruction
        ws[f'A{i}'].font = Font(size=10)
    
    # --- LARGEURS DE COLONNES ---
    col_widths = [35, 15, 20, 20, 25, 25, 12, 12]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    # --- RETOUR DU FICHIER ---
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = "Modele_Import_Taches_Bandoeng.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

