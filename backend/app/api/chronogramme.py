from io import BytesIO
from pathlib import Path
from typing import Optional
import time

import openpyxl
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.cell.cell import MergedCell
from sqlalchemy.orm import Session

from app.core.config import settings
from app.core.db import get_db
from app.services.chronogramme import (
    get_chronogramme_taches,
    generate_csv_taches,
    task_columns,
    get_chronogramme_positions,
)

router = APIRouter(prefix="/chronogramme", tags=["chronogramme"])

_CHRONO_CACHE_TTL = 300  # seconds
_chronogramme_cache = {"rows": [], "timestamp": 0.0}

def _update_cache():
    db_session = next(get_db())
    try:
        rows = get_chronogramme_taches(db_session)
    finally:
        db_session.close()
    _chronogramme_cache["rows"] = rows
    _chronogramme_cache["timestamp"] = time.time()

def _get_cached_rows(force_refresh: bool = False, allow_db: bool = True):
    now = time.time()
    needs_refresh = (
        force_refresh
        or not _chronogramme_cache["rows"]
        or now - _chronogramme_cache["timestamp"] > _CHRONO_CACHE_TTL
    )
    if needs_refresh:
        if not allow_db:
            if _chronogramme_cache["rows"]:
                return _chronogramme_cache["rows"]
            raise RuntimeError("Cache vide et rafraichissement interdit")
        try:
            _update_cache()
        except Exception:
            if _chronogramme_cache["rows"]:
                return _chronogramme_cache["rows"]
            raise
    return _chronogramme_cache["rows"]


@router.get("/taches")
def read_chronogramme_taches(db: Session = Depends(get_db)):
    """
    Retourne les lignes et les colonnes du chronogramme liste des taches.
    """
    try:
        rows = _get_cached_rows(force_refresh=True)
        return {"columns": task_columns, "rows": rows}
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc))


def _refresh_cache_response():
    rows = _get_cached_rows(force_refresh=True)
    return {
        "rows": len(rows),
        "timestamp": _chronogramme_cache["timestamp"],
        "status": "cached",
    }


@router.post("/taches/refresh-cache")
def refresh_chronogramme_cache():
    """
    Force la reconstruction du cache des taches chronogramme.
    """
    try:
        return _refresh_cache_response()
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc))


@router.get("/taches/refresh-cache")
def refresh_chronogramme_cache_get():
    """
    Retourne une version rafraichie du cache (GET compatible).
    """
    try:
        return _refresh_cache_response()
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc))
@router.get("/taches/export-csv")
def export_chronogramme_taches_csv(db: Session = Depends(get_db)):
    """
    Retourne l'export CSV du chronogramme taches.
    """
    try:
        rows = get_chronogramme_taches(db)
        csv_payload = generate_csv_taches(rows)
        body = BytesIO(csv_payload.encode("utf-8"))
        headers = {
            "Content-Disposition": "attachment; filename=chronogramme_taches.csv"
        }
        return StreamingResponse(body, media_type="text/csv; charset=utf-8", headers=headers)
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc))


@router.get("/taches/export-xlsx")
def export_chronogramme_taches_xlsx():
    """
    Retourne l'export Excel du chronogramme taches avec logos.
    """
    try:
        rows = _get_cached_rows(force_refresh=False, allow_db=False)
        if not rows:
            raise RuntimeError("Cache chronogramme vide, lancez /taches ou /taches/refresh-cache")
        buf = _build_chronogramme_workbook(rows)
        headers = {
            "Content-Disposition": 'attachment; filename="chronogramme_taches.xlsx"',
            "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        }
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers=headers,
        )
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc))


@router.get("/positions")
def read_chronogramme_positions(db: Session = Depends(get_db)):
    """
    Retourne l'agrégation par position.
    """
    try:
        return get_chronogramme_positions(db)
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc))


def _resolve_logo_path(name: Optional[str]) -> Optional[str]:
    if not name:
        return None
    candidate = Path(name)
    if not candidate.is_absolute():
        candidate = settings.assets_dir / candidate
    candidate = candidate.resolve(strict=False)
    return str(candidate) if candidate.exists() else None


def _place_logo(ws, path_str: Optional[str], anchor: str, max_width: int, max_height: int):
    if not path_str:
        return
    try:
        img = Image(path_str)
        scale = min(max_width / img.width, max_height / img.height, 1)
        img.width = int(img.width * scale)
        img.height = int(img.height * scale)
        ws.add_image(img, anchor)
    except Exception:
        pass


def _get_column_letter_from_cells(column_cells):
    for cell in column_cells:
        if isinstance(cell, MergedCell):
            continue
        if hasattr(cell, "column_letter"):
            return cell.column_letter
    return None


def _build_chronogramme_workbook(rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Chronogramme"
    ws.row_dimensions[1].height = 60
    ws.row_dimensions[2].height = 22

    _place_logo(_resolve_logo_path(settings.LOGO_BARID), "A1", max_width=150, max_height=60)
    _place_logo(_resolve_logo_path(settings.LOGO_ALMAV), "F1", max_width=140, max_height=55)

    ws.merge_cells("B1:E1")
    ws["B1"] = "Chronogramme de Traitement Unitaire"
    ws["B1"].font = Font(bold=True, size=16, color="005EA8")
    ws["B1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["B2"] = ""
    ws.merge_cells("B2:E2")
    ws["B2"].alignment = Alignment(horizontal="center", vertical="center")

    table_start = 4
    header_fill = PatternFill("solid", fgColor="005EA8")
    header_font = Font(bold=True, color="FFFFFF")
    align = Alignment(horizontal="center", vertical="center")

    for idx, column in enumerate(task_columns, start=1):
        cell = ws.cell(row=table_start, column=idx, value=column)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align

    current_row = table_start + 1
    for row in rows:
        ws.cell(row=current_row, column=1, value=row["tache"])
        ws.cell(row=current_row, column=2, value=row["responsable"])
        ws.cell(row=current_row, column=3, value=row["duree_min"])
        ws.cell(row=current_row, column=4, value=row["duree_sec"])
        ws.cell(row=current_row, column=5, value=row["cum_min"])
        ws.cell(row=current_row, column=6, value=row["cum_sec"])
        ws.cell(row=current_row, column=7, value=row["cum_ms"])
        ws.cell(row=current_row, column=8, value=row["cum_heure"])
        current_row += 1

    for column_cells in ws.columns:
        column_letter = _get_column_letter_from_cells(column_cells)
        if not column_letter:
            continue
        max_len = 0
        for cell in column_cells:
            value = cell.value
            if value is None:
                continue
            max_len = max(max_len, len(str(value)))
        ws.column_dimensions[column_letter].width = max(12, max_len + 2)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
