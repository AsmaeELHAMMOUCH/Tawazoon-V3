import io
import logging
from typing import List, Optional, Union
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy import text
from sqlalchemy.orm import Session
from app.core.db import get_db

logger = logging.getLogger(__name__)

router = APIRouter(
    prefix="/recommande/capacite-nominale",
    tags=["Capacité Nominale Recommandée"],
)

class CapaciteRow(BaseModel):
    position: str
    temps_dossier: str
    dossiers_mois: str
    dossiers_jour: str
    dossiers_heure: str

class CapaciteResponse(BaseModel):
    rows: List[CapaciteRow]
    total: CapaciteRow

def fmt(v: Union[float, str, int, None]) -> str:
    if v == "" or v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v)

TEMPS_FIXES = {
    "Chef Service": 0.00,
    "Chargé Réception dossier": 0.00,
    "Chargé dossier": 9.21,
    "Chargé saisie": 1.19,
    "Chargé Validation": 0.00,
    "Chargé production": 1.50,
    "Chargé envoi": 0.00,
    "Chargé archives": 0.17,
    "Chargé Numérisation": 0.17,
    "Chargé Stock": 0.00,
    "Chargé réclamation et reporting": 0.00,
    "Coordinateur Réseau": 4.12,
    "Chargé codes PIN": 0.00,
    "Chargé Contrôle": 0.00,
    "Détaché agence": 0.00,
    "Automatisation": 1.75,
    "Compta": 1.00
}

EXCLUDED_LOWER = [
    'client', 'agence', 'automatisation', 'compta', 'chef service',
    'détaché agence', 'chargé validation', 'chargé stock',
    'chargé réclamation et reporting', 'chargé réception dossier',
    'chargé envoi', 'chargé codes pin', 'chargé contrôle'
]

@router.get("", response_model=CapaciteResponse)
def get_capacite_nominale_recommande(db: Session = Depends(get_db)):
    param_jour = 480
    param_mois = 22
    
    query = text("""
        SELECT nom_metier_recomande
        FROM METIER_recommande
        ORDER BY nom_metier_recomande
    """)
    
    try:
        results = db.execute(query).fetchall()
    except Exception as e:
        logger.error(f"Error fetching metiers: {e}")
        raise HTTPException(status_code=500, detail="Erreur base de données")

    rows_data = []
    total_temps = 0.0

    # We manually filter to match PyQt casing-insensitive logic provided in prompt
    for res in results:
        poste = res[0] or ""
        if poste.lower().strip() in EXCLUDED_LOWER:
            continue
            
        # Case sensitive lookup for exact match in dictionnaire logic
        temps = TEMPS_FIXES.get(poste, 0.0)
        
        jour_val = ""
        heure_val = ""
        mois_val = ""
        
        if temps > 0:
            jour = round(param_jour / temps, 2)
            heure = round(60 / temps, 1)
            mois = round(jour * param_mois)
            
            jour_val = fmt(jour)
            heure_val = fmt(heure)
            mois_val = fmt(mois)
            total_temps += temps

        rows_data.append(CapaciteRow(
            position=poste,
            temps_dossier=fmt(round(temps, 2)),
            dossiers_mois=mois_val,
            dossiers_jour=jour_val,
            dossiers_heure=heure_val
        ))

    total_row = CapaciteRow(
        position="Total",
        temps_dossier=str(round(total_temps, 2)),
        dossiers_mois="",
        dossiers_jour="",
        dossiers_heure=""
    )

    return CapaciteResponse(rows=rows_data, total=total_row)

@router.get("/export")
def export_capacite_nominale_recommande(db: Session = Depends(get_db)):
    data = get_capacite_nominale_recommande(db)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Capacité Nominale Rec"

    # HEADERS ORDER AS PER SCRIPT LOGIC: Position, Temps, Jour, Heure, Mois
    headers = ["Position", "Temps/Dossier", "Dossiers/Jour", "Dossiers/Heure", "Dossiers/Mois"]
    ws.append(headers)

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="005EA8", end_color="005EA8", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")

    for r in data.rows:
        ws.append([r.position, r.temps_dossier, r.dossiers_jour, r.dossiers_heure, r.dossiers_mois])
        
    # Append Total row
    ws.append([data.total.position, data.total.temps_dossier, data.total.dossiers_jour, data.total.dossiers_heure, data.total.dossiers_mois])
    
    # Style last row (Total)
    last_row = ws.max_row
    for col in range(1, 6):
        cell = ws.cell(row=last_row, column=col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    
    filename = "capacite_nominale_recommande.xlsx"
    headers_dict = {
        "Content-Disposition": f'attachment; filename="{filename}"',
        "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    }
    return StreamingResponse(buf, headers=headers_dict)
