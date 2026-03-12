from typing import Optional, List
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Query, Response, Form
from pydantic import BaseModel, Field
from sqlalchemy.orm import Session
from app.core.db import get_db
from app.services.bandoeng_engine import (
    run_bandoeng_simulation,
    BandoengInputVolumes,
    BandoengParameters,
    BandoengSimulationResult,
    BandoengTaskResult
)

from openpyxl import load_workbook, Workbook
import io
from copy import deepcopy

router = APIRouter(prefix="/bandoeng", tags=["Bandoeng Simulation"])

# Constants
BANDOENG_CENTRE_ID = 1942

from app.models.db_models import (
    normalize_ws, sql_normalize_ws, MappingPosteRecommande, 
    TacheExclueOptimisee, Centre, CentrePoste, Poste, Tache, 
    HierarchiePostes, Ville, Categorie
)

# --- Pydantic Models for API ---

class BandoengVolumesIn(BaseModel):
    amana_import: float = 0.0
    amana_export: float = 0.0
    courrier_ordinaire_import: float = 0.0
    courrier_ordinaire_export: float = 0.0
    courrier_recommande_import: float = 0.0
    courrier_recommande_export: float = 0.0
    gare_import: float = 0.0
    gare_export: float = 0.0
    presse_import: float = 0.0
    presse_export: float = 0.0
    grid_values: dict = {}

class BandoengParamsIn(BaseModel):
    model_config = {"populate_by_name": True}

    ed_percent: float = Field(40.0, alias="edPercent")
    colis_amana_par_canva_sac: float = Field(35.0, alias="colisAmanaParCanvaSac")
    nbr_co_sac: float = Field(350.0, alias="nbrCoSac")
    nbr_cr_sac: float = Field(400.0, alias="nbrCrSac")
    coeff_circ: float = Field(0.0, alias="coeffCirc")
    coeff_geo: float = Field(0.0, alias="coeffGeo")
    pct_retour: float = Field(0.0, alias="pctRetour")
    pct_collecte: float = Field(0.0, alias="pctCollecte")
    pct_axes: float = Field(0.0, alias="pctAxes")
    pct_local: float = Field(0.0, alias="pctLocal")
    pct_international: float = Field(0.0, alias="pctInternational")
    pct_national: float = Field(0.0, alias="pctNational")
    pct_marche_ordinaire: float = Field(0.0, alias="pctMarcheOrdinaire")
    pct_vague_master: float = Field(0.0, alias="pctVagueMaster")
    pct_boite_postale: float = Field(0.0, alias="pctBoitePostale")
    pct_crbt: float = Field(50.0, alias="pctCrbt")
    pct_hors_crbt: float = Field(50.0, alias="pctHorsCrbt")
    productivite: float = 100.0
    idle_minutes: float = Field(0.0, alias="idleMinutes")
    shift: int = 1
    duree_trajet: float = Field(0.0, alias="dureeTrajet")
    has_guichet: int = Field(1, alias="hasGuichet")
    cr_par_caisson: float = Field(40.0, alias="crParCaisson")
    pct_mois: Optional[float] = Field(None, alias="pctMois")
    # --- Saisonnalité mensuelle par flux ---
    pct_mois_amana:   Optional[float] = Field(None, alias="pctMoisAmana")
    pct_mois_co:      Optional[float] = Field(None, alias="pctMoisCo")
    pct_mois_cr:      Optional[float] = Field(None, alias="pctMoisCr")
    pct_mois_lrh:     Optional[float] = Field(None, alias="pctMoisLrh")
    pct_mois_ebarkia: Optional[float] = Field(None, alias="pctMoisEbarkia")
    pct_annee: Optional[float] = Field(None, alias="pctAnnee")
    # --- Taux de croissance par flux ---
    amana_pct_annee:   Optional[float] = Field(None, alias="amanaPctAnnee")
    co_pct_annee:      Optional[float] = Field(None, alias="coPctAnnee")
    cr_pct_annee:      Optional[float] = Field(None, alias="crPctAnnee")
    lrh_pct_annee:     Optional[float] = Field(None, alias="lrhPctAnnee")
    ebarkia_pct_annee: Optional[float] = Field(None, alias="ebarkiaPctAnnee")
    
    # --- Flux specific overrides (amana_) ---
    amana_pct_collecte: Optional[float] = Field(None, alias="amana_pctCollecte")
    amana_pct_retour: Optional[float] = Field(None, alias="amana_pctRetour")
    amana_pct_axes_arrivee: Optional[float] = Field(None, alias="amana_pctAxesArrivee")
    amana_pct_axes_depart: Optional[float] = Field(None, alias="amana_pctAxesDepart")
    amana_pct_national: Optional[float] = Field(None, alias="amana_pctNational")
    amana_pct_international: Optional[float] = Field(None, alias="amana_pctInternational")
    amana_pct_marche_ordinaire: Optional[float] = Field(None, alias="amana_pctMarcheOrdinaire")
    amana_pct_crbt: Optional[float] = Field(None, alias="amana_pctCrbt")
    amana_pct_hors_crbt: Optional[float] = Field(None, alias="amana_pctHorsCrbt")

    # --- Flux specific overrides (co_) ---
    co_pct_collecte: Optional[float] = Field(None, alias="co_pctCollecte")
    co_pct_retour: Optional[float] = Field(None, alias="co_pctRetour")
    co_pct_axes_arrivee: Optional[float] = Field(None, alias="co_pctAxesArrivee")
    co_pct_axes_depart: Optional[float] = Field(None, alias="co_pctAxesDepart")
    co_pct_national: Optional[float] = Field(None, alias="co_pctNational")
    co_pct_international: Optional[float] = Field(None, alias="co_pctInternational")
    co_pct_marche_ordinaire: Optional[float] = Field(None, alias="co_pctMarcheOrdinaire")
    co_pct_vague_master: Optional[float] = Field(None, alias="co_pctVagueMaster")
    co_pct_boite_postale: Optional[float] = Field(None, alias="co_pctBoitePostale")

    # --- Flux specific overrides (cr_) ---
    cr_pct_collecte: Optional[float] = Field(None, alias="cr_pctCollecte")
    cr_pct_retour: Optional[float] = Field(None, alias="cr_pctRetour")
    cr_pct_axes_arrivee: Optional[float] = Field(None, alias="cr_pctAxesArrivee")
    cr_pct_axes_depart: Optional[float] = Field(None, alias="cr_pctAxesDepart")
    cr_pct_national: Optional[float] = Field(None, alias="cr_pctNational")
    cr_pct_international: Optional[float] = Field(None, alias="cr_pctInternational")
    cr_pct_marche_ordinaire: Optional[float] = Field(None, alias="cr_pctMarcheOrdinaire")
    cr_pct_vague_master: Optional[float] = Field(None, alias="cr_pctVagueMaster")
    cr_pct_boite_postale: Optional[float] = Field(None, alias="cr_pctBoitePostale")
    cr_pct_crbt: Optional[float] = Field(None, alias="cr_pctCrbt")
    cr_pct_hors_crbt: Optional[float] = Field(None, alias="cr_pctHorsCrbt")

class BandoengSimulateRequest(BaseModel):
    centre_id: int = 1942
    poste_code: Optional[str] = None
    volumes: BandoengVolumesIn
    params: BandoengParamsIn

class BandoengTaskOut(BaseModel):
    task_id: int
    task_name: str
    unite_mesure: str
    produit: str
    moyenne_min: float
    volume_source: str
    volume_annuel: float
    volume_journalier: float
    heures_calculees: float
    formule: str
    famille: str
    responsable: str
    moy_sec: float
    centre_poste_id: int
    phase: Optional[str] = None

class BandoengSimulateResponse(BaseModel):
    tasks: List[BandoengTaskOut]
    total_heures: float
    heures_net_jour: float
    fte_calcule: float
    fte_arrondi: int
    total_ressources_humaines: float
    ressources_par_poste: dict = {}
    ressources_actuelles_par_poste: dict = {}
    grid_values: Optional[dict] = None
    debug_info: dict = {}

@router.post("/simulate", response_model=BandoengSimulateResponse)
def simulate_bandoeng(request: BandoengSimulateRequest, db: Session = Depends(get_db)):
    try:
        # Convert Request to Engine Inputs
        volumes = BandoengInputVolumes(
            amana_import=request.volumes.amana_import,
            amana_export=request.volumes.amana_export,
            courrier_ordinaire_import=request.volumes.courrier_ordinaire_import,
            courrier_ordinaire_export=request.volumes.courrier_ordinaire_export,
            courrier_recommande_import=request.volumes.courrier_recommande_import,
            courrier_recommande_export=request.volumes.courrier_recommande_export,
            gare_import=request.volumes.gare_import,
            gare_export=request.volumes.gare_export,
            presse_import=request.volumes.presse_import,
            presse_export=request.volumes.presse_export,
            grid_values=request.volumes.grid_values
        )
        
        params = BandoengParameters(
            ed_percent=request.params.ed_percent,
            colis_amana_par_canva_sac=request.params.colis_amana_par_canva_sac,
            nbr_co_sac=request.params.nbr_co_sac,
            nbr_cr_sac=request.params.nbr_cr_sac,
            coeff_circ=request.params.coeff_circ,
            coeff_geo=request.params.coeff_geo,
            pct_retour=request.params.pct_retour,
            pct_collecte=request.params.pct_collecte,
            pct_axes=request.params.pct_axes,
            pct_local=request.params.pct_local,
            pct_international=request.params.pct_international,
            pct_national=request.params.pct_national,
            pct_marche_ordinaire=request.params.pct_marche_ordinaire,
            pct_vague_master=request.params.pct_vague_master,
            pct_boite_postale=request.params.pct_boite_postale,
            pct_crbt=request.params.pct_crbt,
            pct_hors_crbt=request.params.pct_hors_crbt,
            productivite=request.params.productivite,
            idle_minutes=request.params.idle_minutes,
            shift=request.params.shift,
            duree_trajet=request.params.duree_trajet if hasattr(request.params, 'duree_trajet') else 0.0,
            has_guichet=request.params.has_guichet,
            cr_par_caisson=request.params.cr_par_caisson,
            pct_mois=request.params.pct_mois,
            # Saisonnalité par flux
            pct_mois_amana=request.params.pct_mois_amana,
            pct_mois_co=request.params.pct_mois_co,
            pct_mois_cr=request.params.pct_mois_cr,
            pct_mois_lrh=request.params.pct_mois_lrh,
            pct_mois_ebarkia=request.params.pct_mois_ebarkia,
            pct_annee=request.params.pct_annee,
            # Taux par flux
            amana_pct_annee=request.params.amana_pct_annee,
            co_pct_annee=request.params.co_pct_annee,
            cr_pct_annee=request.params.cr_pct_annee,
            lrh_pct_annee=request.params.lrh_pct_annee,
            ebarkia_pct_annee=request.params.ebarkia_pct_annee,
            # AMANA
            amana_pct_collecte=request.params.amana_pct_collecte,
            amana_pct_retour=request.params.amana_pct_retour,
            amana_pct_axes_arrivee=request.params.amana_pct_axes_arrivee,
            amana_pct_axes_depart=request.params.amana_pct_axes_depart,
            amana_pct_national=request.params.amana_pct_national,
            amana_pct_international=request.params.amana_pct_international,
            amana_pct_marche_ordinaire=request.params.amana_pct_marche_ordinaire,
            amana_pct_crbt=request.params.amana_pct_crbt,
            amana_pct_hors_crbt=request.params.amana_pct_hors_crbt,
            # CO
            co_pct_collecte=request.params.co_pct_collecte,
            co_pct_retour=request.params.co_pct_retour,
            co_pct_axes_arrivee=request.params.co_pct_axes_arrivee,
            co_pct_axes_depart=request.params.co_pct_axes_depart,
            co_pct_national=request.params.co_pct_national,
            co_pct_international=request.params.co_pct_international,
            co_pct_marche_ordinaire=request.params.co_pct_marche_ordinaire,
            co_pct_vague_master=request.params.co_pct_vague_master,
            co_pct_boite_postale=request.params.co_pct_boite_postale,
            # CR
            cr_pct_collecte=request.params.cr_pct_collecte,
            cr_pct_retour=request.params.cr_pct_retour,
            cr_pct_axes_arrivee=request.params.cr_pct_axes_arrivee,
            cr_pct_axes_depart=request.params.cr_pct_axes_depart,
            cr_pct_national=request.params.cr_pct_national,
            cr_pct_international=request.params.cr_pct_international,
            cr_pct_marche_ordinaire=request.params.cr_pct_marche_ordinaire,
            cr_pct_vague_master=request.params.cr_pct_vague_master,
            cr_pct_boite_postale=request.params.cr_pct_boite_postale,
            cr_pct_crbt=request.params.cr_pct_crbt,
            cr_pct_hors_crbt=request.params.cr_pct_hors_crbt
        )
        
        print(f"DEBUG: simulate_bandoeng received grid_values: {request.volumes.grid_values}")
        result = run_bandoeng_simulation(db, request.centre_id, volumes, params, request.poste_code)
        
        # Convert Engine Result to Response
        tasks_out = [
            BandoengTaskOut(
                task_id=t.task_id,
                task_name=t.task_name,
                unite_mesure=t.unite_mesure,
                produit=t.produit,
                moyenne_min=t.moyenne_min,
                volume_source=t.volume_source,
                volume_annuel=t.volume_annuel,
                volume_journalier=t.volume_journalier,
                heures_calculees=t.heures_calculees,
                formule=t.formule,
                famille=t.famille,
                responsable=t.responsable,
                moy_sec=t.moy_sec,
                centre_poste_id=t.centre_poste_id,
                phase=t.phase
            ) for t in result.tasks
        ]
        
        return BandoengSimulateResponse(
            tasks=tasks_out,
            total_heures=result.total_heures,
            heures_net_jour=result.heures_net_jour,
            fte_calcule=result.fte_calcule,
            fte_arrondi=result.fte_arrondi,
            total_ressources_humaines=result.total_ressources_humaines,
            ressources_par_poste=result.ressources_par_poste, grid_values=result.grid_values
        )
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


# --- NEW: Simplified Endpoint for VueIntervenant ---
class SimplifiedBandoengRequest(BaseModel):
    """
    Structure simplifiée pour VueIntervenant.
    Accepte directement grid_values et parameters au lieu de volumes/params imbriqués.
    """
    centre_id: int = Field(default=1942, description="ID du centre")
    poste_code: Optional[str] = Field(default=None, description="Code du poste (optionnel)")
    grid_values: dict = Field(default_factory=dict, description="Valeurs de la grille Bandoeng")
    parameters: dict = Field(default_factory=dict, description="Paramètres de simulation")
    mode: str = Field(default="actuel", description="Mode de simulation: 'actuel', 'recommande' ou 'optimise'")

@router.post("/simulate-bandoeng", response_model=BandoengSimulateResponse)
def simulate_bandoeng_direct(request: SimplifiedBandoengRequest, db: Session = Depends(get_db)):
    """
    Endpoint simplifié utilisant bandoeng_engine.py directement.
    Conçu pour VueIntervenant avec une structure de requête plus simple.
    """
    try:
        # 1. Construire BandoengInputVolumes avec grid_values
        volumes = BandoengInputVolumes(
            grid_values=request.grid_values,
            # Champs legacy non utilisés avec grid_values
            amana_import=0.0,
            amana_export=0.0,
            courrier_ordinaire_import=0.0,
            courrier_ordinaire_export=0.0,
            courrier_recommande_import=0.0,
            courrier_recommande_export=0.0,
            gare_import=0.0,
            gare_export=0.0,
            presse_import=0.0,
            presse_export=0.0
        )
        
        # 2. Construire BandoengParameters depuis le dict parameters
        p = request.parameters
        params = BandoengParameters(
            ed_percent=p.get('ed_percent', p.get('edPercent', p.get('pct_sac', 60.0))),
            colis_amana_par_canva_sac=p.get('colis_amana_par_canva_sac', p.get('colisAmanaParCanvaSac', 35.0)),
            nbr_co_sac=p.get('nbr_co_sac', 350.0),
            nbr_cr_sac=p.get('nbr_cr_sac', 400.0),
            # Noms Step4 en priorité, puis anciens alias frontend
            coeff_circ=p.get('coeff_circ', p.get('taux_complexite', p.get('tauxComplexite', 1.0))),
            coeff_geo=p.get('coeff_geo', p.get('nature_geo', p.get('natureGeo', 1.0))),
            pct_retour=p.get('pct_retour', 0.0),
            pct_collecte=p.get('pct_collecte', 0.0),
            pct_axes=p.get('pct_axes', p.get('pct_axes_arrivee', 0.0)),
            pct_local=p.get('pct_local', p.get('pct_axes_depart', 0.0)),
            pct_international=p.get('pct_international', 0.0),
            pct_national=p.get('pct_national', 100.0),
            pct_marche_ordinaire=p.get('pct_marche_ordinaire', 0.0),
            productivite=p.get('productivite', 100.0),
            idle_minutes=p.get('idle_minutes', p.get('idleMinutes', 0.0)),
            shift=int(p.get('shift', 1)),
            duree_trajet=float(p.get('duree_trajet', p.get('dureeTrajet', 0.0))),
            has_guichet=int(p.get('has_guichet', p.get('hasGuichet', 1))),
            pct_mois=p.get('pct_mois'),
            pct_vague_master=p.get('pct_vague_master', p.get('pctVagueMaster', 0.0)),
            pct_boite_postale=p.get('pct_boite_postale', p.get('pctBoitePostale', 0.0)),
            pct_crbt=p.get('pct_crbt', p.get('pctCrbt', 50.0)),
            pct_hors_crbt=p.get('pct_hors_crbt', p.get('pctHorsCrbt', 50.0)),
            # Saisonnalité par flux
            pct_mois_amana=p.get('pct_mois_amana'),
            pct_mois_co=p.get('pct_mois_co'),
            pct_mois_cr=p.get('pct_mois_cr'),
            pct_mois_lrh=p.get('pct_mois_lrh'),
            pct_mois_ebarkia=p.get('pct_mois_ebarkia'),
            pct_annee=p.get('pct_annee'),
            cr_par_caisson=p.get('cr_par_caisson', 40.0),
            # Taux par flux
            amana_pct_annee=p.get('amana_pct_annee'),
            co_pct_annee=p.get('co_pct_annee'),
            cr_pct_annee=p.get('cr_pct_annee'),
            lrh_pct_annee=p.get('lrh_pct_annee'),
            ebarkia_pct_annee=p.get('ebarkia_pct_annee'),
            # AMANA
            amana_pct_collecte=p.get('amana_pct_collecte', p.get('amana_pctCollecte')),
            amana_pct_retour=p.get('amana_pct_retour', p.get('amana_pctRetour')),
            amana_pct_axes_arrivee=p.get('amana_pct_axes_arrivee', p.get('amana_pctAxesArrivee')),
            amana_pct_axes_depart=p.get('amana_pct_axes_depart', p.get('amana_pctAxesDepart')),
            amana_pct_national=p.get('amana_pct_national', p.get('amana_pctNational')),
            amana_pct_international=p.get('amana_pct_international', p.get('amana_pctInternational')),
            amana_pct_marche_ordinaire=p.get('amana_pct_marche_ordinaire', p.get('amana_pctMarcheOrdinaire')),
            amana_pct_crbt=p.get('amana_pct_crbt', p.get('amana_pctCrbt')),
            amana_pct_hors_crbt=p.get('amana_pct_hors_crbt', p.get('amana_pctHorsCrbt')),
            # CO
            co_pct_collecte=p.get('co_pct_collecte', p.get('co_pctCollecte')),
            co_pct_retour=p.get('co_pct_retour', p.get('co_pctRetour')),
            co_pct_axes_arrivee=p.get('co_pct_axes_arrivee', p.get('co_pctAxesArrivee')),
            co_pct_axes_depart=p.get('co_pct_axes_depart', p.get('co_pctAxesDepart')),
            co_pct_national=p.get('co_pct_national', p.get('co_pctNational')),
            co_pct_international=p.get('co_pct_international', p.get('co_pctInternational')),
            co_pct_marche_ordinaire=p.get('co_pct_marche_ordinaire', p.get('co_pctMarcheOrdinaire')),
            co_pct_vague_master=p.get('co_pct_vague_master', p.get('co_pctVagueMaster')),
            co_pct_boite_postale=p.get('co_pct_boite_postale', p.get('co_pctBoitePostale')),
            # CR
            cr_pct_collecte=p.get('cr_pct_collecte', p.get('cr_pctCollecte')),
            cr_pct_retour=p.get('cr_pct_retour', p.get('cr_pctRetour')),
            cr_pct_axes_arrivee=p.get('cr_pct_axes_arrivee', p.get('cr_pctAxesArrivee')),
            cr_pct_axes_depart=p.get('cr_pct_axes_depart', p.get('cr_pctAxesDepart')),
            cr_pct_national=p.get('cr_pct_national', p.get('cr_pctNational')),
            cr_pct_international=p.get('cr_pct_international', p.get('cr_pctInternational')),
            cr_pct_marche_ordinaire=p.get('cr_pct_marche_ordinaire', p.get('cr_pctMarcheOrdinaire')),
            cr_pct_vague_master=p.get('cr_pct_vague_master', p.get('cr_pctVagueMaster')),
            cr_pct_boite_postale=p.get('cr_pct_boite_postale', p.get('cr_pctBoitePostale')),
            cr_pct_crbt=p.get('cr_pct_crbt', p.get('cr_pctCrbt')),
            cr_pct_hors_crbt=p.get('cr_pct_hors_crbt', p.get('cr_pctHorsCrbt'))
        )
        
        # 2.5 Charger le mapping des responsables si mode recommande
        role_mapping = None
        if request.mode == "recommande":
            mappings = db.query(MappingPosteRecommande).all()
            # On construit un dictionnaire {source_code: cible_code}
            # En utilisant les relations pour récupérer les Codes
            role_mapping = {}
            for m in mappings:
                if m.poste_source and m.poste_cible:
                    role_mapping[m.poste_source.Code] = m.poste_cible.Code

        print(f"DEBUG: simulate_bandoeng_direct received grid_values: {request.grid_values}")
        
        # 3.6 Gérer les exclusions pour le mode optimisé
        excluded_task_ids = None
        excluded_task_quadruplets = None
        
        if request.mode == "optimise":
            # --- 1. Exclusions par ID (Centre spécifique) ---
            excl_ids_query = db.query(TacheExclueOptimisee.tache_id).filter(
                TacheExclueOptimisee.centre_id == request.centre_id,
                TacheExclueOptimisee.tache_id.isnot(None)
            )
            excluded_task_ids = [r[0] for r in excl_ids_query.all()]
            
            # --- 2. Exclusions par Typologie (Quadruplet) ---
            centre = db.query(Centre).filter(Centre.id == request.centre_id).first()
            if centre and centre.categorie_id:
                excl_quads_query = db.query(
                    TacheExclueOptimisee.nom_tache,
                    TacheExclueOptimisee.produit,
                    TacheExclueOptimisee.famille_uo,
                    TacheExclueOptimisee.unite_mesure
                ).filter(
                    TacheExclueOptimisee.categorie_id == centre.categorie_id,
                    TacheExclueOptimisee.nom_tache.isnot(None)
                )
                
                # Normalisation pour match efficace dans l'engine
                def clean_q(s):
                    if not s: return ""
                    return "".join(str(s).split()).lower()
                
                excluded_task_quadruplets = [
                    (clean_q(q.nom_tache), clean_q(q.produit), clean_q(q.famille_uo), clean_q(q.unite_mesure))
                    for q in excl_quads_query.all()
                ]

        # 3. Appeler run_bandoeng_simulation
        result = run_bandoeng_simulation(
            db=db,
            centre_id=request.centre_id,
            volumes=volumes,
            params=params,
            poste_code=request.poste_code,
            role_mapping=role_mapping,
            excluded_task_ids=excluded_task_ids,
            excluded_task_quadruplets=excluded_task_quadruplets
        )
        print(f"DEBUG: result.total_heures={result.total_heures}, result.total_ressources_humaines={result.total_ressources_humaines}")
        
        # 3.5 Calculer les ressources actuelles agrégées par poste si mode recommande
        ressources_actuelles_par_poste = {}
        # Récupérer les effectifs actuels de tous les postes du centre
        cps = db.query(CentrePoste).filter(CentrePoste.centre_id == request.centre_id).all()
        
        # Maps robustes pour le calcul
        code_to_val = {}    # Code -> effectif actuel brute
        code_to_label = {}  # Code -> Label affiché
        label_to_val = {}   # Label -> Somme des effectifs (initiale)
        
        for cp in cps:
            if cp.poste:
                code = cp.poste.Code
                lbl = cp.poste.label or cp.poste.nom
                val = float(cp.effectif_actuel or 0)
                
                code_to_val[code] = val
                code_to_label[code] = lbl
                label_to_val[lbl] = label_to_val.get(lbl, 0.0) + val

        if request.mode == "recommande" and role_mapping:
            # On travaille sur une copie de la somme par label
            aggregated = dict(label_to_val)
            for s_code, c_code in role_mapping.items():
                s_lbl = code_to_label.get(s_code)
                c_lbl = code_to_label.get(c_code)
                
                if s_lbl and c_lbl and s_lbl != c_lbl:
                    # On transfère l'effectif spécifique au code source
                    val_to_transfer = code_to_val.get(s_code, 0.0)
                    if val_to_transfer > 0:
                        aggregated[c_lbl] = aggregated.get(c_lbl, 0.0) + val_to_transfer
                        aggregated[s_lbl] = aggregated.get(s_lbl, 0.0) - val_to_transfer
            
            ressources_actuelles_par_poste = aggregated
        else:
            ressources_actuelles_par_poste = label_to_val

        # 4. Convertir le résultat
        tasks_out = [
            BandoengTaskOut(
                task_id=t.task_id,
                task_name=t.task_name,
                unite_mesure=t.unite_mesure,
                produit=t.produit,
                moyenne_min=t.moyenne_min,
                volume_source=t.volume_source,
                volume_annuel=t.volume_annuel,
                volume_journalier=t.volume_journalier,
                heures_calculees=t.heures_calculees,
                formule=t.formule,
                famille=t.famille,
                responsable=t.responsable,
                moy_sec=t.moy_sec,
                centre_poste_id=t.centre_poste_id,
                phase=t.phase
            ) for t in result.tasks
        ]
        
        return BandoengSimulateResponse(
            tasks=tasks_out,
            total_heures=result.total_heures,
            heures_net_jour=result.heures_net_jour,
            fte_calcule=result.fte_calcule,
            fte_arrondi=result.fte_arrondi,
            total_ressources_humaines=result.total_ressources_humaines,
            ressources_par_poste=result.ressources_par_poste,
            ressources_actuelles_par_poste=ressources_actuelles_par_poste,
            grid_values=result.grid_values,
            debug_info=result.debug_info
        )
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Erreur simulation Bandoeng: {str(e)}")



# --- New Response Model for Centre Details ---
class BandoengCentreDetailsResponse(BaseModel):
    centre_id: int
    centre_name: str
    aps: float
    moi_global: int
    mod_global: int
    total_global: int
    task_count: int = 0
    # Paramètres Ville
    nature_geo: Optional[float] = 0.0
    taux_complexite: Optional[float] = 0.0
    duree_trajet: Optional[float] = 0.0
    classe_actuelle: Optional[str] = "SANS"
    categorie_label: Optional[str] = None

@router.get("/centre-details/{centre_id}", response_model=BandoengCentreDetailsResponse)
def get_bandoeng_centre_details(centre_id: int, db: Session = Depends(get_db)):
    """
    Récupère les détails d'effectif du centre (APS, MOI, MOD) depuis la base de données.
    """
    try:
        # 1. Récupérer le Centre
        centre = db.query(Centre).filter(Centre.id == centre_id).first()
        if not centre:
            raise HTTPException(status_code=404, detail="Centre introuvable")
        
        # 2. Récupérer les CentrePostes et leurs types
        centre_postes = (
            db.query(CentrePoste, Poste.type_poste, Poste.label, Poste.Code)
            .join(Poste, CentrePoste.code_resp == Poste.Code)
            .filter(CentrePoste.centre_id == centre_id)
            .all()
        )
        
        moi_sum = 0
        mod_sum = 0
        
        for cp, type_poste, poste_label, poste_code in centre_postes:
            eff = cp.effectif_actuel or 0
            # Logique MOI : Type 'MOI', 'INDIRECT' ou 'STRUCTURE'
            t = (type_poste or "").upper()
            is_moi = t in ["MOI", "INDIRECT", "STRUCTURE"]
            
            if is_moi:
                moi_sum += eff
            else:
                mod_sum += eff
        
        # 3. Récupérer les données de la Ville liée
        ville_data = db.query(Ville).filter(Ville.code == centre.code_ville).first()
        
        # 4. Compter les tâches du centre
        task_count = db.query(Tache).join(CentrePoste).filter(CentrePoste.centre_id == centre_id).count()
                
        return BandoengCentreDetailsResponse(
            centre_id=centre.id,
            centre_name=centre.label,
            aps=float(centre.aps or 0),
            moi_global=int(moi_sum),
            mod_global=int(mod_sum),
            total_global=int(moi_sum + mod_sum),
            task_count=task_count,
            nature_geo=float(ville_data.geographie) if ville_data else 0.0,
            taux_complexite=float(ville_data.circulation) if ville_data else 0.0,
            duree_trajet=float(ville_data.trajet) if ville_data else 0.0,
            classe_actuelle=centre.categorisation.label if centre.categorisation else "SANS",
            categorie_label=centre.categorie.label if getattr(centre, "categorie", None) else None
        )

    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


from sqlalchemy import func

@router.get("/postes")
def list_bandoeng_postes(
    centre_id: int = Query(..., description="ID du centre (Bandoeng = 1942)"),
    db: Session = Depends(get_db),
):
    """
    Retourne la liste des postes pour le dropdown Bandoeng.
    Utilise la jointure via code_resp pour garantir la cohérence avec les autres modules si nécessaire.
    """
    try:
        # Initialisation de la requête : Jointure via code_resp
        query = (
            db.query(
                Poste.id,
                CentrePoste.id.label("centre_poste_id"),
                Poste.label,
                Poste.type_poste,
                func.coalesce(CentrePoste.effectif_actuel, 0).label("effectif_actuel"),
                CentrePoste.code_resp.label("code"), # ✅ Use CentrePoste code as the source of truth
                HierarchiePostes.label.label("categorie"), # ✅ AJOUT: Catégorie Hiérarchique
                Poste.hie_poste.label("raw_hie") # ✅ DEBUG: Raw hierarchy code
            )
            .join(CentrePoste, CentrePoste.code_resp == Poste.Code) # ✅ Join via Code as requested
            .outerjoin(HierarchiePostes, Poste.hie_poste == HierarchiePostes.code) # ✅ Join avec la table Hiérarchie via Code
            .filter(CentrePoste.centre_id == centre_id)
            .order_by(Poste.label)
        )
         
        rows = query.all()

        result = []
        if rows:
            print(f"DEBUG BACKEND: First row category = {rows[0].categorie} | Raw Hie = {rows[0].raw_hie}")
            
        for r in rows:
             result.append({
                "id": r.id,
                "centre_poste_id": r.centre_poste_id,
                "label": r.label,
                "type_poste": r.type_poste,
                "effectif_actuel": r.effectif_actuel,
                "code": r.code,
                "code": r.code,
                "categorie": r.categorie, # ✅ Return category to frontend
                "raw_hie": r.raw_hie # ✅ Debug info
            })
            
        return result

    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Erreur lors de la récupération des postes Bandoeng: {str(e)}")

@router.post("/import/volume-grid")

async def import_bandoeng_volumes(file: UploadFile = File(...)):
    try:
        content = await file.read()
        wb = load_workbook(filename=io.BytesIO(content), data_only=True)
        ws = wb.active
        
        grid_values = {
            "amana": {
                "depot": {
                    "gc": {"global": 0, "local": 0, "axes": 0},
                    "part": {"global": 0, "local": 0, "axes": 0}
                },
                "recu": {
                    "gc": {"global": 0, "local": 0, "axes": 0},
                    "part": {"global": 0, "local": 0, "axes": 0}
                }
            },
            "cr": {
                "med": {"global": 0, "local": 0, "axes": 0},
                "arrive": {"global": 0, "local": 0, "axes": 0}
            },
            "co": {
                "med": {"global": 0, "local": 0, "axes": 0},
                "arrive": {"global": 0, "local": 0, "axes": 0}
            },
            "ebarkia": {
                "med": 0,
                "arrive": 0
            },
            "lrh": {
                "med": 0,
                "arrive": 0
            }
        }
        
        def val(row, col):
            v = ws.cell(row=row, column=col).value
            if v is None: return 0
            if isinstance(v, (int, float)): return v
            try:
                # Handle strings with spaces/commas (thousand separators or decimals)
                cleaned = str(v).replace(' ', '').replace('\u00a0', '').replace('\u202f', '').replace(',', '.')
                return float(cleaned)
            except: 
                return 0

        # --- AMANA (Row 4) ---
        r = 4
        # Depot GC
        grid_values["amana"]["depot"]["gc"]["global"] = val(r, 2)
        grid_values["amana"]["depot"]["gc"]["local"] = val(r, 3)
        grid_values["amana"]["depot"]["gc"]["axes"] = val(r, 4)
        # Depot Part
        grid_values["amana"]["depot"]["part"]["global"] = val(r, 5)
        grid_values["amana"]["depot"]["part"]["local"] = val(r, 6)
        grid_values["amana"]["depot"]["part"]["axes"] = val(r, 7)
        # Recu GC
        grid_values["amana"]["recu"]["gc"]["global"] = val(r, 8)
        grid_values["amana"]["recu"]["gc"]["local"] = val(r, 9)
        grid_values["amana"]["recu"]["gc"]["axes"] = val(r, 10)
        # Recu Part
        grid_values["amana"]["recu"]["part"]["global"] = val(r, 11)
        grid_values["amana"]["recu"]["part"]["local"] = val(r, 12)
        grid_values["amana"]["recu"]["part"]["axes"] = val(r, 13)

        # --- CR and CO (Row 9=CR, 10=CO) ---
        rows_map = {9: "cr", 10: "co"}
        
        for r_idx, key in rows_map.items():
            # MED
            grid_values[key]["med"]["global"] = val(r_idx, 2)
            grid_values[key]["med"]["local"] = val(r_idx, 3)
            grid_values[key]["med"]["axes"] = val(r_idx, 4)
            # ARRIVE
            grid_values[key]["arrive"]["global"] = val(r_idx, 5)
            grid_values[key]["arrive"]["local"] = val(r_idx, 6)
            grid_values[key]["arrive"]["axes"] = val(r_idx, 7)
        
        # --- El Barkia (Row 13: simple MED and Arrivé) ---
        grid_values["ebarkia"]["med"] = val(13, 2)
        grid_values["ebarkia"]["arrive"] = val(13, 3)
        
        # --- LRH (Row 14: simple MED and Arrivé) ---
        grid_values["lrh"]["med"] = val(14, 2)
        grid_values["lrh"]["arrive"] = val(14, 3)

        return grid_values

    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=400, detail=f"Erreur lors de la lecture du fichier Excel: {str(e)}")


@router.get("/import-template")
def get_import_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "Modele Import"
    headers = [
        "Nom de tâche", "Produit", "Famille", "Unité de mesure", 
        "Responsable 1", "Responsable 2", "Temps_min", "Temps_sec", "Base de calcul"
    ]
    ws.append(headers)
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return Response(
        content=buffer.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=Modele_Mise_a_jour_Taches.xlsx"}
    )

@router.post("/import/tasks")
async def import_bandoeng_tasks(
    centre_id: int = Query(..., description="ID du centre cible"),
    file: UploadFile = File(...), 
    db: Session = Depends(get_db)
):
    """
    Importe les tâches depuis un fichier Excel et met à jour les responsables et chronos.
    Gère la duplication des tâches si deux responsables sont fournis.
    """
    try:
        content = await file.read()
        wb = load_workbook(filename=io.BytesIO(content), data_only=True)
        ws = wb.active
        
        updated_count = 0
        duplicate_count = 0
        not_found_count = 0
        failed_rows = [] 
        
        # --- Header for error report ---
        error_headers = [
            "Nom de tâche", "Produit", "Famille", "Unité de mesure", 
            "Responsable 1", "Responsable 2", "Temps_min", "Temps_sec", "base de calcul",
            "Raison du rejet"
        ]

        # --- Helper: Resolve CentrePoste ---
        from app.models.db_models import Tache, CentrePoste, Poste
        from sqlalchemy import func, or_
        
        def resolve_centre_poste(session, c_id, resp_name):
            if not resp_name:
                return None, None
                
            l_resp = normalize_ws(resp_name)
            # 1. Trouver le Poste
            poste = session.query(Poste).filter(sql_normalize_ws(Poste.label) == l_resp).first()
            if not poste:
                return None, f"Poste '{resp_name}' non trouvé"
            
            # 2. Trouver CentrePoste via Code
            if not poste.Code:
                return None, f"Poste '{resp_name}' sans code"
                
            cp = (
                session.query(CentrePoste)
                .filter(CentrePoste.centre_id == c_id)
                .filter(CentrePoste.code_resp == poste.Code)
                .first()
            )
            
            if cp:
                return cp.id, None
            else:
                # 3. Créer CentrePoste
                new_cp = CentrePoste(
                    centre_id=c_id,
                    poste_id=poste.id,
                    code_resp=poste.Code,
                    effectif_actuel=0
                )
                session.add(new_cp)
                session.flush()
                return new_cp.id, None

        def parse_num(val, default=0.0):
            if val is None: return default
            try:
                return float(str(val).replace(',', '.'))
            except:
                return default

        # --- Main Loop ---
        for row_idx in range(2, ws.max_row + 1):
            row_vals = [ws.cell(row=row_idx, column=i).value for i in range(1, 10)]
            if not any(row_vals): continue

            nom_tache, produit, famille, unite_mesure = [str(x or "").strip() for x in row_vals[:4]]
            resp1_raw, resp2_raw = [str(x or "").strip() for x in row_vals[4:6]]
            t_min_raw, t_sec_raw, bc_val = row_vals[6:9]
            # base_calcul logic with percentage detection
            cell_bc = ws.cell(row=row_idx, column=9)
            base_calcul = None
            if bc_val is not None:
                # Si Excel a formaté en pourcentage (ex: 40% -> 0.4), on multiplie par 100 pour avoir "40"
                if isinstance(bc_val, (int, float)) and cell_bc.number_format and '%' in str(cell_bc.number_format):
                    base_calcul = str(int(round(bc_val * 100)))
                else:
                    base_calcul = str(bc_val)
            
            if not nom_tache or nom_tache == "None":
                continue
                
            # Parsing Temps
            t_min = parse_num(t_min_raw)
            t_sec = parse_num(t_sec_raw)
                
            # Recherche Tâches
            q = (
                db.query(Tache)
                .join(CentrePoste, Tache.centre_poste_id == CentrePoste.id)
                .filter(CentrePoste.centre_id == centre_id)
                .filter(sql_normalize_ws(Tache.nom_tache) == normalize_ws(nom_tache))
                .filter(sql_normalize_ws(Tache.famille_uo) == normalize_ws(famille))
                .filter(sql_normalize_ws(Tache.unite_mesure) == normalize_ws(unite_mesure))
            )
            
            # Filtre Produit (LIKE)
            if produit:
                 q = q.filter(sql_normalize_ws(Tache.produit).like(f"%{normalize_ws(produit)}%"))
            
            found_tasks = q.all()
            
            if not found_tasks:
                not_found_count += 1
                failed_rows.append(list(row_vals) + ["Tâche non trouvée"])
                continue
                
            # --- Application Règles ---
            cp_id_1, err1 = resolve_centre_poste(db, centre_id, resp1_raw) if resp1_raw else (None, None)
            cp_id_2, err2 = resolve_centre_poste(db, centre_id, resp2_raw) if resp2_raw else (None, None)
            
            if (resp1_raw and not cp_id_1) or (resp2_raw and not cp_id_2):
                reason = err1 or err2
                failed_rows.append(list(row_vals) + [reason])
                continue

            # Cas A : Deux Responsables
            if resp1_raw and resp2_raw:
                if len(found_tasks) == 1:
                    # 1 seule tâche -> Update T1, Dupliquer T2
                    t1 = found_tasks[0]
                    
                    # Update T1
                    if cp_id_1: t1.centre_poste_id = cp_id_1
                    t1.moyenne_min = str(t_min)
                    t1.moy_sec = str(t_sec)
                    if base_calcul is not None: t1.base_calcul = base_calcul
                    updated_count += 1
                    
                    # Create T2 (Duplicate)
                    if cp_id_2:
                        t2 = Tache(
                            centre_poste_id=cp_id_2,
                            nom_tache=t1.nom_tache,
                            famille_uo=t1.famille_uo,
                            phase=t1.phase,
                            unite_mesure=t1.unite_mesure,
                            etat=t1.etat,
                            produit=t1.produit,
                            segment_id=t1.segment_id,
                            moyenne_min=str(t_min),
                            moy_sec=str(t_sec),
                            base_calcul=base_calcul if base_calcul is not None else t1.base_calcul
                        )
                        db.add(t2)
                        duplicate_count += 1
                    else:
                        failed_rows.append(list(row_vals) + ["Resp2 introuvable"])
                
                elif len(found_tasks) >= 2:
                    # 2+ tâches -> Update T1 et T2
                    # Tâche 1 -> Resp 1
                    t1 = found_tasks[0]
                    if cp_id_1: t1.centre_poste_id = cp_id_1
                    t1.moyenne_min = str(t_min)
                    t1.moy_sec = str(t_sec)
                    if base_calcul is not None: t1.base_calcul = base_calcul
                    updated_count += 1
                    
                    # Tâche 2 -> Resp 2
                    t2 = found_tasks[1]
                    if cp_id_2: t2.centre_poste_id = cp_id_2
                    t2.moyenne_min = str(t_min)
                    t2.moy_sec = str(t_sec)
                    if base_calcul is not None: t2.base_calcul = base_calcul
                    updated_count += 1
                    
                    # Les autres -> Supprimer
                    for t_extra in found_tasks[2:]:
                         db.delete(t_extra)

            # Cas B : Un Seul Responsable
            elif resp1_raw and not resp2_raw:
                t1 = found_tasks[0]
                if cp_id_1: t1.centre_poste_id = cp_id_1
                t1.moyenne_min = str(t_min)
                t1.moy_sec = str(t_sec)
                if base_calcul is not None: t1.base_calcul = base_calcul
                updated_count += 1
                
                if len(found_tasks) >= 2:
                    for t_other in found_tasks[1:]:
                        db.delete(t_other)
            
            else:
                 for t in found_tasks:
                     t.moyenne_min = str(t_min)
                     t.moy_sec = str(t_sec)
                     if base_calcul is not None: t.base_calcul = base_calcul
                     updated_count += 1

        db.commit()

        if failed_rows:
            ewb = Workbook()
            ews = ewb.active
            ews.title = "Rejets Mise à jour"
            ews.append(error_headers)
            for fr in failed_rows:
                ews.append(fr)
            
            eout = io.BytesIO()
            ewb.save(eout)
            eout.seek(0)
            
            return Response(
                content=eout.read(),
                headers={
                    'Content-Disposition': 'attachment; filename="rejet_mise_a_jour_taches.xlsx"',
                    'X-Error-Count': str(len(failed_rows)),
                    'X-Updated-Count': str(updated_count)
                },
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        
        return {
            "success": True,
            "updated_count": updated_count,
            "duplicate_count": duplicate_count,
            "not_found_count": not_found_count,
            "failed_count": 0
        }

    except Exception as e:
        db.rollback()
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=400, detail=f"Erreur import: {str(e)}")


@router.get("/new-tasks-template")
def get_new_tasks_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "Modele Nouvelles Taches"
    headers = [
        "Nom de tâche", "Produit", "Famille", "Phase", 
        "Unité de mesure", "base de calcul", 
        "Responsable 1", "Responsable 2", 
        "Temps_min", "Temps_sec"
    ]
    ws.append(headers)
    
    # Optional: adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except: pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width if adjusted_width > 15 else 15

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    headers_resp = {
        'Content-Disposition': 'attachment; filename="modele_nouvelles_taches.xlsx"'
    }
    return Response(content=output.read(), headers=headers_resp, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")



def _process_task_import_workbook(db: Session, centre_id: int, ws):
    from app.models.db_models import Tache, CentrePoste, Poste
    
    created_count = 0
    failed_rows = []
    
    def resolve_cp(session, c_id, resp_name):
        if not resp_name: return None, None
        l_resp = normalize_ws(resp_name)
        # 1. Check in Postes
        poste = session.query(Poste).filter(sql_normalize_ws(Poste.label) == l_resp).first()
        if not poste:
            return None, f"Poste '{resp_name}' non trouvé dans le référentiel"
        
        if not poste.Code:
            return None, f"Le poste '{resp_name}' n'a pas de code associé"
        
        # 2. Check in CentrePoste
        cp = (
            session.query(CentrePoste)
            .filter(CentrePoste.centre_id == c_id)
            .filter(CentrePoste.code_resp == poste.Code)
            .first()
        )
        
        if cp:
            return cp.id, None
        else:
            # 3. Create CentrePoste
            new_cp = CentrePoste(
                centre_id=c_id,
                poste_id=poste.id,
                code_resp=poste.Code,
                effectif_actuel=0
            )
            session.add(new_cp)
            session.flush()
            return new_cp.id, None

    def parse_numeric(val, default="0.0"):
        if val is None: return default
        try:
            # Convertir en float pour nettoyer (gère les formats bizarres d'Excel)
            clean_val = float(str(val).replace(',', '.'))
            return str(clean_val)
        except:
            return default

    # Main loop
    for row_idx in range(2, ws.max_row + 1):
        row_data = [ws.cell(row=row_idx, column=i).value for i in range(1, 11)]
        if not any(row_data): continue # Skip empty rows
        
        nom_tache, produit, famille, phase, unit, base, r1_name, r2_name, t_min, t_sec = row_data
        
        # Resolution Resp 1
        cp_id_1, err1 = resolve_cp(db, centre_id, r1_name)
        
        # Resolution Resp 2
        cp_id_2, err2 = (None, None)
        if r2_name:
            cp_id_2, err2 = resolve_cp(db, centre_id, r2_name)
        
        # Error handling
        if (r1_name and not cp_id_1) or (r2_name and not cp_id_2):
            reason = err1 if (r1_name and not cp_id_1) else err2
            failed_rows.append(list(row_data) + [reason])
            continue
        
        if not cp_id_1 and not cp_id_2:
            failed_rows.append(list(row_data) + ["Aucun responsable valide fourni"])
            continue

        # Insertion
        tasks_to_create = []
        if cp_id_1:
            tasks_to_create.append(cp_id_1)
        if cp_id_2:
            tasks_to_create.append(cp_id_2)
        
        # Parsing numeric values properly
        parsed_min = parse_numeric(t_min)
        parsed_sec = parse_numeric(t_sec)
        
        # Base de calcul with percentage detection
        cell_base = ws.cell(row=row_idx, column=6)
        if base is not None:
            if isinstance(base, (int, float)) and cell_base.number_format and '%' in str(cell_base.number_format):
                parsed_base = str(int(round(base * 100)))
            else:
                parsed_base = parse_numeric(base, "100")
        else:
            parsed_base = "100"

        for cp_id in tasks_to_create:
            new_t = Tache(
                nom_tache=str(nom_tache or ""),
                produit=str(produit or ""),
                famille_uo=str(famille or ""),
                phase=str(phase or ""),
                unite_mesure=str(unit or ""),
                base_calcul=parsed_base,
                moyenne_min=parsed_min,
                moy_sec=parsed_sec,
                centre_poste_id=cp_id,
                etat="ACTIF"
            )
            db.add(new_t)
            created_count += 1
    
    return created_count, failed_rows

@router.post("/import-new-tasks")
async def import_new_tasks(
    centre_id: int = Query(..., description="ID du centre cible"),
    file: UploadFile = File(...), 
    db: Session = Depends(get_db)
):
    import openpyxl
    from openpyxl import Workbook
    
    try:
        content = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        ws = wb.active
        
        created_count, failed_rows = _process_task_import_workbook(db, centre_id, ws)
        
        db.commit()

        # If errors, return Excel file
        if failed_rows:
            error_headers = [
                "Nom de tâche", "Produit", "Famille", "Phase", 
                "Unité de mesure", "base de calcul", 
                "Responsable 1", "Responsable 2", 
                "Temps_min", "Temps_sec", "Raison du rejet"
            ]
            ewb = Workbook()
            ews = ewb.active
            ews.title = "Taches non creees"
            ews.append(error_headers)
            for fr in failed_rows:
                ews.append(fr)
            
            # Style errors
            for cell in ews[1]:
                cell.font = openpyxl.styles.Font(bold=True)
            
            eout = io.BytesIO()
            ewb.save(eout)
            eout.seek(0)
            
            return Response(
                content=eout.read(),
                headers={
                    'Content-Disposition': 'attachment; filename="rejet_import_taches.xlsx"',
                    'X-Error-Count': str(len(failed_rows)),
                    'X-Created-Count': str(created_count)
                },
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        return {
            "success": True,
            "created_count": created_count,
            "failed_count": 0
        }

    except Exception as e:
        db.rollback()
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=400, detail=f"Erreur import: {str(e)}")

@router.post("/auto-import-tasks")
async def auto_import_tasks(
    centre_id: int = Query(..., description="ID du centre cible"),
    db: Session = Depends(get_db)
):
    from app.models.db_models import Centre, Categorie
    import openpyxl
    import os
    
    try:
        # 1. Vérifier si le centre existe et récupérer sa typologie
        centre = db.query(Centre).filter(Centre.id == centre_id).first()
        if not centre:
            raise HTTPException(status_code=404, detail="Centre non trouvé")
        
        # 2. Déterminer le fichier Excel à utiliser
        typology_label = ""
        if centre.categorie:
            typology_label = str(centre.categorie.label).upper()
        
        # Mapping simple basé sur les préfixes ou mots clés
        filename = "Standard.xlsx"
        if "AGENCE MESSAGERIE" in typology_label or typology_label.startswith("AM"):
            filename = "AM.xlsx"
        elif "CENTRE MESSAGERIE" in typology_label or typology_label.startswith("CM"):
            filename = "CM.xlsx"
        elif "CENTRE DE DISTRIBUTION" in typology_label or typology_label.startswith("CD"):
            filename = "CD.xlsx"
        elif "CENTRE COURRIER COLIS" in typology_label or typology_label.startswith("CCC"):
            filename = "CCC.xlsx"
        elif "CELLULE DE DISTRIBUTION" in typology_label or typology_label.startswith("CLD"):
            filename = "CLD.xlsx"
        elif "CENTRE DE TRAITEMENT ET DISTRIBUTION" in typology_label or typology_label.startswith("CTD"):
            filename = "CTD.xlsx"
            
        resources_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), "resources", "typologies")
        file_path = os.path.join(resources_dir, filename)
        
        if not os.path.exists(file_path):
            # Fallback to Standard.xlsx if exists
            file_path = os.path.join(resources_dir, "Standard.xlsx")
            if not os.path.exists(file_path):
                raise HTTPException(status_code=404, detail=f"Fichier de typologie non trouvé pour {typology_label}")

        # 3. Charger le workbook et importer les tâches
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active
        
        created_count, failed_rows = _process_task_import_workbook(db, centre_id, ws)
        
        db.commit()
        
        return {
            "success": True,
            "created_count": created_count,
            "failed_count": len(failed_rows),
            "typology_used": filename,
            "failed_rows": failed_rows # Return all failed rows for the report
        }

    except Exception as e:
        db.rollback()
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Erreur lors de l'auto-import : {str(e)}")

class ExportRejectionsRequest(BaseModel):
    failed_rows: List[list]

@router.post("/export-rejections")
async def export_rejections(request: ExportRejectionsRequest):
    """
    Génère un fichier Excel à partir des lignes en échec.
    """
    try:
        from openpyxl import Workbook
        import io
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Tâches rejetées"
        
        headers = [
            "Nom de tâche", "Produit", "Famille", "Phase", 
            "Unité de mesure", "base de calcul", 
            "Responsable 1", "Responsable 2", 
            "Temps_min", "Temps_sec", "Raison du rejet"
        ]
        ws.append(headers)
        
        for row in request.failed_rows:
            ws.append(row)
            
        # Optional: adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except: pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width if adjusted_width > 15 else 15

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return Response(
            content=buffer.read(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=rejet_auto_import.xlsx"}
        )
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Erreur lors de l'export des rejets : {str(e)}")


#    return mapping


# --- Task Exclusions (Optimized Process) ---

# --- Task Exclusions (Optimized Process) ---

class ToggleExclusionIn(BaseModel):
    tache_id: Optional[int] = None
    centre_id: Optional[int] = None
    # For Typology
    categorie_id: Optional[int] = None
    nom_tache: Optional[str] = None
    famille_uo: Optional[str] = None
    produit: Optional[str] = None
    unite_mesure: Optional[str] = None

@router.get("/pm/exclusions")
def get_exclusions(centre_id: Optional[int] = None, categorie_id: Optional[int] = None, db: Session = Depends(get_db)):
    query = db.query(TacheExclueOptimisee)
    if centre_id:
        query = query.filter(TacheExclueOptimisee.centre_id == centre_id)
    if categorie_id:
        query = query.filter(TacheExclueOptimisee.categorie_id == categorie_id)
    
    exclusions = query.all()
    
    # On renvoie à la fois les IDs (si centre) et les quadruplets (si typologie)
    ids = [e.tache_id for e in exclusions if e.tache_id]
    quads = [
        {
            "nom": e.nom_tache,
            "famille": e.famille_uo,
            "produit": e.produit,
            "unite": e.unite_mesure
        }
        for e in exclusions if e.nom_tache
    ]
    return {"ids": ids, "quadruplets": quads}

@router.post("/pm/exclusions/toggle")
def toggle_exclusion(data: ToggleExclusionIn, db: Session = Depends(get_db)):
    if data.tache_id:
        # Mode par centre
        existing = db.query(TacheExclueOptimisee).filter(
            TacheExclueOptimisee.tache_id == data.tache_id,
            TacheExclueOptimisee.centre_id == data.centre_id
        ).first()
    else:
        # Mode par typologie (Quadruplet)
        existing = db.query(TacheExclueOptimisee).filter(
            TacheExclueOptimisee.categorie_id == data.categorie_id,
            TacheExclueOptimisee.nom_tache == data.nom_tache,
            TacheExclueOptimisee.famille_uo == data.famille_uo,
            TacheExclueOptimisee.produit == data.produit,
            TacheExclueOptimisee.unite_mesure == data.unite_mesure
        ).first()
    
    if existing:
        db.delete(existing)
        db.commit()
        return {"status": "removed"}
    else:
        new_excl = TacheExclueOptimisee(**data.dict())
        db.add(new_excl)
        db.commit()
        return {"status": "added"}

@router.get("/categories")
def get_categories(db: Session = Depends(get_db)):
    cats = db.query(Categorie).all()
    return [{"id": c.id, "label": c.label} for c in cats]

@router.get("/pm/standard-tasks/{categorie_id}")
def get_standard_tasks(categorie_id: int, db: Session = Depends(get_db)):
    """Lit les tâches standards depuis les fichiers Excel de typologie."""
    import os
    import openpyxl
    from pathlib import Path

    cat = db.query(Categorie).filter(Categorie.id == categorie_id).first()
    if not cat:
        raise HTTPException(status_code=404, detail="Catégorie non trouvée")

    typology_label = str(cat.label).upper()
    filename = "Standard.xlsx"
    
    # Mapping logic
    if "AGENCE MESSAGERIE" in typology_label or typology_label.startswith("AM"): filename = "AM.xlsx"
    elif "CENTRE MESSAGERIE" in typology_label or typology_label.startswith("CM"): filename = "CM.xlsx"
    elif "CENTRE DE DISTRIBUTION" in typology_label or typology_label.startswith("CD"): filename = "CD.xlsx"
    elif "CENTRE COURRIER COLIS" in typology_label or typology_label.startswith("CCC"): filename = "CCC.xlsx"
    elif "CELLULE DE DISTRIBUTION" in typology_label or typology_label.startswith("CLD"): filename = "CLD.xlsx"
    elif "CENTRE DE TRAITEMENT ET DISTRIBUTION" in typology_label or typology_label.startswith("CTD"): filename = "CTD.xlsx"

    # Absolute path resolution
    base_dir = Path(__file__).resolve().parent.parent # app/
    resources_dir = base_dir / "resources" / "typologies"
    file_path = resources_dir / filename

    if not file_path.exists():
        file_path = resources_dir / "Standard.xlsx"

    print(f"DEBUG PM: Loading typology tasks from {file_path} for category {typology_label}")

    tasks = []
    if not file_path.exists():
        print(f"DEBUG PM: File NOT FOUND at {file_path}")
        return tasks

    try:
        # data_only=True pour les valeurs calculées, read_only pour la performance
        wb = openpyxl.load_workbook(str(file_path), data_only=True)
        # Utiliser la première feuille au lieu de 'active' pour plus de robustesse
        ws = wb.worksheets[0]
        
        # Parcourir les lignes à partir de la 2ème (ignorer l'entête)
        # On utilise une boucle manuelle au cas où max_row est imprécis
        for row_idx in range(2, 1000): # Limite raisonnable pour éviter les boucles infinies sur fichiers corrompus
            if row_idx > ws.max_row and ws.max_row > 1: break # Respecter max_row s'il est cohérent
            
            # Lecture des 5 colonnes : Nom, Produit, Famille, Phase, Unité
            cells = [ws.cell(row=row_idx, column=col_idx).value for col_idx in range(1, 6)]
            
            # Si le nom de la tâche est vide, on arrête (fin des données)
            if not cells[0]: break
            
            tasks.append({
                "nom": str(cells[0] or "").strip(),
                "produit": str(cells[1] or "").strip(),
                "famille": str(cells[2] or "").strip(),
                "phase": str(cells[3] or "").strip(),
                "unite": str(cells[4] or "").strip()
            })
        wb.close()
    except Exception as e:
        print(f"Error reading {file_path}: {e}")
        import traceback
        traceback.print_exc()
        
    return tasks
@router.get("/exclusions/template")
def download_exclusion_template():
    import openpyxl
    from io import BytesIO
    from fastapi.responses import StreamingResponse

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Template Exclusions"
    
    # Headers
    headers = ["Nom Tâche", "Produit", "Famille / UO", "Unité Mesure"]
    ws.append(headers)
    
    # Sample Row
    ws.append(["EXEMPLE TACHE", "EXEMPLE PRODUIT", "EXEMPLE FAMILLE", "H"])

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=template_exclusions_optimise.xlsx"}
    )

@router.post("/exclusions/import/{categorie_id}")
async def import_exclusions(
    categorie_id: int,
    file: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    import openpyxl
    from io import BytesIO
    from fastapi.responses import StreamingResponse

    cat = db.query(Categorie).filter(Categorie.id == categorie_id).first()
    if not cat:
        raise HTTPException(status_code=404, detail="Typologie non trouvée")

    content = await file.read()
    wb = openpyxl.load_workbook(BytesIO(content), data_only=True)
    ws = wb.worksheets[0]

    # Get standard tasks for this typology to validate
    standard_tasks = get_standard_tasks(categorie_id, db)
    standard_set = set()
    for st in standard_tasks:
        key = (
            str(st['nom'] or '').strip().upper(),
            str(st['produit'] or '').strip().upper(),
            str(st['famille'] or '').strip().upper(),
            str(st['unite'] or '').strip().upper()
        )
        standard_set.add(key)

    created_count = 0
    failed_rows = []

    # Iterate rows (skip header)
    for row_idx in range(2, ws.max_row + 1):
        cells = [ws.cell(row=row_idx, column=c).value for c in range(1, 5)]
        if not cells[0]: continue

        nom, produit, famille, unite = [str(c or '').strip() for c in cells]
        key = (nom.upper(), produit.upper(), famille.upper(), unite.upper())

        if key not in standard_set:
            failed_rows.append({
                "line": row_idx,
                "data": cells,
                "error": "Tâche non trouvée dans le référentiel standard de cette typologie"
            })
            continue

        # Check if already excluded
        existing = db.query(TacheExclueOptimisee).filter(
            TacheExclueOptimisee.categorie_id == categorie_id,
            TacheExclueOptimisee.nom_tache == nom,
            TacheExclueOptimisee.produit == produit,
            TacheExclueOptimisee.famille_uo == famille,
            TacheExclueOptimisee.unite_mesure == unite
        ).first()

        if not existing:
            new_excl = TacheExclueOptimisee(
                categorie_id=categorie_id,
                nom_tache=nom,
                produit=produit,
                famille_uo=famille,
                unite_mesure=unite
            )
            db.add(new_excl)
            created_count += 1

    db.commit()

    if failed_rows:
        # Generate Rejection Excel
        wb_rej = openpyxl.Workbook()
        ws_rej = wb_rej.active
        ws_rej.title = "Rejets Exclusions"
        ws_rej.append(["Ligne", "Nom Tâche", "Produit", "Famille", "Unité", "Erreur"])
        
        for f in failed_rows:
            ws_rej.append([f['line']] + f['data'] + [f['error']])
        
        output_rej = BytesIO()
        wb_rej.save(output_rej)
        output_rej.seek(0)
        
        return StreamingResponse(
            output_rej,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": "attachment; filename=rejets_exclusions.xlsx",
                "X-Created-Count": str(created_count),
                "X-Error-Count": str(len(failed_rows))
            }
        )

    return {"status": "success", "created": created_count}


# --- Simulation Endpoints ---

class MappingPosteIn(BaseModel):
    poste_source_id: int
    poste_cible_id: int
    centre_id: Optional[int] = None

class MappingPosteOut(BaseModel):
    id: int
    poste_source_id: int
    poste_cible_id: int
    poste_source_code: Optional[str] = None
    poste_cible_code: Optional[str] = None
    centre_id: Optional[int] = None
    poste_source_label: Optional[str] = None
    poste_cible_label: Optional[str] = None

    class Config:
        from_attributes = True

@router.get("/mappings", response_model=List[MappingPosteOut])
def get_mappings(db: Session = Depends(get_db)):
    mappings = db.query(MappingPosteRecommande).all()
    out = []
    for m in mappings:
        item = MappingPosteOut.from_orm(m)
        if m.poste_source:
             item.poste_source_label = m.poste_source.label
             item.poste_source_code = m.poste_source.Code
        if m.poste_cible:
             item.poste_cible_label = m.poste_cible.label
             item.poste_cible_code = m.poste_cible.Code
        out.append(item)
    return out

@router.post("/mappings", response_model=MappingPosteOut)
def create_mapping(mapping: MappingPosteIn, db: Session = Depends(get_db)):
    # Vérifier si un mapping existe déjà pour ce poste source
    existing = db.query(MappingPosteRecommande).filter(
        MappingPosteRecommande.poste_source_id == mapping.poste_source_id,
        MappingPosteRecommande.centre_id == mapping.centre_id
    ).first()
    
    if existing:
        existing.poste_cible_id = mapping.poste_cible_id
    else:
        db_mapping = MappingPosteRecommande(**mapping.dict())
        db.add(db_mapping)
        existing = db_mapping
        
    db.commit()
    db.refresh(existing)
    
    # Enrich for response
    out = MappingPosteOut.from_orm(existing)
    if existing.poste_source:
        out.poste_source_label = existing.poste_source.label
        out.poste_source_code = existing.poste_source.Code
    if existing.poste_cible:
        out.poste_cible_label = existing.poste_cible.label
        out.poste_cible_code = existing.poste_cible.Code
    return out

@router.delete("/mappings/{mapping_id}")
def delete_mapping(mapping_id: int, db: Session = Depends(get_db)):
    db_mapping = db.query(MappingPosteRecommande).filter(MappingPosteRecommande.id == mapping_id).first()
    if not db_mapping:
        raise HTTPException(status_code=404, detail="Mapping non trouvé")
    db.delete(db_mapping)
    db.commit()
    return {"success": True}

@router.get("/mappings/template")
def get_mappings_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "Modele Mapping"
    ws.append(["Poste Source (Actuel)", "Poste Cible (Recommande)", "Centre ID (Optionnel)"])
    
    # Exemples
    ws.append(["AGENT DE FACTURATION", "AGENTS SUPPORT", ""])
    ws.append(["CONTROLEUR CABINE DES CHARGEMENTS", "AGENTS CONTRÔLE", ""])
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return Response(
        content=buffer.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=Modele_Mapping_Postes.xlsx"}
    )

@router.post("/mappings/import")
async def import_mappings(file: UploadFile = File(...), db: Session = Depends(get_db)):
    try:
        content = await file.read()
        wb = load_workbook(io.BytesIO(content), data_only=True)
        ws = wb.active
        
        created = 0
        updated = 0
        failed_rows = []
        
        # Headers pour le rapport de rejets
        headers = ["Poste Source (Actuel)", "Poste Cible (Recommande)", "Centre ID (Optionnel)", "Raison du rejet"]
        
        for row_idx in range(2, ws.max_row + 1):
            source_lbl = ws.cell(row=row_idx, column=1).value
            target_lbl = ws.cell(row=row_idx, column=2).value
            c_id = ws.cell(row=row_idx, column=3).value
            
            if source_lbl is None and target_lbl is None:
                continue
                
            if not source_lbl or not target_lbl:
                failed_rows.append([source_lbl, target_lbl, c_id, "Libellé source ou cible manquant"])
                continue
                
            # Resolution Source
            p_source = db.query(Poste).filter(sql_normalize_ws(Poste.label) == normalize_ws(str(source_lbl))).first()
            # Resolution Cible
            p_target = db.query(Poste).filter(sql_normalize_ws(Poste.label) == normalize_ws(str(target_lbl))).first()
            
            if not p_source or not p_target:
                missing = []
                if not p_source: missing.append(f"Source '{source_lbl}'")
                if not p_target: missing.append(f"Cible '{target_lbl}'")
                failed_rows.append([source_lbl, target_lbl, c_id, f"Non trouvé: {', '.join(missing)}"])
                continue
            
            centre_id = None
            if c_id:
                try: centre_id = int(c_id)
                except: pass

            # Update or Create
            existing = db.query(MappingPosteRecommande).filter(
                MappingPosteRecommande.poste_source_id == p_source.id,
                MappingPosteRecommande.centre_id == centre_id
            ).first()
            
            if existing:
                existing.poste_cible_id = p_target.id
                updated += 1
            else:
                new_m = MappingPosteRecommande(
                    poste_source_id=p_source.id,
                    poste_cible_id=p_target.id,
                    centre_id=centre_id
                )
                db.add(new_m)
                created += 1
        
        db.commit()
        
        if failed_rows:
            ewb = Workbook()
            ews = ewb.active
            ews.title = "Rejets Mapping"
            ews.append(headers)
            for fr in failed_rows:
                ews.append(fr)
            
            eout = io.BytesIO()
            ewb.save(eout)
            eout.seek(0)
            
            return Response(
                content=eout.read(),
                headers={
                    'Content-Disposition': 'attachment; filename="rejet_mapping_postes.xlsx"',
                    'X-Error-Count': str(len(failed_rows)),
                    'X-Created-Count': str(created),
                    'X-Updated-Count': str(updated)
                },
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        return {"created": created, "updated": updated, "errors": []}
    except Exception as e:
        db.rollback()
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/mappings/unmapped-report")
def get_unmapped_report(db: Session = Depends(get_db)):
    # Récupérer tous les postes qui existent dans CentrePoste (donc actifs dans un centre)
    # et qui n'ont pas de mapping global (centre_id IS NULL)
    
    sql = """
    SELECT DISTINCT p.id, p.label, p.Code
    FROM dbo.postes p
    INNER JOIN dbo.centre_postes cp ON p.Code = cp.code_resp
    LEFT JOIN dbo.mapping_postes_recommandes m ON p.id = m.poste_source_id AND m.centre_id IS NULL
    WHERE m.id IS NULL
    ORDER BY p.label
    """
    
    from sqlalchemy import text
    unmapped = db.execute(text(sql)).mappings().all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Postes Non Mappes"
    ws.append(["ID Poste", "Libellé Poste", "Code Poste"])
    
    for row in unmapped:
        ws.append([row['id'], row['label'], row['Code']])
        
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return Response(
        content=buffer.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=Postes_Non_Mappes.xlsx"}
    )
