import os
import openpyxl
from sqlalchemy.orm import Session
from app.models.db_models import Tache, CentrePoste, Poste, Centre, normalize_ws, sql_normalize_ws

def auto_import_tasks_if_empty(db: Session, centre_id: int):
    """
    Checks if a center has no tasks. If so, automatically imports them
    from an Excel template based on the center's typology.
    """
    # 1. Check if center already has tasks
    exists = db.query(Tache).join(CentrePoste).filter(CentrePoste.centre_id == centre_id).first()
    if exists:
        return 0, []

    # 2. Identify typology
    centre = db.query(Centre).filter(Centre.id == centre_id).first()
    if not centre or not centre.categorie:
        print(f"Centre {centre_id} has no typology/category. Skipping auto-import.")
        return 0, []

    typology_label = str(centre.categorie.label).upper()
    filename = "Standard.xlsx"
    if "AGENCE MESSAGERIE" in typology_label or typology_label.startswith("AM"):
        filename = "AM.xlsx"
    elif "CENTRE MESSAGERIE" in typology_label or typology_label.startswith("CM"):
        filename = "CM.xlsx"
    elif "CENTRE DE DISTRIBUTION" in typology_label or typology_label.startswith("CD"):
        filename = "CD.xlsx"
    elif "CENTRE COURRIER COLIS" in typology_label or typology_label.startswith("CCC"):
        filename = "CCC.xlsx"
    elif "CELLULE DE DISTRIBUTION" in typology_label or typology_label.startswith("CLD"):
        filename = "CLD.xlsx"
    elif "CENTRE DE TRAITEMENT ET DISTRIBUTION" in typology_label or typology_label.startswith("CTD"):
        filename = "CTD.xlsx"

    # resources are in app/resources/typologies
    current_dir = os.path.dirname(os.path.abspath(__file__))
    resources_dir = os.path.join(os.path.dirname(current_dir), "resources", "typologies")
    file_path = os.path.join(resources_dir, filename)

    if not os.path.exists(file_path):
        print(f"Template {filename} not found in {resources_dir}. Falling back to Standard.xlsx.")
        file_path = os.path.join(resources_dir, "Standard.xlsx")
        if not os.path.exists(file_path):
            print(f"Standard template not found. Aborting auto-import for center {centre_id}.")
            return 0, [f"Template not found: {filename}"]

    # 3. Import
    print(f"Auto-importing tasks for center {centre_id} using {filename}")
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active
        created_count, failed_rows = _process_task_import_workbook(db, centre_id, ws)
        db.commit()
        return created_count, failed_rows
    except Exception as e:
        db.rollback()
        print(f"Error during auto-import for center {centre_id}: {str(e)}")
        return 0, [str(e)]

def _process_task_import_workbook(db: Session, centre_id: int, ws):
    """
    Internal logic to process the Excel rows and create DB records.
    Adapted from bandoeng_router.py.
    """
    created_count = 0
    failed_rows = []
    
    def resolve_cp(session, c_id, resp_name):
        if not resp_name: return None, None
        l_resp = normalize_ws(resp_name)
        
        # Match poste label robustly
        poste = session.query(Poste).filter(sql_normalize_ws(Poste.label) == l_resp).first()
        if not poste:
            return None, f"Poste '{resp_name}' non trouvé"
        
        if not poste.Code:
            return None, f"Le poste '{resp_name}' n'a pas de code associé"
        
        # Check if center already has this poste
        cp = session.query(CentrePoste).filter(
            CentrePoste.centre_id == c_id,
            CentrePoste.code_resp == poste.Code
        ).first()
        
        if cp:
            return cp.id, None
        else:
            # Create the link
            new_cp = CentrePoste(
                centre_id=c_id,
                poste_id=poste.id,
                code_resp=poste.Code,
                effectif_actuel=0
            )
            session.add(new_cp)
            session.flush()
            return new_cp.id, None

    def parse_numeric(val, default="0.0"):
        if val is None: return default
        try:
            clean_val = float(str(val).replace(',', '.'))
            return str(clean_val)
        except:
            return default

    # Skip header (starting at row 2)
    for row_idx in range(2, ws.max_row + 1):
        row_data = [ws.cell(row=row_idx, column=i).value for i in range(1, 11)]
        if not any(row_data): continue # Skip empty rows
        
        nom_tache, produit, famille, phase, unit, base, r1_name, r2_name, t_min, t_sec = row_data
        
        # Resolution Resp 1
        cp_id_1, err1 = resolve_cp(db, centre_id, r1_name)
        
        # Resolution Resp 2
        cp_id_2, err2 = (None, None)
        if r2_name:
            cp_id_2, err2 = resolve_cp(db, centre_id, r2_name)
        
        # Error handling
        if (r1_name and not cp_id_1) or (r2_name and not cp_id_2):
            reason = err1 if (r1_name and not cp_id_1) else err2
            failed_rows.append(list(row_data) + [reason])
            continue
        
        if not cp_id_1 and not cp_id_2:
            failed_rows.append(list(row_data) + ["Aucun responsable valide fourni"])
            continue

        # Insertion
        tasks_to_create = []
        if cp_id_1: tasks_to_create.append(cp_id_1)
        if cp_id_2: tasks_to_create.append(cp_id_2)
        
        parsed_min = parse_numeric(t_min)
        parsed_sec = parse_numeric(t_sec)
        
        # Base de calcul with percentage detection
        cell_base = ws.cell(row=row_idx, column=6)
        if base is not None:
            if isinstance(base, (int, float)) and cell_base.number_format and '%' in str(cell_base.number_format):
                parsed_base = str(int(round(base * 100)))
            else:
                parsed_base = parse_numeric(base, "100")
        else:
            parsed_base = "100"

        for cp_id in tasks_to_create:
            new_t = Tache(
                nom_tache=str(nom_tache or ""),
                produit=str(produit or ""),
                famille_uo=str(famille or ""),
                phase=str(phase or ""),
                unite_mesure=str(unit or ""),
                base_calcul=parsed_base,
                moyenne_min=parsed_min,
                moy_sec=parsed_sec,
                centre_poste_id=cp_id,
                etat="ACTIF"
            )
            db.add(new_t)
            created_count += 1
    
    return created_count, failed_rows
