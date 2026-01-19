"""
Script pour générer le template Excel d'import des volumes
Compatible avec la structure de l'application TAWAZOON RH
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

def create_import_template():
    """
    Crée un fichier Excel template pour l'import des volumes
    avec 2 feuilles : Niveau Intervenant et Niveau Centre
    """
    
    # Créer le workbook
    wb = openpyxl.Workbook()
    
    # Supprimer la feuille par défaut
    wb.remove(wb.active)
    
    # ========================================
    # FEUILLE 1 : NIVEAU INTERVENANT (Centre)
    # ========================================
    ws_intervenant = wb.create_sheet("Import Niveau Intervenant")
    
    # Styles
    header_fill = PatternFill(start_color="005EA8", end_color="005EA8", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=10)
    section_fill = PatternFill(start_color="E8F4F8", end_color="E8F4F8", fill_type="solid")
    section_font = Font(bold=True, size=9, color="005EA8")
    border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )
    
    # Titre principal
    ws_intervenant.merge_cells('A1:P1')
    cell = ws_intervenant['A1']
    cell.value = "IMPORT VOLUMES - NIVEAU INTERVENANT (CENTRE)"
    cell.font = Font(bold=True, size=14, color="005EA8")
    cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Instructions
    ws_intervenant.merge_cells('A2:P2')
    cell = ws_intervenant['A2']
    cell.value = "Remplissez les volumes globaux par centre. Laissez vide les colonnes non utilisées."
    cell.font = Font(italic=True, size=9, color="666666")
    cell.alignment = Alignment(horizontal='center')
    
    # Ligne vide
    current_row = 4
    
    # En-têtes de colonnes
    headers = [
        "Centre ID",
        "Nom du Centre",
        # FLUX ARRIVÉE
        "Amana Arrivée",
        "CO Arrivée", 
        "CR Arrivée",
        "E-Barkia Arrivée",
        "LRH Arrivée",
        # GUICHET
        "Guichet Dépôt",
        "Guichet Récup",
        # FLUX DÉPART
        "Amana Départ",
        "CO Départ",
        "CR Départ",
        "E-Barkia Départ",
        "LRH Départ",
        # AUTRES
        "Sacs",
        "Colis"
    ]
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws_intervenant.cell(row=current_row, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
    
    # Ajuster la hauteur de la ligne d'en-tête
    ws_intervenant.row_dimensions[current_row].height = 30
    
    # Lignes d'exemple
    current_row += 1
    example_data = [
        [1, "Centre Casablanca", 100, 200, 50, 30, 20, 150, 80, 90, 180, 45, 25, 15, 50, 30],
        [2, "Centre Rabat", 120, 220, 60, 35, 25, 160, 90, 100, 200, 55, 30, 18, 60, 35],
        [3, "Centre Marrakech", "", "", "", "", "", "", "", "", "", "", "", "", "", ""],
    ]
    
    for row_data in example_data:
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws_intervenant.cell(row=current_row, column=col_idx)
            cell.value = value
            cell.border = border
            cell.alignment = Alignment(horizontal='center' if col_idx > 2 else 'left', vertical='center')
            if col_idx > 2 and value != "":
                cell.number_format = '#,##0'
        current_row += 1
    
    # Ajuster les largeurs de colonnes
    ws_intervenant.column_dimensions['A'].width = 12
    ws_intervenant.column_dimensions['B'].width = 25
    for col in range(3, 17):
        ws_intervenant.column_dimensions[get_column_letter(col)].width = 12
    
    # ========================================
    # FEUILLE 2 : NIVEAU CENTRE (Détaillé par Poste)
    # ========================================
    ws_centre = wb.create_sheet("Import Niveau Centre")
    
    # Titre principal
    ws_centre.merge_cells('A1:V1')
    cell = ws_centre['A1']
    cell.value = "IMPORT VOLUMES - NIVEAU CENTRE (DÉTAIL PAR POSTE)"
    cell.font = Font(bold=True, size=14, color="005EA8")
    cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Instructions
    ws_centre.merge_cells('A2:V2')
    cell = ws_centre['A2']
    cell.value = "Remplissez les volumes détaillés par flux, sens et segment pour chaque poste du centre."
    cell.font = Font(italic=True, size=9, color="666666")
    cell.alignment = Alignment(horizontal='center')
    
    current_row = 4
    
    # En-têtes avec structure Flux/Sens/Segment
    headers_centre = [
        "Centre ID",
        "Centre Poste ID",
        "Nom du Centre",
        "Nom du Poste",
        # FLUX ARRIVÉE - Segments
        "Amana Arr GLOBAL",
        "Amana Arr PART",
        "Amana Arr PRO",
        "Amana Arr DIST",
        "Amana Arr AXES",
        "CO Arr GLOBAL",
        "CO Arr PART",
        "CO Arr PRO",
        "CO Arr DIST",
        "CO Arr AXES",
        # GUICHET
        "Guichet DÉPÔT",
        "Guichet RÉCUP",
        # FLUX DÉPART - Segments
        "Amana Dép GLOBAL",
        "Amana Dép PART",
        "Amana Dép PRO",
        "Amana Dép DIST",
        "Amana Dép AXES",
        "CO Dép GLOBAL",
        "CO Dép PART",
        "CO Dép PRO",
        "CO Dép DIST",
        "CO Dép AXES",
    ]
    
    for col_idx, header in enumerate(headers_centre, start=1):
        cell = ws_centre.cell(row=current_row, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
    
    ws_centre.row_dimensions[current_row].height = 40
    
    # Lignes d'exemple
    current_row += 1
    example_data_centre = [
        [1, 101, "Centre Casablanca", "Guichetier", 20, 15, 10, 5, 0, 40, 30, 20, 10, 0, 50, 30, 18, 12, 8, 4, 0, 35, 25, 15, 8, 0],
        [1, 102, "Centre Casablanca", "Trieur", 30, 20, 15, 8, 0, 60, 40, 25, 12, 0, 0, 0, 25, 18, 12, 6, 0, 50, 35, 20, 10, 0],
        [2, 201, "Centre Rabat", "Guichetier", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", ""],
    ]
    
    for row_data in example_data_centre:
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws_centre.cell(row=current_row, column=col_idx)
            cell.value = value
            cell.border = border
            cell.alignment = Alignment(horizontal='center' if col_idx > 4 else 'left', vertical='center')
            if col_idx > 4 and value != "":
                cell.number_format = '#,##0'
        current_row += 1
    
    # Ajuster les largeurs de colonnes
    ws_centre.column_dimensions['A'].width = 12
    ws_centre.column_dimensions['B'].width = 15
    ws_centre.column_dimensions['C'].width = 25
    ws_centre.column_dimensions['D'].width = 20
    for col in range(5, 27):
        ws_centre.column_dimensions[get_column_letter(col)].width = 11
    
    # ========================================
    # FEUILLE 3 : GUIDE & MAPPING
    # ========================================
    ws_guide = wb.create_sheet("Guide & Mapping")
    
    # Titre
    ws_guide.merge_cells('A1:D1')
    cell = ws_guide['A1']
    cell.value = "GUIDE D'UTILISATION & MAPPING DES DONNÉES"
    cell.font = Font(bold=True, size=14, color="005EA8")
    cell.alignment = Alignment(horizontal='center', vertical='center')
    
    current_row = 3
    
    # Section 1 : Flux
    ws_guide.merge_cells(f'A{current_row}:D{current_row}')
    cell = ws_guide[f'A{current_row}']
    cell.value = "1. FLUX DISPONIBLES"
    cell.font = section_font
    cell.fill = section_fill
    current_row += 1
    
    flux_data = [
        ["ID", "Code", "Nom", "Description"],
        [1, "AMANA", "Amana", "Colis Amana"],
        [2, "CO", "Courrier Ordinaire", "Courrier standard"],
        [3, "CR", "Courrier Recommandé", "Courrier avec accusé"],
        [4, "EBARKIA", "E-Barkia", "Service E-Barkia"],
        [5, "LRH", "LRH", "Lettres recommandées avec AR"],
    ]
    
    for row_data in flux_data:
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws_guide.cell(row=current_row, column=col_idx)
            cell.value = value
            cell.border = border
            if current_row == 4:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
        current_row += 1
    
    current_row += 1
    
    # Section 2 : Sens
    ws_guide.merge_cells(f'A{current_row}:D{current_row}')
    cell = ws_guide[f'A{current_row}']
    cell.value = "2. SENS DE FLUX"
    cell.font = section_font
    cell.fill = section_fill
    current_row += 1
    
    sens_data = [
        ["ID", "Code", "Nom", "Description"],
        [1, "ARRIVEE", "Arrivée", "Flux entrant"],
        [2, "GUICHET", "Guichet", "Opérations au guichet"],
        [3, "DEPART", "Départ", "Flux sortant"],
    ]
    
    for row_data in sens_data:
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws_guide.cell(row=current_row, column=col_idx)
            cell.value = value
            cell.border = border
            if current_row == current_row - len(sens_data) + 1:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
        current_row += 1
    
    current_row += 1
    
    # Section 3 : Segments
    ws_guide.merge_cells(f'A{current_row}:D{current_row}')
    cell = ws_guide[f'A{current_row}']
    cell.value = "3. SEGMENTS"
    cell.font = section_font
    cell.fill = section_fill
    current_row += 1
    
    segment_data = [
        ["ID", "Code", "Nom", "Description"],
        [1, "GLOBAL", "Global", "Volume global non segmenté"],
        [2, "PART", "Particuliers", "Segment particuliers"],
        [3, "PRO", "Professionnels", "Segment professionnels"],
        [4, "DIST", "Distribution", "Segment distribution"],
        [5, "AXES", "Axes", "Segment axes stratégiques"],
        [6, "DEPOT", "Dépôt", "Opération de dépôt (guichet)"],
        [7, "RECUP", "Récupération", "Opération de récupération (guichet)"],
    ]
    
    for row_data in segment_data:
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws_guide.cell(row=current_row, column=col_idx)
            cell.value = value
            cell.border = border
            if current_row == current_row - len(segment_data) + 1:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
        current_row += 1
    
    # Ajuster les largeurs
    for col in ['A', 'B', 'C', 'D']:
        ws_guide.column_dimensions[col].width = 20
    
    # ========================================
    # SAUVEGARDER LE FICHIER
    # ========================================
    filename = f"Template_Import_Volumes_TAWAZOON_RH_{datetime.now().strftime('%Y%m%d')}.xlsx"
    wb.save(filename)
    print(f"✅ Template créé avec succès : {filename}")
    return filename

if __name__ == "__main__":
    create_import_template()
