"""
Script pour générer le template Excel d'import des volumes
STRUCTURE MATRICIELLE - Reproduction exacte de l'UI
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

def create_matrix_import_template():
    """
    Crée un fichier Excel template avec structure matricielle
    identique à l'interface utilisateur
    """
    
    # Créer le workbook
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    
    # Styles
    header_fill = PatternFill(start_color="005EA8", end_color="005EA8", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=10)
    section_fill = PatternFill(start_color="E8F4F8", end_color="E8F4F8", fill_type="solid")
    section_font = Font(bold=True, size=10, color="005EA8")
    flux_fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
    flux_font = Font(bold=True, size=9)
    border_thick = Border(
        left=Side(style='medium', color='005EA8'),
        right=Side(style='medium', color='005EA8'),
        top=Side(style='medium', color='005EA8'),
        bottom=Side(style='medium', color='005EA8')
    )
    border_thin = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )
    
    # ========================================
    # FEUILLE : IMPORT VOLUMES (STRUCTURE MATRICIELLE)
    # ========================================
    ws = wb.create_sheet("Import Volumes")
    
    # Titre principal
    ws.merge_cells('A1:L1')
    cell = ws['A1']
    cell.value = "IMPORT VOLUMES - STRUCTURE MATRICIELLE"
    cell.font = Font(bold=True, size=16, color="005EA8")
    cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 25
    
    # Instructions
    ws.merge_cells('A2:L2')
    cell = ws['A2']
    cell.value = "Remplissez les volumes selon la structure Flux × Segment (comme dans l'interface)"
    cell.font = Font(italic=True, size=9, color="666666")
    cell.alignment = Alignment(horizontal='center')
    
    # Informations du centre
    row = 4
    ws['A' + str(row)] = "Centre ID:"
    ws['B' + str(row)] = 1
    ws['A' + str(row)].font = Font(bold=True)
    ws['B' + str(row)].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    
    row += 1
    ws['A' + str(row)] = "Nom du Centre:"
    ws['B' + str(row)] = "Centre Casablanca"
    ws['A' + str(row)].font = Font(bold=True)
    ws['B' + str(row)].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    
    row += 1
    ws['A' + str(row)] = "Poste ID (optionnel):"
    ws['B' + str(row)] = ""
    ws['A' + str(row)].font = Font(bold=True, italic=True, size=8)
    ws['B' + str(row)].fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
    
    row += 2  # Espace
    
    # ========================================
    # SECTION A : FLUX ARRIVÉE
    # ========================================
    start_row = row
    
    # Titre de section
    ws.merge_cells(f'A{row}:F{row}')
    cell = ws[f'A{row}']
    cell.value = "A) FLUX ARRIVÉE"
    cell.font = section_font
    cell.fill = section_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[row].height = 20
    row += 1
    
    # En-têtes des colonnes (Segments)
    ws[f'A{row}'] = "FLUX \\ SEGMENT"
    ws[f'A{row}'].font = header_font
    ws[f'A{row}'].fill = header_fill
    ws[f'A{row}'].alignment = Alignment(horizontal='center', vertical='center')
    
    segments = ["GLOBAL", "PART.", "PRO", "DIST.", "AXES"]
    for col_idx, segment in enumerate(segments, start=2):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = segment
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border_thick
    
    ws.row_dimensions[row].height = 20
    row += 1
    
    # Lignes de flux
    flux_list = ["Amana", "CO", "CR", "E-Barkia", "LRH"]
    for flux in flux_list:
        ws[f'A{row}'] = flux
        ws[f'A{row}'].font = flux_font
        ws[f'A{row}'].fill = flux_fill
        ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center')
        ws[f'A{row}'].border = border_thin
        
        # Cellules de saisie pour chaque segment
        for col_idx in range(2, 7):  # B à F
            cell = ws.cell(row=row, column=col_idx)
            cell.value = ""
            cell.border = border_thin
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.number_format = '#,##0'
        
        row += 1
    
    row += 1  # Espace entre sections
    
    # ========================================
    # SECTION B : GUICHET
    # ========================================
    
    # Titre de section
    ws.merge_cells(f'A{row}:C{row}')
    cell = ws[f'A{row}']
    cell.value = "B) GUICHET"
    cell.font = section_font
    cell.fill = section_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[row].height = 20
    row += 1
    
    # En-têtes
    ws[f'A{row}'] = "OPÉRATION"
    ws[f'A{row}'].font = header_font
    ws[f'A{row}'].fill = header_fill
    ws[f'A{row}'].alignment = Alignment(horizontal='center', vertical='center')
    
    guichet_ops = ["DÉPÔT", "RÉCUP."]
    for col_idx, op in enumerate(guichet_ops, start=2):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = op
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border_thick
    
    ws.row_dimensions[row].height = 20
    row += 1
    
    # Ligne de saisie
    ws[f'A{row}'] = "Volume"
    ws[f'A{row}'].font = flux_font
    ws[f'A{row}'].fill = flux_fill
    ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center')
    
    for col_idx in range(2, 4):  # B à C
        cell = ws.cell(row=row, column=col_idx)
        cell.value = ""
        cell.border = border_thin
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.number_format = '#,##0'
    
    row += 2  # Espace entre sections
    
    # ========================================
    # SECTION C : FLUX DÉPART
    # ========================================
    
    # Titre de section
    ws.merge_cells(f'A{row}:F{row}')
    cell = ws[f'A{row}']
    cell.value = "C) FLUX DÉPART"
    cell.font = section_font
    cell.fill = section_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[row].height = 20
    row += 1
    
    # En-têtes des colonnes (Segments)
    ws[f'A{row}'] = "FLUX \\ SEGMENT"
    ws[f'A{row}'].font = header_font
    ws[f'A{row}'].fill = header_fill
    ws[f'A{row}'].alignment = Alignment(horizontal='center', vertical='center')
    
    for col_idx, segment in enumerate(segments, start=2):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = segment
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border_thick
    
    ws.row_dimensions[row].height = 20
    row += 1
    
    # Lignes de flux (même structure que Arrivée)
    for flux in flux_list:
        ws[f'A{row}'] = flux
        ws[f'A{row}'].font = flux_font
        ws[f'A{row}'].fill = flux_fill
        ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center')
        ws[f'A{row}'].border = border_thin
        
        # Cellules de saisie pour chaque segment
        for col_idx in range(2, 7):  # B à F
            cell = ws.cell(row=row, column=col_idx)
            cell.value = ""
            cell.border = border_thin
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.number_format = '#,##0'
        
        row += 1
    
    # Ajuster les largeurs de colonnes
    ws.column_dimensions['A'].width = 18
    for col in ['B', 'C', 'D', 'E', 'F']:
        ws.column_dimensions[col].width = 12
    
    # ========================================
    # FEUILLE : GUIDE DE REMPLISSAGE
    # ========================================
    ws_guide = wb.create_sheet("Guide")
    
    # Titre
    ws_guide.merge_cells('A1:D1')
    cell = ws_guide['A1']
    cell.value = "GUIDE DE REMPLISSAGE"
    cell.font = Font(bold=True, size=14, color="005EA8")
    cell.alignment = Alignment(horizontal='center', vertical='center')
    
    row = 3
    
    # Instructions
    instructions = [
        ("1. IDENTIFICATION", ""),
        ("", "Remplissez obligatoirement :"),
        ("", "  - Centre ID : Identifiant numérique du centre"),
        ("", "  - Nom du Centre : Nom complet du centre"),
        ("", "  - Poste ID (optionnel) : Si import au niveau poste"),
        ("", ""),
        ("2. STRUCTURE DES DONNÉES", ""),
        ("", "Le template reproduit exactement l'interface :"),
        ("", ""),
        ("A) FLUX ARRIVÉE", "Matrice 5 flux × 5 segments"),
        ("", "  Flux : Amana, CO, CR, E-Barkia, LRH"),
        ("", "  Segments : GLOBAL, PART., PRO, DIST., AXES"),
        ("", ""),
        ("B) GUICHET", "2 opérations uniquement"),
        ("", "  - DÉPÔT : Volume des dépôts"),
        ("", "  - RÉCUP. : Volume des récupérations"),
        ("", ""),
        ("C) FLUX DÉPART", "Même matrice que Flux Arrivée"),
        ("", "  Flux : Amana, CO, CR, E-Barkia, LRH"),
        ("", "  Segments : GLOBAL, PART., PRO, DIST., AXES"),
        ("", ""),
        ("3. RÈGLES DE SAISIE", ""),
        ("", "✓ Saisir uniquement des nombres entiers ou décimaux"),
        ("", "✓ Laisser vide si volume = 0"),
        ("", "✓ Ne pas modifier la structure du tableau"),
        ("", "✓ Ne pas ajouter/supprimer de lignes ou colonnes"),
        ("", ""),
        ("4. MAPPING DES SEGMENTS", ""),
        ("GLOBAL", "Volume global non segmenté"),
        ("PART.", "Segment Particuliers"),
        ("PRO", "Segment Professionnels"),
        ("DIST.", "Segment Distribution"),
        ("AXES", "Segment Axes stratégiques"),
        ("", ""),
        ("5. IMPORT DANS L'APPLICATION", ""),
        ("", "1. Remplir le template"),
        ("", "2. Aller dans Vue Nationale ou Vue Direction"),
        ("", "3. Cliquer sur 'Importer'"),
        ("", "4. Sélectionner ce fichier"),
        ("", "5. Valider l'import"),
    ]
    
    for instruction in instructions:
        ws_guide[f'A{row}'] = instruction[0]
        ws_guide[f'B{row}'] = instruction[1]
        
        if instruction[0] and not instruction[0].startswith(" "):
            ws_guide[f'A{row}'].font = Font(bold=True, color="005EA8")
        
        row += 1
    
    # Ajuster les largeurs
    ws_guide.column_dimensions['A'].width = 25
    ws_guide.column_dimensions['B'].width = 50
    
    # ========================================
    # SAUVEGARDER LE FICHIER
    # ========================================
    filename = f"Template_Import_Volumes_Matriciel_{datetime.now().strftime('%Y%m%d')}.xlsx"
    wb.save(filename)
    print(f"✅ Template matriciel créé avec succès : {filename}")
    return filename

if __name__ == "__main__":
    create_matrix_import_template()
